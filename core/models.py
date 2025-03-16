import csv
import os
import uuid
from datetime import date
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import models
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render

from inventory.models import Warehouse
from organization.models import Department


class Tax(models.Model):
    """Modell für Mehrwertsteuersätze."""
    name = models.CharField(max_length=100, verbose_name="Name")
    code = models.CharField(max_length=20, unique=True, verbose_name="Steuerkürzel")
    rate = models.DecimalField(max_digits=5, decimal_places=2, verbose_name="Steuersatz (%)")
    description = models.TextField(blank=True, verbose_name="Beschreibung")
    is_default = models.BooleanField(default=False, verbose_name="Standard-Steuersatz")
    is_active = models.BooleanField(default=True, verbose_name="Aktiv")
    created_at = models.DateTimeField(auto_now_add=True, verbose_name="Erstellt am")
    updated_at = models.DateTimeField(auto_now=True, verbose_name="Aktualisiert am")

    class Meta:
        verbose_name = "Mehrwertsteuersatz"
        verbose_name_plural = "Mehrwertsteuersätze"
        ordering = ['rate']

    def __str__(self):
        return f"{self.name} ({self.rate}%)"

    def save(self, *args, **kwargs):
        # Wenn dieser Steuersatz als Standard markiert wird, alle anderen auf nicht-Standard setzen
        if self.is_default:
            Tax.objects.filter(is_default=True).exclude(pk=self.pk).update(is_default=False)
        super().save(*args, **kwargs)

    @classmethod
    def get_default_tax(cls):
        """Gibt den Standard-Steuersatz zurück. Falls keiner existiert, wird der erste aktive Steuersatz zurückgegeben."""
        default_tax = cls.objects.filter(is_default=True, is_active=True).first()
        if not default_tax:
            default_tax = cls.objects.filter(is_active=True).first()
        return default_tax

class Category(models.Model):
    name = models.CharField(max_length=100)
    description = models.TextField(blank=True)

    def __str__(self):
        return self.name

    class Meta:
        verbose_name = "Kategorie"
        verbose_name_plural = "Kategorien"
        ordering = ['name']

class Product(models.Model):
    name = models.CharField(max_length=200)
    sku = models.CharField(max_length=50, unique=True)
    barcode = models.CharField(max_length=100, blank=True)
    description = models.TextField(blank=True)
    category = models.ForeignKey(Category, on_delete=models.SET_NULL, null=True)
    minimum_stock = models.IntegerField(default=0)
    unit = models.CharField(max_length=20, default="Stück")
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)
    warehouses = models.ManyToManyField(Warehouse, through='ProductWarehouse')
    has_variants = models.BooleanField(default=False)
    has_serial_numbers = models.BooleanField(default=False)
    has_batch_tracking = models.BooleanField(default=False)
    has_expiry_tracking = models.BooleanField(default=False)

    tax = models.ForeignKey(Tax, on_delete=models.SET_NULL, null=True, blank=True,
                            verbose_name="Mehrwertsteuersatz",
                            help_text="Der auf dieses Produkt anzuwendende Mehrwertsteuersatz")

    def __str__(self):
        return self.name
    
    @property
    def total_stock(self):
        """Berechnet den Gesamtbestand über alle Lager."""
        from django.db.models import Sum
        return self.productwarehouse_set.aggregate(Sum('quantity'))['quantity__sum'] or 0

    @property
    def get_tax_rate(self):
        """Gibt den anzuwendenden Steuersatz zurück oder den Standard falls keiner hinterlegt ist."""
        if self.tax and self.tax.is_active:
            return self.tax.rate
        # Standard-Steuersatz verwenden
        default_tax = Tax.get_default_tax()
        if default_tax:
            return default_tax.rate
        # Fallback, falls kein Steuersatz definiert ist
        return 0


class ProductWarehouse(models.Model):
    product = models.ForeignKey(Product, on_delete=models.CASCADE)
    warehouse = models.ForeignKey(Warehouse, on_delete=models.CASCADE)
    quantity = models.DecimalField(max_digits=10, decimal_places=2, default=0)

    class Meta:
        unique_together = ('product', 'warehouse')


# Updated ImportLog model in core/models.py

# Updated ImportLog model in core/models.py

class ImportLog(models.Model):
    STATUS_CHOICES = (
        ('completed', 'Completed'),
        ('completed_with_errors', 'Completed with Errors'),
        ('processing', 'Processing'),
        ('failed', 'Failed'),
    )

    IMPORT_TYPE_CHOICES = (
        ('products', 'Products'),
        ('categories', 'Categories'),
        ('suppliers', 'Suppliers'),
        ('supplier_products', 'Supplier Products'),
        ('warehouses', 'Warehouses'),
        ('departments', 'Departments'),
        ('warehouse_products', 'Warehouse Products'),
    )

    file_name = models.CharField(max_length=255)
    import_type = models.CharField(max_length=50, choices=IMPORT_TYPE_CHOICES, default='products')
    status = models.CharField(max_length=25, choices=STATUS_CHOICES, default='processing')

    # Keep original fields for compatibility
    total_records = models.IntegerField(default=0)
    success_count = models.IntegerField(default=0)
    error_count = models.IntegerField(default=0)

    # New more descriptive fields
    rows_processed = models.IntegerField(default=0)
    rows_created = models.IntegerField(default=0)
    rows_updated = models.IntegerField(default=0)
    rows_error = models.IntegerField(default=0)

    error_file = models.FileField(upload_to='import_errors/', null=True, blank=True)
    created_at = models.DateTimeField(auto_now_add=True)
    created_by = models.ForeignKey(User, on_delete=models.CASCADE, related_name='import_logs')

    # Additional fields for error handling
    notes = models.TextField(blank=True, null=True)
    error_details = models.TextField(blank=True, null=True)

    # Properties for backwards compatibility
    @property
    def successful_rows(self):
        return self.success_count

    @successful_rows.setter
    def successful_rows(self, value):
        self.success_count = value
        self.rows_created = value  # Update the new field too

    @property
    def failed_rows(self):
        return self.error_count

    @failed_rows.setter
    def failed_rows(self, value):
        self.error_count = value
        self.rows_error = value  # Update the new field too

    @property
    def total_rows(self):
        return self.total_records

    @total_rows.setter
    def total_rows(self, value):
        self.total_records = value
        self.rows_processed = value  # Update the new field too

    @property
    def user(self):
        return self.created_by

    @user.setter
    def user(self, value):
        self.created_by = value

    @property
    def started_by(self):
        return self.created_by

    @started_by.setter
    def started_by(self, value):
        self.created_by = value

    @property
    def has_errors(self):
        return self.error_count > 0

    @property
    def filename(self):
        return self.file_name

    def success_rate(self):
        if self.total_records > 0:
            return round((self.success_count / self.total_records) * 100, 1)
        return 0

    class Meta:
        ordering = ['-created_at']


@login_required
def import_log_list(request):
    # Get all users for the filter dropdown
    users = User.objects.all()

    # Base queryset
    queryset = ImportLog.objects.all()

    # Apply filters
    if 'search' in request.GET and request.GET['search']:
        search = request.GET['search']
        queryset = queryset.filter(file_name__icontains=search)

    if 'status' in request.GET and request.GET['status']:
        status = request.GET['status']
        queryset = queryset.filter(status=status)

    if 'user' in request.GET and request.GET['user']:
        user_id = request.GET['user']
        queryset = queryset.filter(created_by_id=user_id)

    if 'daterange' in request.GET and request.GET['daterange']:
        date_range = request.GET['daterange'].split(' - ')
        if len(date_range) == 2:
            start_date = datetime.datetime.strptime(date_range[0], '%Y-%m-%d')
            end_date = datetime.datetime.strptime(date_range[1], '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            queryset = queryset.filter(created_at__range=[start_date, end_date])

    # Apply sorting
    sort_param = request.GET.get('sort', '-created_at')
    if sort_param:
        queryset = queryset.order_by(sort_param)

    # Pagination
    paginator = Paginator(queryset, 25)  # Show 25 logs per page
    page_number = request.GET.get('page')
    import_logs = paginator.get_page(page_number)

    context = {
        'import_logs': import_logs,
        'users': users,
    }

    return render(request, 'import_log_list.html', context)


@login_required
def import_log_detail(request, log_id):
    log = get_object_or_404(ImportLog, id=log_id)
    return render(request, 'import_log_detail.html', {'log': log})


@login_required
def download_error_file(request, log_id):
    log = get_object_or_404(ImportLog, id=log_id)
    if log.error_file:
        response = HttpResponse(log.error_file.read(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{log.file_name}_errors.csv"'
        return response
    return HttpResponse("Error file not found", status=404)


@login_required
def delete_import_log(request, log_id):
    if request.method == 'POST':
        log = get_object_or_404(ImportLog, id=log_id)
        log.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=405)


@login_required
def bulk_delete_import_logs(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')
        ImportLog.objects.filter(id__in=ids).delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=405)


@login_required
def export_import_logs(request):
    format_type = request.GET.get('format', 'csv')

    # Filter logs based on request parameters
    queryset = ImportLog.objects.all()

    # Check if specific IDs were requested
    if 'ids' in request.GET and request.GET['ids']:
        ids = request.GET['ids'].split(',')
        queryset = queryset.filter(id__in=ids)
    else:
        # Apply the same filters as the list view
        if 'search' in request.GET and request.GET['search']:
            queryset = queryset.filter(file_name__icontains=request.GET['search'])

        if 'status' in request.GET and request.GET['status']:
            queryset = queryset.filter(status=request.GET['status'])

        if 'user' in request.GET and request.GET['user']:
            queryset = queryset.filter(created_by_id=request.GET['user'])

        if 'daterange' in request.GET and request.GET['daterange']:
            date_range = request.GET['daterange'].split(' - ')
            if len(date_range) == 2:
                start_date = datetime.datetime.strptime(date_range[0], '%Y-%m-%d')
                end_date = datetime.datetime.strptime(date_range[1], '%Y-%m-%d')
                end_date = end_date.replace(hour=23, minute=59, second=59)
                queryset = queryset.filter(created_at__range=[start_date, end_date])

    # CSV export
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="import_logs.csv"'

        writer = csv.writer(response)
        writer.writerow(['ID', 'Date', 'Filename', 'Status', 'Items Processed',
                         'Success Count', 'Error Count', 'User'])

        for log in queryset:
            writer.writerow([
                log.id,
                log.created_at.strftime('%Y-%m-%d %H:%M'),
                log.file_name,
                log.status,
                log.total_records,
                log.success_count,
                log.error_count,
                log.created_by.username
            ])

        return response

    # Excel export
    elif format_type == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Logs"

        # Add header row with bold font
        headers = ['ID', 'Date', 'Filename', 'Status', 'Items Processed',
                   'Success Count', 'Error Count', 'User']

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Add data rows
        for row_num, log in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=log.id)
            ws.cell(row=row_num, column=2, value=log.created_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=3, value=log.file_name)
            ws.cell(row=row_num, column=4, value=log.status)
            ws.cell(row=row_num, column=5, value=log.total_records)
            ws.cell(row=row_num, column=6, value=log.success_count)
            ws.cell(row=row_num, column=7, value=log.error_count)
            ws.cell(row=row_num, column=8, value=log.created_by.username)

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="import_logs.xlsx"'

        # Save workbook to response
        wb.save(response)
        return response

    return HttpResponse("Unsupported format", status=400)


class ImportError(models.Model):
    import_log = models.ForeignKey(ImportLog, related_name='errors', on_delete=models.CASCADE)
    row_number = models.IntegerField()
    error_message = models.TextField()
    row_data = models.TextField(blank=True)
    field_name = models.CharField(max_length=100, blank=True)
    field_value = models.TextField(blank=True)

    class Meta:
        ordering = ['row_number']


class UserProfile(models.Model):
    user = models.OneToOneField(User, on_delete=models.CASCADE, related_name='profile')
    departments = models.ManyToManyField(Department, related_name='user_profiles', blank=True)

    def __str__(self):
        return f"Profil von {self.user.username}"


# ProductPhoto - für Produktfotos
class ProductPhoto(models.Model):
    """Model for product photos."""
    product = models.ForeignKey(Product, on_delete=models.CASCADE, related_name='photos')
    image = models.ImageField(upload_to='product_photos/')
    is_primary = models.BooleanField(default=False)
    caption = models.CharField(max_length=255, blank=True)
    upload_date = models.DateTimeField(auto_now_add=True)

    def __str__(self):
        return f"Foto für {self.product.name}"

    def save(self, *args, **kwargs):
        # Wenn dieses Foto als primär markiert wird, alle anderen Fotos dieses Produkts auf nicht-primär setzen
        if self.is_primary:
            ProductPhoto.objects.filter(product=self.product, is_primary=True).update(is_primary=False)
        super().save(*args, **kwargs)

    class Meta:
        ordering = ['-is_primary', '-upload_date']


# ProductAttachment - für Dokumente und andere Anhänge
def attachment_path(instance, filename):
    # Datei wird hochgeladen zu MEDIA_ROOT/product_attachments/product_<id>/<filename>
    return os.path.join('product_attachments', f'product_{instance.product.id}', filename)


class ProductAttachment(models.Model):
    """Model for product attachments (documents, manuals, etc.)."""
    product = models.ForeignKey(Product, on_delete=models.CASCADE, related_name='attachments')
    file = models.FileField(upload_to=attachment_path)
    title = models.CharField(max_length=255)
    description = models.TextField(blank=True)
    file_type = models.CharField(max_length=100, blank=True)  # PDF, DOCX, etc.
    upload_date = models.DateTimeField(auto_now_add=True)

    def __str__(self):
        return self.title

    def save(self, *args, **kwargs):
        # Automatisch den Dateityp anhand der Dateierweiterung bestimmen
        if self.file and not self.file_type:
            extension = os.path.splitext(self.file.name)[1].lower()[1:]
            self.file_type = extension
        super().save(*args, **kwargs)

    class Meta:
        ordering = ['-upload_date']


# ProductVariantType - definiert Variationstypen wie "Farbe", "Größe", etc.
class ProductVariantType(models.Model):
    """Model for types of product variants (e.g., color, size)."""
    name = models.CharField(max_length=100, unique=True)
    description = models.TextField(blank=True)

    def __str__(self):
        return self.name

    class Meta:
        ordering = ['name']


class ProductVariant(models.Model):
    """Model for product variants (e.g., 'XL', 'Red')."""
    parent_product = models.ForeignKey(Product, on_delete=models.CASCADE, related_name='variants')
    sku = models.CharField(max_length=50, unique=True)  # Eigene SKU für jede Variante
    name = models.CharField(max_length=255)  # Name der konkreten Variante z.B. "T-Shirt Rot XL"
    variant_type = models.ForeignKey(ProductVariantType, on_delete=models.CASCADE)
    value = models.CharField(max_length=100)  # Der Wert für den Typ, z.B. "Rot" für Farbe
    price_adjustment = models.DecimalField(max_digits=10, decimal_places=2, default=0)  # Preisanpassung (+/-)
    barcode = models.CharField(max_length=100, blank=True)
    is_active = models.BooleanField(default=True)
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)

    def __str__(self):
        return self.name

    class Meta:
        unique_together = ('parent_product', 'variant_type', 'value')
        ordering = ['parent_product', 'variant_type', 'value']
    
    @property
    def total_stock(self):
        """Berechnet den Gesamtbestand über alle Lager."""
        from django.db.models import Sum
        from inventory.models import VariantWarehouse
        
        # Prüfen, ob ein VariantWarehouse-Model existiert
        try:
            return VariantWarehouse.objects.filter(variant=self).aggregate(Sum('quantity'))['quantity__sum'] or 0
        except:
            # Fallback, falls kein VariantWarehouse existiert
            return 0


# SerialNumber - für Seriennummern
class SerialNumber(models.Model):
    """Model for tracking serial numbers."""
    product = models.ForeignKey(Product, on_delete=models.CASCADE, related_name='serial_numbers')
    variant = models.ForeignKey(ProductVariant, on_delete=models.CASCADE, related_name='serial_numbers',
                                null=True, blank=True)  # Optional, falls die Seriennummer zu einer Variante gehört
    serial_number = models.CharField(max_length=100, unique=True)
    status_choices = [
        ('in_stock', 'Auf Lager'),
        ('sold', 'Verkauft'),
        ('reserved', 'Reserviert'),
        ('defective', 'Defekt'),
        ('returned', 'Zurückgegeben'),
    ]
    status = models.CharField(max_length=20, choices=status_choices, default='in_stock')
    purchase_date = models.DateField(null=True, blank=True)  # Einkaufsdatum
    expiry_date = models.DateField(null=True, blank=True)  # Verfallsdatum
    notes = models.TextField(blank=True)
    warehouse = models.ForeignKey('inventory.Warehouse', on_delete=models.SET_NULL, null=True, blank=True)
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)

    def __str__(self):
        return self.serial_number

    @property
    def is_expired(self):
        """Check if the product is expired."""
        if self.expiry_date:
            return self.expiry_date < date.today()
        return False

    @property
    def days_until_expiry(self):
        """Calculate days until expiry."""
        if self.expiry_date:
            delta = self.expiry_date - date.today()
            return delta.days
        return None

    class Meta:
        ordering = ['-created_at']


# Batch/Lot - für Chargennummern
class BatchNumber(models.Model):
    """Model for tracking batch or lot numbers."""
    product = models.ForeignKey(Product, on_delete=models.CASCADE, related_name='batches')
    variant = models.ForeignKey(ProductVariant, on_delete=models.CASCADE, related_name='batches',
                                null=True, blank=True)  # Optional, falls die Charge zu einer Variante gehört
    batch_number = models.CharField(max_length=100)
    quantity = models.DecimalField(max_digits=10, decimal_places=2, default=0)  # Menge in dieser Charge
    production_date = models.DateField(null=True, blank=True)
    expiry_date = models.DateField(null=True, blank=True)
    supplier = models.ForeignKey('suppliers.Supplier', on_delete=models.SET_NULL, null=True, blank=True)
    warehouse = models.ForeignKey('inventory.Warehouse', on_delete=models.SET_NULL, null=True, blank=True)
    notes = models.TextField(blank=True)
    created_at = models.DateTimeField(auto_now_add=True)

    def __str__(self):
        return f"{self.product.name} - {self.batch_number}"

    @property
    def is_expired(self):
        """Check if the batch is expired."""
        if self.expiry_date:
            return self.expiry_date < date.today()
        return False

    @property
    def days_until_expiry(self):
        """Calculate days until expiry."""
        if self.expiry_date:
            delta = self.expiry_date - date.today()
            return delta.days
        return None

    class Meta:
        unique_together = ('product', 'batch_number', 'warehouse')
        ordering = ['-created_at']


class Currency(models.Model):
    """Model for currencies used in the system."""
    code = models.CharField(max_length=3, unique=True,
                            help_text="ISO 4217 currency code (e.g., EUR, USD)")
    name = models.CharField(max_length=100)
    symbol = models.CharField(max_length=5)
    decimal_places = models.IntegerField(default=2)
    is_default = models.BooleanField(default=False)
    is_active = models.BooleanField(default=True)
    exchange_rate = models.DecimalField(max_digits=10, decimal_places=6, default=1.0,
                                        help_text="Exchange rate relative to the default currency")
    created_at = models.DateTimeField(auto_now_add=True)
    updated_at = models.DateTimeField(auto_now=True)

    class Meta:
        verbose_name = "Währung"
        verbose_name_plural = "Währungen"
        ordering = ['code']

    def __str__(self):
        return f"{self.name} ({self.code})"

    def save(self, *args, **kwargs):
        # If this currency is set as default, set all others to non-default
        if self.is_default:
            Currency.objects.filter(is_default=True).exclude(pk=self.pk).update(is_default=False)

        # If no default currency exists, make this the default
        elif not Currency.objects.filter(is_default=True).exists():
            self.is_default = True

        super().save(*args, **kwargs)

    @classmethod
    def get_default_currency(cls):
        """Returns the default currency. If none is set, returns the first active currency."""
        default_currency = cls.objects.filter(is_default=True, is_active=True).first()
        if not default_currency:
            default_currency = cls.objects.filter(is_active=True).first()
        return default_currency


