import csv
import io
import mimetypes
import os
from datetime import datetime
from datetime import timedelta
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import User, Group, Permission
from django.db import models
from django.db.models import Count, Q, Sum, OuterRef, Subquery, F, Value, DecimalField
from django.db.models.functions import Coalesce
from django.http import FileResponse, Http404
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404
from django.shortcuts import render, redirect
from django.templatetags.static import static
from django.utils import timezone

from accessmanagement.decorators import permission_required
from accessmanagement.models import WarehouseAccess
from accessmanagement.permissions import PERMISSION_AREAS
from inventory.models import StockMovement, Warehouse, Department, StockTake
from order.models import PurchaseOrder, OrderSuggestion
from suppliers.models import Supplier, SupplierProduct
from .forms import ProductForm, CategoryForm, SupplierProductImportForm, CategoryImportForm, SupplierImportForm, \
    ProductImportForm, WarehouseImportForm, DepartmentImportForm, WarehouseProductImportForm, ProductPhotoForm, \
    ProductAttachmentForm, ProductVariantTypeForm, ProductVariantForm, SerialNumberForm, BulkSerialNumberForm, \
    BatchNumberForm, CurrencyForm, SerialNumberImportForm, BatchNumberImportForm
from .importers.batch_numbers import BatchNumberImporter
from .importers.categories import CategoryImporter
from .importers.products import ProductImporter, SupplierProductImporter
from .importers.serialnumbers import SerialNumberImporter
from .importers.suppliers import SupplierImporter
from .models import Product, Category, ImportLog, ProductWarehouse, ProductPhoto, ProductAttachment, ProductVariantType, \
    ProductVariant, SerialNumber, BatchNumber, Currency
from .utils.deletion import handle_delete_view
from .utils.files import delete_file_if_exists
from .utils.filters import filter_product_serials, apply_exact_filter, \
    apply_expiry_filter
from .utils.forms import handle_form_view
from .utils.imports import handle_csv_import
from .utils.pagination import paginate_queryset
from .utils.products import get_filtered_products
from .utils.stock import get_accessible_stock
from .utils.view_helpers import handle_model_delete, handle_model_update


@login_required
def dashboard(request):
    # Existierende Daten
    total_products = Product.objects.count()
    total_categories = Category.objects.count()
    total_suppliers = Supplier.objects.count()
    
    # Lager, auf die der Benutzer Zugriff hat
    if request.user.is_superuser:
        accessible_warehouses = Warehouse.objects.filter(is_active=True)
    else:
        accessible_warehouses = [w for w in Warehouse.objects.filter(is_active=True)
                                 if WarehouseAccess.has_access(request.user, w, 'view')]

    try:
        from order.services import generate_order_suggestions
        generate_order_suggestions()
    except ImportError:
        pass  # Ignorieren, falls die Funktion nicht importiert werden kann

    # Produkte mit niedrigem Bestand basierend auf zugänglichen Lagern
    low_stock_products = Product.objects.annotate(
        accessible_stock=Coalesce(
            Subquery(
                ProductWarehouse.objects.filter(
                    product=OuterRef('pk'),
                    warehouse__in=accessible_warehouses  # Korrekte Variable verwenden
                ).values('product')
                .annotate(total=Sum('quantity'))
                .values('total')[:1]
            ),
            Value(0, output_field=DecimalField(max_digits=10, decimal_places=2))  # DecimalField statt FloatField
        )
    ).filter(
        accessible_stock__lt=F('minimum_stock')
    )
    
    low_stock_count = low_stock_products.count()

    # Neue Daten für Lager und Abteilungen
    total_warehouses = accessible_warehouses.count()
    total_departments = Department.objects.count()
    active_stock_takes_count = StockTake.objects.filter(status='in_progress').count()

    # Erweiterte Daten für Tabellen
    recent_movements = StockMovement.objects.select_related('product', 'warehouse', 'created_by').order_by(
        '-created_at')[:10]

    # Kritische Bestände nach Lager
    low_stock_products_detail = []
    for product in low_stock_products[:10]:  # Limit auf 10 Produkte
        # Finde das Lager mit dem größten Bestand für dieses Produkt
        main_warehouse = ProductWarehouse.objects.filter(
            product=product,
            warehouse__in=accessible_warehouses
        ).order_by('-quantity').first()

        if main_warehouse:
            # Erstelle ein Dummy-Objekt mit den benötigten Eigenschaften
            class ProductWarehouseInfo:
                def __init__(self, product, warehouse, quantity):
                    self.product = product
                    self.warehouse = warehouse
                    self.quantity = quantity

            # Verwende das Dummy-Objekt statt main_warehouse direkt
            product_warehouse_info = ProductWarehouseInfo(
                product=product,
                warehouse=main_warehouse.warehouse,
                quantity=main_warehouse.quantity
            )

            low_stock_products_detail.append(product_warehouse_info)
        else:
            # Wenn kein Lager gefunden wurde, erstelle einen Eintrag mit Menge 0
            class ProductWarehouseInfo:
                def __init__(self, product, warehouse, quantity):
                    self.product = product
                    self.warehouse = None
                    self.quantity = quantity

            product_warehouse_info = ProductWarehouseInfo(
                product=product,
                warehouse=None,
                quantity=0
            )

            low_stock_products_detail.append(product_warehouse_info)

    # Ersetze low_stock_items durch unsere neue Liste
    low_stock_items = low_stock_products_detail

    # Aktive Inventuren
    active_stock_takes = StockTake.objects.select_related('warehouse').filter(status='in_progress')[:5]

    # Lagerübersicht mit Produktanzahl
    warehouses = accessible_warehouses
    for warehouse in warehouses:
        warehouse.product_count = ProductWarehouse.objects.filter(warehouse=warehouse, quantity__gt=0).count()

    pending_count = PurchaseOrder.objects.filter(status='pending').count()
    sent_count = PurchaseOrder.objects.filter(status='sent').count()
    partial_count = PurchaseOrder.objects.filter(status='partially_received').count()
    suggestion_count = OrderSuggestion.objects.count()

    # Letzte Bestellungen
    recent_orders = PurchaseOrder.objects.select_related('supplier').order_by('-order_date')[:5]

    context = {
        'total_products': total_products,
        'total_categories': total_categories,
        'total_suppliers': total_suppliers,
        'low_stock_count': low_stock_count,
        'total_warehouses': total_warehouses,
        'total_departments': total_departments,
        'active_stock_takes_count': active_stock_takes_count,
        'recent_movements': recent_movements,
        'low_stock_items': low_stock_items,
        'active_stock_takes': active_stock_takes,
        'warehouses': warehouses[:5],
        'pending_count': pending_count,
        'sent_count': sent_count,
        'partial_count': partial_count,
        'suggestion_count': suggestion_count,
        'recent_orders': recent_orders,
    }

    return render(request, 'dashboard.html', context)


@login_required
@permission_required('product', 'view')
def product_list(request):
    filtered_products = get_filtered_products(request)
    products = paginate_queryset(filtered_products, request.GET.get('page'), per_page=25)

    categories = Category.objects.all().order_by('name')
    context = {
        'products': products,
        'categories': categories,
        'search_query': request.GET.get('search', ''),
        'category_id': request.GET.get('category', ''),
        'stock_status': request.GET.get('stock_status', ''),
    }
    return render(request, 'core/product_list.html', context)



@login_required
@permission_required('product', 'view')
def low_stock_list(request):
    products = get_filtered_products(request, filter_low_stock=True)
    return render(request, 'core/low_stock_list.html', {'products': products, 'title': 'Kritische Bestände'})



@login_required
@permission_required('product', 'create')
def product_create(request):
    """Create a new product."""
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save()

            # Bestandsbewegung für Anfangsbestand erstellen, wenn > 0
            initial_stock = form.cleaned_data.get('initial_stock', 0)
            if initial_stock > 0:
                # Standardlager abrufen oder erstellen
                try:
                    warehouse = Warehouse.objects.filter(is_active=True).first()
                    if not warehouse:
                        warehouse = Warehouse.objects.create(
                            name='Hauptlager',
                            code='MAIN'
                        )
                    
                    # Produktlager-Eintrag erstellen
                    ProductWarehouse.objects.create(
                        product=product,
                        warehouse=warehouse,
                        quantity=initial_stock
                    )
                    
                    # Bestandsbewegung erstellen
                    StockMovement.objects.create(
                        product=product,
                        warehouse=warehouse,
                        quantity=initial_stock,
                        movement_type='in',
                        reference='Anfangsbestand',
                        created_by=request.user
                    )
                except Exception as e:
                    messages.error(request, f'Fehler bei der Erstellung des Anfangsbestands: {str(e)}')

            messages.success(request, f'Produkt "{product.name}" wurde erfolgreich erstellt.')
            return redirect('product_detail', pk=product.pk)
    else:
        form = ProductForm()

    context = {
        'form': form,
    }

    return render(request, 'core/product_form.html', context)


@login_required
@permission_required('product', 'edit')
def product_update(request, pk):
    """Update an existing product."""
    product = get_object_or_404(Product, pk=pk)

    # Lager abrufen, auf die der Benutzer Zugriff hat
    if request.user.is_superuser:
        accessible_warehouses = Warehouse.objects.filter(is_active=True)
    else:
        user_departments = request.user.profile.departments.all()
        warehouse_access = WarehouseAccess.objects.filter(
            department__in=user_departments
        ).values_list('warehouse', flat=True)
        accessible_warehouses = Warehouse.objects.filter(id__in=warehouse_access, is_active=True)

    # Aktuellen Bestand berechnen
    total_accessible_stock = ProductWarehouse.objects.filter(
        product=product,
        warehouse__in=accessible_warehouses
    ).aggregate(total=Sum('quantity'))['total'] or 0

    # Prüfen, ob Funktionen deaktivierbar sind
    has_variants = False
    has_serial_numbers = False
    has_batches = False
    has_expiry_dates = False

    if hasattr(product, 'variants'):
        has_variants = product.variants.exists()

    if hasattr(product, 'serial_numbers'):
        has_serial_numbers = product.serial_numbers.exists()
        # Prüfen, ob Seriennummern mit Verfallsdaten existieren
        has_expiry_dates = has_expiry_dates or product.serial_numbers.filter(expiry_date__isnull=False).exists()

    if hasattr(product, 'batches'):
        has_batches = product.batches.exists()
        # Prüfen, ob Chargen mit Verfallsdaten existieren
        has_expiry_dates = has_expiry_dates or product.batches.filter(expiry_date__isnull=False).exists()

    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            # Das Produkt ohne Bestandsänderung speichern
            product = form.save()

            messages.success(request, f'Produkt "{product.name}" wurde erfolgreich aktualisiert.')
            return redirect('product_detail', pk=product.pk)
    else:
        form = ProductForm(instance=product)

    context = {
        'form': form,
        'total_accessible_stock': total_accessible_stock,
        'has_variants': has_variants,
        'has_serial_numbers': has_serial_numbers,
        'has_batches': has_batches,
        'has_expiry_dates': has_expiry_dates
    }

    return render(request, 'core/product_form.html', context)


@login_required
@permission_required('product', 'view')
def category_list(request):
    """List all categories."""
    categories = Category.objects.all().order_by('name')

    # Für jede Kategorie die Anzahl der zugehörigen Produkte ermitteln
    categories_with_counts = []
    for category in categories:
        product_count = Product.objects.filter(category=category).count()
        categories_with_counts.append({
            'category': category,
            'product_count': product_count
        })

    context = {
        'categories': categories_with_counts,
    }

    return render(request, 'core/category_list.html', context)


@login_required
@permission_required('product', 'create')
def category_create(request):
    """Create a new category."""
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Kategorie "{category.name}" wurde erfolgreich erstellt.')
            return redirect('category_list')
    else:
        form = CategoryForm()

    context = {
        'form': form,
    }

    return render(request, 'core/category_form.html', context)


@login_required
def profile(request):
    """User profile view."""
    # Hier könnten Sie die Form für die Profilbearbeitung hinzufügen
    # Für das MVP zeigen wir nur grundlegende Benutzerinformationen an

    context = {
        'user': request.user,
    }

    return render(request, 'auth/profile.html', context)


@login_required
@permission_required('product', 'edit')
def category_update(request, pk):
    """Update an existing category."""
    category = get_object_or_404(Category, pk=pk)

    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Kategorie "{category.name}" wurde erfolgreich aktualisiert.')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)

    context = {
        'form': form,
        'category': category,
    }

    return render(request, 'core/category_form.html', context)


# Ergänzung zu core/views.py

@login_required
@permission_required('import', 'view')
def import_dashboard(request):
    """Dashboard for all import functionality."""
    # Get recent import logs
    recent_imports = ImportLog.objects.all()[:10]

    context = {
        'recent_imports': recent_imports,
    }

    return render(request, 'core/import/dashboard.html', context)


@login_required
@permission_required('import', 'create')
def import_products(request):
    return handle_csv_import(
        form_class=ProductImportForm,
        importer_class=ProductImporter,
        request=request,
        template_name='core/import/import_form.html',
        success_redirect='import_log_detail',
        extra_context={
            'title': 'Produkte importieren',
            'template_file_url': static('files/product_import_template.csv'),
        }
    )



@login_required
@permission_required('import', 'create')
def import_suppliers(request):
    """Import suppliers from CSV."""
    if request.method == 'POST':
        form = SupplierImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                file_obj = request.FILES['file']
                delimiter = form.cleaned_data['delimiter']
                encoding = form.cleaned_data['encoding']
                skip_header = form.cleaned_data['skip_header']
                update_existing = form.cleaned_data['update_existing']

                importer = SupplierImporter(
                    file_obj=file_obj,
                    delimiter=delimiter,
                    encoding=encoding,
                    skip_header=skip_header,
                    update_existing=update_existing,
                    user=request.user
                )

                import_log = importer.run_import()

                messages.success(
                    request,
                    f"Import abgeschlossen. {import_log.successful_rows} von {import_log.total_rows} "
                    f"Lieferanten erfolgreich importiert ({import_log.success_rate()}%)."
                )

                if import_log.failed_rows > 0:
                    messages.warning(
                        request,
                        f"{import_log.failed_rows} Zeilen konnten nicht importiert werden. "
                        f"Siehe Details für weitere Informationen."
                    )

                return redirect('import_log_detail', pk=import_log.pk)

            except Exception as e:
                messages.error(request, f"Fehler beim Import: {str(e)}")
    else:
        form = SupplierImportForm()

    context = {
        'form': form,
        'title': 'Lieferanten importieren',
        'template_file_url': static('files/supplier_import_template.csv'),
    }

    return render(request, 'core/import/import_form.html', context)


@login_required
@permission_required('import', 'create')
def import_categories(request):
    """Import categories from CSV."""
    if request.method == 'POST':
        form = CategoryImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                file_obj = request.FILES['file']
                delimiter = form.cleaned_data['delimiter']
                encoding = form.cleaned_data['encoding']
                skip_header = form.cleaned_data['skip_header']
                update_existing = form.cleaned_data['update_existing']

                importer = CategoryImporter(
                    file_obj=file_obj,
                    delimiter=delimiter,
                    encoding=encoding,
                    skip_header=skip_header,
                    update_existing=update_existing,
                    user=request.user
                )

                import_log = importer.run_import()

                messages.success(
                    request,
                    f"Import abgeschlossen. {import_log.successful_rows} von {import_log.total_rows} "
                    f"Kategorien erfolgreich importiert ({import_log.success_rate()}%)."
                )

                if import_log.failed_rows > 0:
                    messages.warning(
                        request,
                        f"{import_log.failed_rows} Zeilen konnten nicht importiert werden. "
                        f"Siehe Details für weitere Informationen."
                    )

                return redirect('import_log_detail', pk=import_log.pk)

            except Exception as e:
                messages.error(request, f"Fehler beim Import: {str(e)}")
    else:
        form = CategoryImportForm()

    context = {
        'form': form,
        'title': 'Kategorien importieren',
        'template_file_url': static('files/category_import_template.csv'),
    }

    return render(request, 'core/import/import_form.html', context)


@login_required
@permission_required('import', 'create')
def import_supplier_products(request):
    """Import supplier-product relationships from CSV."""
    if request.method == 'POST':
        form = SupplierProductImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                file_obj = request.FILES['file']
                delimiter = form.cleaned_data['delimiter']
                encoding = form.cleaned_data['encoding']
                skip_header = form.cleaned_data['skip_header']
                update_existing = form.cleaned_data['update_existing']

                importer = SupplierProductImporter(
                    file_obj=file_obj,
                    delimiter=delimiter,
                    encoding=encoding,
                    skip_header=skip_header,
                    update_existing=update_existing,
                    user=request.user
                )

                import_log = importer.run_import()

                messages.success(
                    request,
                    f"Import abgeschlossen. {import_log.successful_rows} von {import_log.total_rows} "
                    f"Produkt-Lieferanten-Zuordnungen erfolgreich importiert ({import_log.success_rate()}%)."
                )

                if import_log.failed_rows > 0:
                    messages.warning(
                        request,
                        f"{import_log.failed_rows} Zeilen konnten nicht importiert werden. "
                        f"Siehe Details für weitere Informationen."
                    )

                return redirect('import_log_detail', pk=import_log.pk)

            except Exception as e:
                messages.error(request, f"Fehler beim Import: {str(e)}")
    else:
        form = SupplierProductImportForm()

    context = {
        'form': form,
        'title': 'Produkt-Lieferanten-Zuordnungen importieren',
        'template_file_url': static('files/supplier_product_import_template.csv'),
    }

    return render(request, 'core/import/import_form.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
@permission_required('import', 'create')
def import_warehouses(request):
    """Import warehouses from CSV file."""
    if request.method == 'POST':
        form = WarehouseImportForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']

            # Überprüfen des Dateityps
            if not csv_file.name.endswith('.csv'):
                messages.error(request, 'Bitte laden Sie eine CSV-Datei hoch.')
                return redirect('import_warehouses')

            # Datei in Speicher lesen
            file_data = csv_file.read().decode('utf-8')

            # Import-Log erstellen
            import_log = ImportLog.objects.create(
                file_name=csv_file.name,
                import_type='warehouses',
                started_by=request.user,
                status='processing'
            )

            try:
                lines = file_data.split('\n')
                # CSV-Header überprüfen
                header = lines[0].strip().split(',')
                expected_header = ['name', 'location', 'description', 'is_active']

                if not all(column in header for column in expected_header):
                    import_log.status = 'failed'
                    import_log.notes = f'Ungültiger CSV-Header. Erwarteter Header: {", ".join(expected_header)}'
                    import_log.save()
                    messages.error(request, import_log.notes)
                    return redirect('import_warehouses')

                # Index der Spalten ermitteln
                name_idx = header.index('name')
                location_idx = header.index('location')
                description_idx = header.index('description') if 'description' in header else None
                is_active_idx = header.index('is_active') if 'is_active' in header else None

                # Import starten
                created_count = 0
                updated_count = 0
                error_count = 0
                error_msgs = []

                # CSV-Zeilen verarbeiten (ohne Header)
                for i, line in enumerate(lines[1:], 1):
                    if not line.strip():  # Leere Zeile überspringen
                        continue

                    try:
                        # Zeile parsen
                        fields = line.strip().split(',')

                        name = fields[name_idx].strip()
                        location = fields[location_idx].strip()
                        description = fields[
                            description_idx].strip() if description_idx is not None and description_idx < len(
                            fields) else ''

                        # is_active Feld parsen
                        is_active = True  # Standardwert
                        if is_active_idx is not None and is_active_idx < len(fields):
                            is_active_value = fields[is_active_idx].strip().lower()
                            is_active = is_active_value in ['true', '1', 'yes', 'ja', 'y', 'j']

                        # Lager erstellen oder aktualisieren
                        warehouse, created = Warehouse.objects.update_or_create(
                            name=name,
                            defaults={
                                'location': location,
                                'description': description,
                                'is_active': is_active
                            }
                        )

                        if created:
                            created_count += 1
                        else:
                            updated_count += 1

                    except Exception as e:
                        error_count += 1
                        error_msg = f'Fehler in Zeile {i + 1}: {str(e)}'
                        error_msgs.append(error_msg)

                # Import-Log aktualisieren
                import_log.status = 'completed' if error_count == 0 else 'completed_with_errors'
                import_log.rows_processed = created_count + updated_count + error_count
                import_log.rows_created = created_count
                import_log.rows_updated = updated_count
                import_log.rows_error = error_count
                import_log.notes = f"Import abgeschlossen: {created_count} erstellt, {updated_count} aktualisiert, {error_count} Fehler"

                if error_msgs:
                    import_log.error_details = '\n'.join(error_msgs)

                import_log.save()

                if error_count > 0:
                    messages.warning(request,
                                     f'Import mit Fehlern abgeschlossen. {created_count} Lager erstellt, {updated_count} aktualisiert, {error_count} Fehler.')
                else:
                    messages.success(request,
                                     f'Import erfolgreich abgeschlossen. {created_count} Lager erstellt, {updated_count} aktualisiert.')

                return redirect('import_log_detail', pk=import_log.pk)

            except Exception as e:
                import_log.status = 'failed'
                import_log.notes = f'Fehler beim Import: {str(e)}'
                import_log.save()
                messages.error(request, import_log.notes)
                return redirect('import_warehouses')
    else:
        form = WarehouseImportForm()

    context = {
        'form': form,
        'title': 'Lager importieren',
        'description': 'Importieren Sie Lager aus einer CSV-Datei.',
        'expected_format': 'name,location,description,is_active',
        'example': 'Hauptlager,Berlin,Unser Hauptlager,true',
        'required_columns': ['name', 'location'],
        'optional_columns': ['description', 'is_active'],
    }

    return render(request, 'core/import/import_form.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
@permission_required('import', 'create')
def import_departments(request):
    """Import departments from CSV file."""
    if request.method == 'POST':
        form = DepartmentImportForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']

            # Überprüfen des Dateityps
            if not csv_file.name.endswith('.csv'):
                messages.error(request, 'Bitte laden Sie eine CSV-Datei hoch.')
                return redirect('import_departments')

            # Datei in Speicher lesen
            file_data = csv_file.read().decode('utf-8')

            # Import-Log erstellen
            import_log = ImportLog.objects.create(
                file_name=csv_file.name,
                import_type='departments',
                started_by=request.user,
                status='processing'
            )

            try:
                lines = file_data.split('\n')
                # CSV-Header überprüfen
                header = lines[0].strip().split(',')
                expected_header = ['name', 'code', 'manager_username']

                if not all(column in header for column in expected_header):
                    import_log.status = 'failed'
                    import_log.notes = f'Ungültiger CSV-Header. Erwarteter Header: {", ".join(expected_header)}'
                    import_log.save()
                    messages.error(request, import_log.notes)
                    return redirect('import_departments')

                # Index der Spalten ermitteln
                name_idx = header.index('name')
                code_idx = header.index('code')
                manager_username_idx = header.index('manager_username')

                # Import starten
                created_count = 0
                updated_count = 0
                error_count = 0
                error_msgs = []

                # CSV-Zeilen verarbeiten (ohne Header)
                for i, line in enumerate(lines[1:], 1):
                    if not line.strip():  # Leere Zeile überspringen
                        continue

                    try:
                        # Zeile parsen
                        fields = line.strip().split(',')

                        name = fields[name_idx].strip()
                        code = fields[code_idx].strip()
                        manager_username = fields[manager_username_idx].strip() if manager_username_idx < len(
                            fields) else ''

                        # Manager finden, falls angegeben
                        manager = None
                        if manager_username:
                            try:
                                manager = User.objects.get(username=manager_username)
                            except User.DoesNotExist:
                                error_count += 1
                                error_msg = f'Fehler in Zeile {i + 1}: Benutzer "{manager_username}" nicht gefunden.'
                                error_msgs.append(error_msg)
                                continue

                        # Abteilung erstellen oder aktualisieren
                        department, created = Department.objects.update_or_create(
                            code=code,
                            defaults={
                                'name': name,
                                'manager': manager
                            }
                        )

                        if created:
                            created_count += 1
                        else:
                            updated_count += 1

                    except Exception as e:
                        error_count += 1
                        error_msg = f'Fehler in Zeile {i + 1}: {str(e)}'
                        error_msgs.append(error_msg)

                # Import-Log aktualisieren
                import_log.status = 'completed' if error_count == 0 else 'completed_with_errors'
                import_log.rows_processed = created_count + updated_count + error_count
                import_log.rows_created = created_count
                import_log.rows_updated = updated_count
                import_log.rows_error = error_count
                import_log.notes = f"Import abgeschlossen: {created_count} erstellt, {updated_count} aktualisiert, {error_count} Fehler"

                if error_msgs:
                    import_log.error_details = '\n'.join(error_msgs)

                import_log.save()

                if error_count > 0:
                    messages.warning(request,
                                     f'Import mit Fehlern abgeschlossen. {created_count} Abteilungen erstellt, {updated_count} aktualisiert, {error_count} Fehler.')
                else:
                    messages.success(request,
                                     f'Import erfolgreich abgeschlossen. {created_count} Abteilungen erstellt, {updated_count} aktualisiert.')

                return redirect('import_log_detail', pk=import_log.pk)

            except Exception as e:
                import_log.status = 'failed'
                import_log.notes = f'Fehler beim Import: {str(e)}'
                import_log.save()
                messages.error(request, import_log.notes)
                return redirect('import_departments')
    else:
        form = DepartmentImportForm()

    context = {
        'form': form,
        'title': 'Abteilungen importieren',
        'description': 'Importieren Sie Abteilungen aus einer CSV-Datei.',
        'expected_format': 'name,code,manager_username',
        'example': 'Vertrieb,VT01,max.mustermann',
        'required_columns': ['name', 'code'],
        'optional_columns': ['manager_username'],
    }

    return render(request, 'core/import/import_form.html', context)


@login_required
@user_passes_test(lambda u: u.is_staff)
@permission_required('import', 'create')
def import_warehouse_products(request):
    """Import product stock to warehouses from CSV file."""
    if request.method == 'POST':
        form = WarehouseProductImportForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']

            # Überprüfen des Dateityps
            if not csv_file.name.endswith('.csv'):
                messages.error(request, 'Bitte laden Sie eine CSV-Datei hoch.')
                return redirect('import_warehouse_products')

            # Datei in Speicher lesen
            file_data = csv_file.read().decode('utf-8')

            # Import-Log erstellen
            import_log = ImportLog.objects.create(
                file_name=csv_file.name,
                import_type='warehouse_products',
                started_by=request.user,
                status='processing'
            )

            try:
                lines = file_data.split('\n')
                # CSV-Header überprüfen
                header = lines[0].strip().split(',')
                expected_header = ['product_sku', 'warehouse_name', 'quantity', 'reference']

                if not all(column in header for column in expected_header[:3]):  # Referenz ist optional
                    import_log.status = 'failed'
                    import_log.notes = f'Ungültiger CSV-Header. Erwarteter Header: {", ".join(expected_header[:3])}'
                    import_log.save()
                    messages.error(request, import_log.notes)
                    return redirect('import_warehouse_products')

                # Index der Spalten ermitteln
                product_sku_idx = header.index('product_sku')
                warehouse_name_idx = header.index('warehouse_name')
                quantity_idx = header.index('quantity')
                reference_idx = header.index('reference') if 'reference' in header else None

                # Zugriffsrechte prüfen und verfügbare Lager für den Benutzer ermitteln
                if request.user.is_superuser:
                    accessible_warehouses = Warehouse.objects.filter(is_active=True)
                else:
                    user_departments = request.user.departments.all()
                    warehouse_access = WarehouseAccess.objects.filter(
                        department__in=user_departments,
                        can_manage_stock=True
                    ).values_list('warehouse', flat=True)
                    accessible_warehouses = Warehouse.objects.filter(pk__in=warehouse_access, is_active=True)

                # Import starten
                created_count = 0
                updated_count = 0
                error_count = 0
                error_msgs = []

                # CSV-Zeilen verarbeiten (ohne Header)
                for i, line in enumerate(lines[1:], 1):
                    if not line.strip():  # Leere Zeile überspringen
                        continue

                    try:
                        # Zeile parsen
                        fields = line.strip().split(',')

                        product_sku = fields[product_sku_idx].strip()
                        warehouse_name = fields[warehouse_name_idx].strip()
                        quantity_str = fields[quantity_idx].strip()
                        reference = fields[reference_idx].strip() if reference_idx is not None and reference_idx < len(
                            fields) else 'CSV-Import'

                        # Produkt und Lager finden
                        try:
                            product = Product.objects.get(sku=product_sku)
                        except Product.DoesNotExist:
                            error_count += 1
                            error_msg = f'Fehler in Zeile {i + 1}: Produkt mit SKU "{product_sku}" nicht gefunden.'
                            error_msgs.append(error_msg)
                            continue

                        try:
                            warehouse = Warehouse.objects.get(name=warehouse_name, is_active=True)
                        except Warehouse.DoesNotExist:
                            error_count += 1
                            error_msg = f'Fehler in Zeile {i + 1}: Lager "{warehouse_name}" nicht gefunden oder inaktiv.'
                            error_msgs.append(error_msg)
                            continue

                        # Prüfen, ob der Benutzer Zugriff auf das Lager hat
                        if warehouse not in accessible_warehouses:
                            error_count += 1
                            error_msg = f'Fehler in Zeile {i + 1}: Keine Berechtigung für Lager "{warehouse_name}".'
                            error_msgs.append(error_msg)
                            continue

                        # Menge parsen
                        try:
                            quantity = Decimal(quantity_str)
                            if quantity < 0:
                                raise ValueError("Menge kann nicht negativ sein")
                        except ValueError:
                            error_count += 1
                            error_msg = f'Fehler in Zeile {i + 1}: Ungültige Menge "{quantity_str}".'
                            error_msgs.append(error_msg)
                            continue

                        # Aktuelle Bestandsmenge im Lager ermitteln
                        try:
                            product_warehouse = ProductWarehouse.objects.get(product=product, warehouse=warehouse)
                            current_quantity = product_warehouse.quantity
                            is_new = False
                        except ProductWarehouse.DoesNotExist:
                            current_quantity = Decimal('0')
                            is_new = True

                        # Bestandskorrektur nur vornehmen, wenn sich die Menge ändert
                        if current_quantity != quantity:
                            # Bestandsbewegung erstellen
                            StockMovement.objects.create(
                                product=product,
                                warehouse=warehouse,
                                quantity=quantity,
                                movement_type='adj',
                                reference=reference,
                                notes=f'CSV-Import: Bestand von {current_quantity} auf {quantity} angepasst',
                                created_by=request.user
                            )

                            # Bestand im Lager aktualisieren
                            if is_new:
                                ProductWarehouse.objects.create(
                                    product=product,
                                    warehouse=warehouse,
                                    quantity=quantity
                                )
                                created_count += 1
                            else:
                                product_warehouse.quantity = quantity
                                product_warehouse.save()
                                updated_count += 1

                            # Gesamtbestand des Produkts aktualisieren
                            total_stock = get_accessible_stock(product, accessible_warehouses)
                            product.current_stock = total_stock
                            product.save()
                        else:
                            # Keine Änderung notwendig
                            if is_new:
                                created_count += 1
                            else:
                                updated_count += 1

                    except Exception as e:
                        error_count += 1
                        error_msg = f'Fehler in Zeile {i + 1}: {str(e)}'
                        error_msgs.append(error_msg)

                # Import-Log aktualisieren
                import_log.status = 'completed' if error_count == 0 else 'completed_with_errors'
                import_log.rows_processed = created_count + updated_count + error_count
                import_log.rows_created = created_count
                import_log.rows_updated = updated_count
                import_log.rows_error = error_count
                import_log.notes = f"Import abgeschlossen: {created_count} erstellt, {updated_count} aktualisiert, {error_count} Fehler"

                if error_msgs:
                    import_log.error_details = '\n'.join(error_msgs)

                import_log.save()

                if error_count > 0:
                    messages.warning(request,
                                     f'Import mit Fehlern abgeschlossen. {created_count} Bestände erstellt, {updated_count} aktualisiert, {error_count} Fehler.')
                else:
                    messages.success(request,
                                     f'Import erfolgreich abgeschlossen. {created_count} Bestände erstellt, {updated_count} aktualisiert.')

                return redirect('import_log_detail', pk=import_log.pk)

            except Exception as e:
                import_log.status = 'failed'
                import_log.notes = f'Fehler beim Import: {str(e)}'
                import_log.save()
                messages.error(request, import_log.notes)
                return redirect('import_warehouse_products')
    else:
        form = WarehouseProductImportForm()

    context = {
        'form': form,
        'title': 'Produkt-Lager-Bestände importieren',
        'description': 'Importieren Sie Produktbestände für Lager aus einer CSV-Datei.',
        'expected_format': 'product_sku,warehouse_name,quantity,reference',
        'example': 'P1001,Hauptlager,25.5,Anfangsbestand',
        'required_columns': ['product_sku', 'warehouse_name', 'quantity'],
        'optional_columns': ['reference'],
    }

    return render(request, 'core/import/import_form.html', context)


# Updated import-related views in core/views.py

# Updated import-related views in core/views.py

@login_required
@permission_required('import', 'view')
def import_log_list(request):
    """List all import logs."""
    # Base queryset
    queryset = ImportLog.objects.all()

    # Apply filters
    if 'search' in request.GET and request.GET['search']:
        search = request.GET['search']
        queryset = queryset.filter(file_name__icontains=search)

    if 'status' in request.GET and request.GET['status']:
        status = request.GET['status']
        queryset = queryset.filter(status=status)

    if 'user' in request.GET and request.GET['user']:
        user_id = request.GET['user']
        queryset = queryset.filter(created_by_id=user_id)

    if 'type' in request.GET and request.GET['type']:
        import_type = request.GET['type']
        queryset = queryset.filter(import_type=import_type)

    # Apply sorting
    sort_param = request.GET.get('sort', '-created_at')
    if sort_param:
        queryset = queryset.order_by(sort_param)

    # Pagination
    queryset = ImportLog.objects.all()
    import_logs = paginate_queryset(queryset, request.GET.get('page'), per_page=20)

    context = {
        'import_logs': import_logs,
        'users': User.objects.all(),
    }

    return render(request, 'core/import/import_log_list.html', context)


@login_required
@permission_required('import', 'view')
def import_log_detail(request, pk):
    """Show details of an import log."""
    log = get_object_or_404(ImportLog, pk=pk)
    errors = log.errors.all() if hasattr(log, 'errors') else ImportError.objects.filter(import_log=log)

    context = {
        'log': log,
        'errors': errors,
    }

    return render(request, 'core/import/import_log_detail.html', context)


@login_required
@permission_required('import', 'view')
def download_error_file(request, log_id):
    """Download the error file for an import log."""
    log = get_object_or_404(ImportLog, pk=log_id)

    # Check if actual error file exists and return it if it does
    if log.error_file and log.error_file.name:
        response = HttpResponse(log.error_file.read(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{log.file_name}_errors.csv"'
        return response

    # Otherwise generate a CSV with errors
    errors = log.errors.all() if hasattr(log, 'errors') else ImportError.objects.filter(import_log=log)

    if not errors.exists():
        messages.warning(request, 'Keine Fehler zum Herunterladen vorhanden.')
        return redirect('import_log_detail', pk=log_id)

    # CSV-Datei erstellen
    response = HttpResponse(content_type='text/csv')
    response[
        'Content-Disposition'] = f'attachment; filename="import_fehler_{log.pk}_{datetime.now().strftime("%Y%m%d")}.csv"'

    # CSV-Writer einrichten
    writer = csv.writer(response)

    # Header schreiben
    writer.writerow(['Zeile', 'Feld', 'Fehler', 'Wert'])

    # Fehler schreiben
    for error in errors:
        writer.writerow([error.row_number, error.field_name, error.error_message, error.field_value])

    return response


@login_required
@permission_required('import', 'delete')
def delete_import_log(request, log_id):
    """Delete an import log."""
    if request.method == 'POST':
        log = get_object_or_404(ImportLog, pk=log_id)

        # Save name for confirmation message
        log_name = f"Import #{log.pk} ({log.import_type})"

        # Delete all errors associated with the log
        from core.models import ImportError as ImportErrorModel
        ImportErrorModel.objects.filter(import_log=log).delete()

        # Delete the log
        log.delete()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # For AJAX requests
            return JsonResponse({'success': True, 'message': f'{log_name} wurde gelöscht.'})
        else:
            # For normal requests
            messages.success(request, f'{log_name} wurde gelöscht.')
            return redirect('import_log_list')

    # If not a POST request, 405 Method Not Allowed
    return JsonResponse({'success': False, 'message': 'Nur POST-Anfragen sind erlaubt.'}, status=405)


@login_required
@permission_required('import', 'delete')
def bulk_delete_import_logs(request):
    """Delete multiple import logs."""
    if request.method == 'POST':
        # Get IDs from the request (comma-separated list)
        ids_str = request.POST.get('ids', '')

        if not ids_str:
            return JsonResponse({'success': False, 'message': 'Keine IDs angegeben.'}, status=400)

        try:
            # Convert IDs to a list
            ids = [int(id_str.strip()) for id_str in ids_str.split(',') if id_str.strip()]

            # Delete errors associated with the logs
            from core.models import ImportError as ImportErrorModel
            ImportErrorModel.objects.filter(import_log_id__in=ids).delete()

            # Delete logs
            deleted_count = ImportLog.objects.filter(pk__in=ids).delete()[0]

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # For AJAX requests
                return JsonResponse({'success': True, 'message': f'{deleted_count} Import-Logs wurden gelöscht.'})
            else:
                # For normal requests
                messages.success(request, f'{deleted_count} Import-Logs wurden gelöscht.')
                return redirect('import_log_list')

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Fehler beim Löschen: {str(e)}'}, status=500)

    # If not a POST request, 405 Method Not Allowed
    return JsonResponse({'success': False, 'message': 'Nur POST-Anfragen sind erlaubt.'}, status=405)


@login_required
@permission_required('report', 'view')
def export_import_logs(request):
    """Export import logs to CSV or Excel."""
    # Base queryset
    queryset = ImportLog.objects.all()

    # Apply filters
    import_type = request.GET.get('type')
    if import_type:
        queryset = queryset.filter(import_type=import_type)

    # Check for specific IDs
    ids_str = request.GET.get('ids', '')
    if ids_str:
        ids = [int(id_str.strip()) for id_str in ids_str.split(',') if id_str.strip()]
        queryset = queryset.filter(pk__in=ids)

    # Determine format
    export_format = request.GET.get('format', 'csv').lower()

    # CSV export
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="import_logs_{datetime.now().strftime("%Y%m%d")}.csv"'

        writer = csv.writer(response)

        # Write header
        writer.writerow(['ID', 'Type', 'Filename', 'Status', 'Total Rows', 'Successful Rows',
                         'Error Rows', 'Success Rate', 'User', 'Created At'])

        # Write data
        for log in queryset:
            writer.writerow([
                log.pk,
                log.import_type,
                log.file_name,
                log.status,
                log.total_records,
                log.success_count,
                log.error_count,
                f"{log.success_rate()}%",
                log.created_by.username if log.created_by else 'System',
                log.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])

        return response

    # Excel export
    elif export_format == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            messages.error(request, 'Excel-Export ist nicht verfügbar. Bitte installieren Sie openpyxl.')
            return redirect('import_log_list')

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Logs"

        # Define headers
        headers = ['ID', 'Type', 'Filename', 'Status', 'Total Rows', 'Successful Rows',
                   'Error Rows', 'Success Rate', 'User', 'Created At']

        # Define styles for header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Write and format header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for row_num, log in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=log.pk)
            ws.cell(row=row_num, column=2, value=log.import_type)
            ws.cell(row=row_num, column=3, value=log.file_name)
            ws.cell(row=row_num, column=4, value=log.status)
            ws.cell(row=row_num, column=5, value=log.total_records)
            ws.cell(row=row_num, column=6, value=log.success_count)
            ws.cell(row=row_num, column=7, value=log.error_count)
            ws.cell(row=row_num, column=8, value=f"{log.success_rate()}%")
            ws.cell(row=row_num, column=9, value=log.created_by.username if log.created_by else 'System')
            ws.cell(row=row_num, column=10, value=log.created_at.strftime('%Y-%m-%d %H:%M:%S'))

        # Adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            # Minimum width based on header + some space
            max_length = len(header) + 2

            # Check width based on data
            for row_num in range(2, len(queryset) + 2):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)) + 2)

            ws.column_dimensions[column_letter].width = max_length

        # Write Excel file to a BytesIO stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="import_logs_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        return response

    # Unsupported format
    else:
        messages.error(request, f'Das Format "{export_format}" wird nicht unterstützt.')
        return redirect('import_log_list')


@login_required
@user_passes_test(lambda u: u.is_superuser)
def get_user_permissions(request):
    """AJAX-Endpunkt zum Laden der Benutzerberechtigungen."""
    user_id = request.GET.get('user_id')

    try:
        user = User.objects.get(pk=user_id)

        # Gruppen
        groups = list(user.groups.values_list('id', flat=True))

        # Direkte Berechtigungen
        direct_permissions = list(user.user_permissions.values_list('id', flat=True))

        # Effektive Berechtigungen (inkl. Gruppen)
        effective_permissions = []

        # Aus Gruppen
        for group in user.groups.all():
            for perm in group.permissions.all():
                effective_permissions.append({
                    'id': perm.id,
                    'name': perm.name,
                    'codename': perm.codename,
                    'source': f'Gruppe: {group.name}'
                })

        # Direkte
        for perm in user.user_permissions.all():
            effective_permissions.append({
                'id': perm.id,
                'name': perm.name,
                'codename': perm.codename,
                'source': 'Direkt zugewiesen'
            })

        return JsonResponse({
            'groups': groups,
            'direct_permissions': direct_permissions,
            'effective_permissions': effective_permissions
        })

    except User.DoesNotExist:
        return JsonResponse({'error': 'Benutzer nicht gefunden'}, status=404)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def permission_management(request):
    """Verwaltung der Benutzerberechtigungen."""
    users = User.objects.all().order_by('username')
    groups = Group.objects.all().order_by('name')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'assign_group':
            user_id = request.POST.get('user_id')
            group_id = request.POST.get('group_id')

            try:
                user = User.objects.get(pk=user_id)
                group = Group.objects.get(pk=group_id)

                if request.POST.get('assign') == 'true':
                    user.groups.add(group)
                    messages.success(request, f'Benutzer {user.username} wurde zur Gruppe {group.name} hinzugefügt.')
                else:
                    user.groups.remove(group)
                    messages.success(request, f'Benutzer {user.username} wurde aus der Gruppe {group.name} entfernt.')

            except (User.DoesNotExist, Group.DoesNotExist):
                messages.error(request, "Benutzer oder Gruppe nicht gefunden.")

        elif action == 'direct_permission':
            user_id = request.POST.get('user_id')
            perm_id = request.POST.get('permission_id')

            try:
                user = User.objects.get(pk=user_id)
                perm = Permission.objects.get(pk=perm_id)

                if request.POST.get('assign') == 'true':
                    user.user_permissions.add(perm)
                    messages.success(request, f'Berechtigung {perm.name} wurde {user.username} direkt zugewiesen.')
                else:
                    user.user_permissions.remove(perm)
                    messages.success(request, f'Berechtigung {perm.name} wurde von {user.username} entfernt.')

            except (User.DoesNotExist, Permission.DoesNotExist):
                messages.error(request, "Benutzer oder Berechtigung nicht gefunden.")

    # Berechtigungen nach Bereichen gruppieren
    permissions = {}
    for area in PERMISSION_AREAS.keys():
        area_perms = Permission.objects.filter(codename__contains=f'_{area}').order_by('codename')
        permissions[area] = area_perms

    context = {
        'users': users,
        'groups': groups,
        'permissions': permissions,
        'permission_areas': PERMISSION_AREAS,
    }

    return render(request, 'core/permission_management.html', context)

#111
# ------------------------------------------------------------------------------
# Produktfotos
# ------------------------------------------------------------------------------

@login_required
@permission_required('product', 'view')
def product_photos(request, pk):
    """Zeigt alle Fotos eines Produkts an."""
    product = get_object_or_404(Product, pk=pk)
    photos = product.photos.all()

    context = {
        'product': product,
        'photos': photos,
    }

    return render(request, 'core/product/product_photos.html', context)


@login_required
@permission_required('product', 'edit')
def product_photo_add(request, pk):
    product = get_object_or_404(Product, pk=pk)

    def form_instance_save(form):
        photo = form.save(commit=False)
        photo.product = product
        photo.save()
        messages.success(request, 'Foto wurde erfolgreich hinzugefügt.')
        return redirect('product_photos', pk=product.pk)

    return handle_form_view(
        request,
        form_class=ProductPhotoForm,
        template='core/product/product_photo_form.html',
        redirect_url=f'/produkte/{pk}/fotos',  # oder benutze reverse()
        context_extra={'product': product}
    )



@login_required
@permission_required('product', 'delete')
def product_photo_delete(request, pk, photo_id):
    product = get_object_or_404(Product, pk=pk)
    photo = get_object_or_404(ProductPhoto, pk=photo_id, product=product)

    return handle_delete_view(
        request,
        obj=photo,
        redirect_url='product_photos',  # oder f'/produkte/{pk}/fotos'
        confirm_template='core/product/product_photo_confirm_delete.html',
        context={'product': product, 'photo': photo},
        filefield=photo.image,
        delete_file_func=delete_file_if_exists
    )



@login_required
@permission_required('product', 'edit')
def product_photo_set_primary(request, pk, photo_id):
    """Setzt ein Foto als Hauptfoto des Produkts."""
    product = get_object_or_404(Product, pk=pk)
    photo = get_object_or_404(ProductPhoto, pk=photo_id, product=product)

    # Erst alle anderen Fotos als nicht-primär markieren
    ProductPhoto.objects.filter(product=product).update(is_primary=False)

    # Dann dieses Foto als primär markieren
    photo.is_primary = True
    photo.save()

    messages.success(request, 'Hauptfoto wurde erfolgreich gesetzt.')

    # AJAX-Request oder normale Anfrage unterscheiden
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    else:
        return redirect('product_photos', pk=product.pk)


# ------------------------------------------------------------------------------
# Produktanhänge
# ------------------------------------------------------------------------------

@login_required
@permission_required('product', 'view')
def product_attachments(request, pk):
    """Zeigt alle Anhänge eines Produkts an."""
    product = get_object_or_404(Product, pk=pk)
    attachments = product.attachments.all()

    context = {
        'product': product,
        'attachments': attachments,
    }

    return render(request, 'core/product/product_attachments.html', context)


@login_required
@permission_required('product', 'edit')
def product_attachment_add(request, pk):
    """Fügt einen Anhang zu einem Produkt hinzu."""
    product = get_object_or_404(Product, pk=pk)

    if request.method == 'POST':
        form = ProductAttachmentForm(request.POST, request.FILES)
        if form.is_valid():
            attachment = form.save(commit=False)
            attachment.product = product
            attachment.save()

            messages.success(request, 'Anhang wurde erfolgreich hinzugefügt.')
            return redirect('product_attachments', pk=product.pk)
    else:
        form = ProductAttachmentForm()

    context = {
        'product': product,
        'form': form,
    }

    return render(request, 'core/product/product_attachment_form.html', context)


@login_required
@permission_required('product', 'delete')
def product_attachment_delete(request, pk, attachment_id):
    product = get_object_or_404(Product, pk=pk)
    attachment = get_object_or_404(ProductAttachment, pk=attachment_id, product=product)

    # Datei löschen
    if request.method == 'POST' and attachment.file and os.path.isfile(attachment.file.path):
        os.remove(attachment.file.path)

    context = handle_model_delete(
        request,
        instance=attachment,
        model_name_singular='Anhang',
        success_redirect='product_attachments',
        context_extra={'product': product}
    )

    return render(request, 'core/product/product_attachment_confirm_delete.html', context)



@login_required
@permission_required('product', 'view')
def product_attachment_download(request, pk, attachment_id):
    """Lädt einen Produktanhang herunter."""
    product = get_object_or_404(Product, pk=pk)
    attachment = get_object_or_404(ProductAttachment, pk=attachment_id, product=product)

    if not attachment.file:
        raise Http404("Die angeforderte Datei existiert nicht.")

    file_path = attachment.file.path

    if not os.path.exists(file_path):
        raise Http404("Die angeforderte Datei existiert nicht.")

    # Mime-Type bestimmen
    content_type, encoding = mimetypes.guess_type(file_path)
    if content_type is None:
        content_type = 'application/octet-stream'

    # Datei öffnen und als Response senden
    try:
        response = FileResponse(open(file_path, 'rb'), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
        return response
    except Exception as e:
        messages.error(request, f"Fehler beim Herunterladen der Datei: {str(e)}")
        return redirect('product_attachments', pk=product.pk)


# ------------------------------------------------------------------------------
# Variantentypen
# ------------------------------------------------------------------------------

@login_required
@permission_required('product', 'view')
def variant_type_list(request):
    """Zeigt alle Variantentypen an."""
    variant_types = ProductVariantType.objects.all()

    # Varianten pro Typ zählen
    variant_types_with_count = []
    for vt in variant_types:
        variant_count = ProductVariant.objects.filter(variant_type=vt).count()
        variant_types_with_count.append({
            'type': vt,
            'variant_count': variant_count
        })

    context = {
        'variant_types': variant_types_with_count,
    }

    return render(request, 'core/variant_type/variant_type_list.html', context)


@login_required
@permission_required('product', 'create')
def variant_type_add(request):
    """Erstellt einen neuen Variantentyp."""
    if request.method == 'POST':
        form = ProductVariantTypeForm(request.POST)
        if form.is_valid():
            variant_type = form.save()
            messages.success(request, f'Variantentyp "{variant_type.name}" wurde erfolgreich erstellt.')
            return redirect('variant_type_list')
    else:
        form = ProductVariantTypeForm()

    context = {
        'form': form,
    }

    return render(request, 'core/variant_type/variant_type_form.html', context)


@login_required
@permission_required('product', 'edit')
def variant_type_update(request, pk):
    """Aktualisiert einen Variantentyp."""
    variant_type = get_object_or_404(ProductVariantType, pk=pk)

    if request.method == 'POST':
        form = ProductVariantTypeForm(request.POST, instance=variant_type)
        if form.is_valid():
            variant_type = form.save()
            messages.success(request, f'Variantentyp "{variant_type.name}" wurde erfolgreich aktualisiert.')
            return redirect('variant_type_list')
    else:
        form = ProductVariantTypeForm(instance=variant_type)

    context = {
        'form': form,
        'variant_type': variant_type,
    }

    return render(request, 'core/variant_type/variant_type_form.html', context)


@login_required
@permission_required('product', 'delete')
def variant_type_delete(request, pk):
    """Löscht einen Variantentyp."""
    variant_type = get_object_or_404(ProductVariantType, pk=pk)

    # Prüfen, ob Varianten diesen Typ verwenden
    variant_count = ProductVariant.objects.filter(variant_type=variant_type).count()

    if request.method == 'POST':
        if variant_count > 0 and 'confirm_delete' not in request.POST:
            messages.error(request, f'Dieser Variantentyp wird von {variant_count} Varianten verwendet. '
                                    f'Löschen bestätigen, um trotzdem fortzufahren.')
            return redirect('variant_type_delete', pk=variant_type.pk)

        variant_type.delete()
        messages.success(request, f'Variantentyp "{variant_type.name}" wurde erfolgreich gelöscht.')
        return redirect('variant_type_list')

    context = {
        'variant_type': variant_type,
        'variant_count': variant_count,
    }

    return render(request, 'core/variant_type/variant_type_confirm_delete.html', context)


# ------------------------------------------------------------------------------
# Produktvarianten
# ------------------------------------------------------------------------------

@login_required
@permission_required('product', 'view')
def product_variants(request, pk):
    """Zeigt alle Varianten eines Produkts an."""
    product = get_object_or_404(Product, pk=pk)
    variants = product.variants.all()

    context = {
        'product': product,
        'variants': variants,
    }

    return render(request, 'core/product/product_variants.html', context)


@login_required
@permission_required('product', 'view')
def product_variant_detail(request, pk, variant_id):
    """Zeigt Details zu einer Produktvariante an."""
    product = get_object_or_404(Product, pk=pk)
    variant = get_object_or_404(ProductVariant, pk=variant_id, parent_product=product)

    # Seriennummern für diese Variante
    serial_numbers = SerialNumber.objects.filter(variant=variant)

    # Chargen für diese Variante
    batches = BatchNumber.objects.filter(variant=variant)

    context = {
        'product': product,
        'variant': variant,
        'serial_numbers': serial_numbers,
        'batches': batches,
    }

    return render(request, 'core/product/product_variant_detail.html', context)


@login_required
@permission_required('product', 'create')
def product_variant_add(request, pk):
    """Fügt eine Variante zu einem Produkt hinzu."""
    product = get_object_or_404(Product, pk=pk)

    # Prüfen, ob das Produkt für Varianten konfiguriert ist
    if not product.has_variants:
        product.has_variants = True
        product.save()
        messages.info(request, 'Das Produkt wurde für Varianten aktiviert.')

    if request.method == 'POST':
        form = ProductVariantForm(request.POST)
        if form.is_valid():
            variant = form.save(commit=False)
            variant.parent_product = product
            variant.save()
            
            # Bestandsbewegung für Anfangsbestand erstellen, wenn > 0
            initial_stock = form.cleaned_data.get('initial_stock', 0)
            if initial_stock > 0:
                # Standardlager abrufen oder erstellen
                try:
                    warehouse = Warehouse.objects.filter(is_active=True).first()
                    if not warehouse:
                        warehouse = Warehouse.objects.create(
                            name='Hauptlager',
                            code='MAIN'
                        )
                    
                    # Versuchen, ein VariantWarehouse-Modell zu verwenden, falls vorhanden
                    try:
                        from inventory.models import VariantWarehouse
                        
                        # Varianten-Lager-Eintrag erstellen
                        VariantWarehouse.objects.create(
                            variant=variant,
                            warehouse=warehouse,
                            quantity=initial_stock
                        )
                        
                        # Optional: Bestandsbewegung erstellen, falls dein System dies unterstützt
                        try:
                            from inventory.models import VariantStockMovement
                            VariantStockMovement.objects.create(
                                variant=variant,
                                warehouse=warehouse,
                                quantity=initial_stock,
                                movement_type='in',
                                reference='Anfangsbestand',
                                created_by=request.user
                            )
                        except (ImportError, AttributeError):
                            # Keine Bewegung erstellen, wenn kein VariantStockMovement existiert
                            pass
                            
                    except (ImportError, AttributeError):
                        # Wenn kein VariantWarehouse-Modell existiert, speichern wir den Bestand nicht
                        messages.warning(request, 'Bestandserfassung für Varianten ist nicht konfiguriert.')
                        
                except Exception as e:
                    messages.error(request, f'Fehler bei der Erstellung des Anfangsbestands: {str(e)}')

            messages.success(request, 'Produktvariante wurde erfolgreich hinzugefügt.')
            return redirect('product_variants', pk=product.pk)
    else:
        # Name-Vorschlag generieren
        form = ProductVariantForm(initial={
            'name': product.name,
            'initial_stock': 0
        })

    context = {
        'product': product,
        'form': form,
    }

    return render(request, 'core/product/product_variant_form.html', context)


@login_required
@permission_required('product', 'edit')
def product_variant_update(request, pk, variant_id):
    product = get_object_or_404(Product, pk=pk)
    variant = get_object_or_404(ProductVariant, pk=variant_id, parent_product=product)

    context = handle_model_update(
        request,
        instance=variant,
        form_class=ProductVariantForm,
        model_name_singular='Produktvariante',
        success_redirect='product_variants',
        context_extra={'product': product, 'variant': variant}
    )

    return render(request, 'core/product/product_variant_form.html', context)



@login_required
@permission_required('product', 'delete')
def product_variant_delete(request, pk, variant_id):
    """Löscht eine Produktvariante."""
    product = get_object_or_404(Product, pk=pk)
    variant = get_object_or_404(ProductVariant, pk=variant_id, parent_product=product)

    if request.method == 'POST':
        variant.delete()
        messages.success(request, 'Produktvariante wurde erfolgreich gelöscht.')

        # Wenn keine Varianten mehr übrig sind, has_variants zurücksetzen
        if not product.variants.exists():
            product.has_variants = False
            product.save()

        return redirect('product_variants', pk=product.pk)

    context = {
        'product': product,
        'variant': variant,
    }

    return render(request, 'core/product/product_variant_confirm_delete.html', context)


# ------------------------------------------------------------------------------
# Seriennummern
# ------------------------------------------------------------------------------

@login_required
@permission_required('product', 'view')
def product_serials(request, pk):
    """Zeigt alle Seriennummern eines Produkts an."""
    product = get_object_or_404(Product, pk=pk)

    filters = {
        'status': request.GET.get('status', ''),
        'warehouse': request.GET.get('warehouse', ''),
        'variant': request.GET.get('variant', ''),
        'search': request.GET.get('search', ''),
    }

    serials = SerialNumber.objects.filter(product=product)
    serials = filter_product_serials(serials, filters)
    serials = paginate_queryset(serials, request.GET.get('page'), per_page=50)

    status_counts = SerialNumber.objects.filter(product=product).values('status').annotate(
        count=Count('status')).order_by('status')
    status_stats = {item['status']: item['count'] for item in status_counts}

    context = {
        'product': product,
        'serials': serials,
        'status_stats': status_stats,
        'warehouses': Warehouse.objects.filter(is_active=True),
        'variants': product.variants.all(),
        'status_choices': SerialNumber.status_choices,
        **filters,
    }

    return render(request, 'core/product/product_serials.html', context)


@login_required
@permission_required('product', 'create')
def product_serial_add(request, pk):
    """Fügt eine Seriennummer zu einem Produkt hinzu."""
    product = get_object_or_404(Product, pk=pk)

    # Prüfen, ob das Produkt für Seriennummern konfiguriert ist
    if not product.has_serial_numbers:
        product.has_serial_numbers = True
        product.save()
        messages.info(request, 'Das Produkt wurde für Seriennummern aktiviert.')

    if request.method == 'POST':
        form = SerialNumberForm(request.POST, product=product)
        if form.is_valid():
            serial = form.save(commit=False)
            serial.product = product
            serial.save()

            messages.success(request, 'Seriennummer wurde erfolgreich hinzugefügt.')
            return redirect('product_serials', pk=product.pk)
    else:
        form = SerialNumberForm(product=product)

    context = {
        'product': product,
        'form': form,
    }

    return render(request, 'core/product/product_serial_form.html', context)


@login_required
@permission_required('product', 'create')
def product_serial_bulk_add(request, pk):
    """Fügt mehrere Seriennummern zu einem Produkt hinzu."""
    product = get_object_or_404(Product, pk=pk)

    # Prüfen, ob das Produkt für Seriennummern konfiguriert ist
    if not product.has_serial_numbers:
        product.has_serial_numbers = True
        product.save()
        messages.info(request, 'Das Produkt wurde für Seriennummern aktiviert.')

    if request.method == 'POST':
        form = BulkSerialNumberForm(request.POST, product=product)
        if form.is_valid():
            prefix = form.cleaned_data.get('prefix', '')
            start_number = form.cleaned_data.get('start_number')
            count = form.cleaned_data.get('count')
            digits = form.cleaned_data.get('digits')
            status = form.cleaned_data.get('status')
            warehouse = form.cleaned_data.get('warehouse')
            purchase_date = form.cleaned_data.get('purchase_date')
            expiry_date = form.cleaned_data.get('expiry_date')
            variant = form.cleaned_data.get('variant')

            # Format für Seriennummern erstellen
            serial_format = f"{prefix}{{:0{digits}d}}"

            # Prüfen, ob Seriennummern bereits existieren
            existing_numbers = []
            for i in range(start_number, start_number + count):
                serial_number = serial_format.format(i)
                if SerialNumber.objects.filter(serial_number=serial_number).exists():
                    existing_numbers.append(serial_number)

            if existing_numbers:
                messages.error(request,
                               f'Folgende Seriennummern existieren bereits: {", ".join(existing_numbers[:10])}' +
                               (f' und {len(existing_numbers) - 10} weitere' if len(existing_numbers) > 10 else ''))
                context = {
                    'product': product,
                    'form': form,
                }
                return render(request, 'core/product/product_serial_bulk_form.html', context)

            # Seriennummern erstellen
            created_count = 0
            for i in range(start_number, start_number + count):
                serial_number = serial_format.format(i)
                SerialNumber.objects.create(
                    product=product,
                    serial_number=serial_number,
                    status=status,
                    warehouse=warehouse,
                    purchase_date=purchase_date,
                    expiry_date=expiry_date,
                    variant=variant
                )
                created_count += 1

            messages.success(request, f'{created_count} Seriennummern wurden erfolgreich erstellt.')
            return redirect('product_serials', pk=product.pk)
    else:
        form = BulkSerialNumberForm(product=product)

    context = {
        'product': product,
        'form': form,
    }

    return render(request, 'core/product/product_serial_bulk_form.html', context)


@login_required
@permission_required('product', 'edit')
def product_serial_update(request, pk, serial_id):
    """Aktualisiert eine Seriennummer."""
    product = get_object_or_404(Product, pk=pk)
    serial = get_object_or_404(SerialNumber, pk=serial_id, product=product)

    if request.method == 'POST':
        form = SerialNumberForm(request.POST, instance=serial, product=product)
        if form.is_valid():
            serial = form.save()
            messages.success(request, 'Seriennummer wurde erfolgreich aktualisiert.')
            return redirect('product_serials', pk=product.pk)
    else:
        form = SerialNumberForm(instance=serial, product=product)

    context = {
        'product': product,
        'serial': serial,
        'form': form,
    }

    return render(request, 'core/product/product_serial_form.html', context)


@login_required
@permission_required('product', 'delete')
def product_serial_delete(request, pk, serial_id):
    """Löscht eine Seriennummer."""
    product = get_object_or_404(Product, pk=pk)
    serial = get_object_or_404(SerialNumber, pk=serial_id, product=product)

    if request.method == 'POST':
        serial.delete()
        messages.success(request, 'Seriennummer wurde erfolgreich gelöscht.')
        return redirect('product_serials', pk=product.pk)

    context = {
        'product': product,
        'serial': serial,
    }

    return render(request, 'core/product/product_serial_confirm_delete.html', context)


# ------------------------------------------------------------------------------
# Chargen
# ------------------------------------------------------------------------------

@login_required
@permission_required('product', 'view')
def product_batches(request, pk):
    """Zeigt alle Chargen eines Produkts an."""
    product = get_object_or_404(Product, pk=pk)

    # Filteroptionen
    warehouse_filter = request.GET.get('warehouse', '')
    variant_filter = request.GET.get('variant', '')
    search_query = request.GET.get('search', '')
    expiry_filter = request.GET.get('expiry', '')

    batches = BatchNumber.objects.filter(product=product)

    if warehouse_filter:
        batches = batches.filter(warehouse_id=warehouse_filter)

    if variant_filter:
        batches = batches.filter(variant_id=variant_filter)

    if search_query:
        batches = batches.filter(batch_number__icontains=search_query)

    # Filtern nach Verfallsdatum
    today = timezone.now().date()
    if expiry_filter == 'expired':
        batches = batches.filter(expiry_date__lt=today)
    elif expiry_filter == 'expiring_soon':
        batches = batches.filter(expiry_date__gte=today, expiry_date__lte=today + timedelta(days=30))
    elif expiry_filter == 'valid':
        batches = batches.filter(expiry_date__gt=today + timedelta(days=30))

    # Paginierung
    batches = paginate_queryset(batches, request.GET.get('page'), per_page=25)


    # Verfallsstatistiken
    expired_count = BatchNumber.objects.filter(product=product, expiry_date__lt=today).count()
    expiring_soon_count = BatchNumber.objects.filter(
        product=product,
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30)
    ).count()
    valid_count = BatchNumber.objects.filter(
        product=product,
        expiry_date__gt=today + timedelta(days=30)
    ).count()

    context = {
        'product': product,
        'batches': batches,
        'warehouses': Warehouse.objects.filter(is_active=True),
        'variants': product.variants.all(),
        'warehouse_filter': warehouse_filter,
        'variant_filter': variant_filter,
        'search_query': search_query,
        'expiry_filter': expiry_filter,
        'expired_count': expired_count,
        'expiring_soon_count': expiring_soon_count,
        'valid_count': valid_count,
        'today': today,
    }

    return render(request, 'core/product/product_batches.html', context)


@login_required
@permission_required('product', 'create')
def product_batch_add(request, pk):
    """Fügt eine Charge zu einem Produkt hinzu."""
    product = get_object_or_404(Product, pk=pk)

    # Prüfen, ob das Produkt für Chargen konfiguriert ist
    if not product.has_batch_tracking:
        product.has_batch_tracking = True
        product.save()
        messages.info(request, 'Das Produkt wurde für Chargenverfolgung aktiviert.')

    if request.method == 'POST':
        form = BatchNumberForm(request.POST, product=product)
        if form.is_valid():
            batch = form.save(commit=False)
            batch.product = product
            batch.save()

            messages.success(request, 'Charge wurde erfolgreich hinzugefügt.')
            return redirect('product_batches', pk=product.pk)
    else:
        form = BatchNumberForm(product=product)

    context = {
        'product': product,
        'form': form,
    }

    return render(request, 'core/product/product_batch_form.html', context)


@login_required
@permission_required('product', 'edit')
def product_batch_update(request, pk, batch_id):
    """Aktualisiert eine Charge."""
    product = get_object_or_404(Product, pk=pk)
    batch = get_object_or_404(BatchNumber, pk=batch_id, product=product)

    if request.method == 'POST':
        form = BatchNumberForm(request.POST, instance=batch, product=product)
        if form.is_valid():
            batch = form.save()
            messages.success(request, 'Charge wurde erfolgreich aktualisiert.')
            return redirect('product_batches', pk=product.pk)
    else:
        form = BatchNumberForm(instance=batch, product=product)

    context = {
        'product': product,
        'batch': batch,
        'form': form,
    }

    return render(request, 'core/product/product_batch_form.html', context)


@login_required
@permission_required('product', 'delete')
def product_batch_delete(request, pk, batch_id):
    """Löscht eine Charge."""
    product = get_object_or_404(Product, pk=pk)
    batch = get_object_or_404(BatchNumber, pk=batch_id, product=product)

    if request.method == 'POST':
        batch.delete()
        messages.success(request, 'Charge wurde erfolgreich gelöscht.')
        return redirect('product_batches', pk=product.pk)

    context = {
        'product': product,
        'batch': batch,
    }

    return render(request, 'core/product/product_batch_confirm_delete.html', context)

#7777
# ------------------------------------------------------------------------------
# Verfallsdaten-Verwaltung
# ------------------------------------------------------------------------------

from core.utils.filters import filter_expiring_serials, filter_expiring_batches

@login_required
@permission_required('product', 'view')
def expiry_management(request):
    """Zentrale Verwaltung aller Produkte mit Verfallsdaten."""

    today = timezone.now().date()
    days_threshold = request.GET.get('days_threshold', '30')
    try:
        days_threshold = int(days_threshold)
    except ValueError:
        days_threshold = 30

    filters = {
        'expiry': request.GET.get('filter', 'all'),
        'search': request.GET.get('search', ''),
        'category': request.GET.get('category', ''),
        'days': days_threshold,
    }

    serials = SerialNumber.objects.filter(expiry_date__isnull=False)
    batches = BatchNumber.objects.filter(expiry_date__isnull=False)

    serials = filter_expiring_serials(serials, filters, today)
    batches = filter_expiring_batches(batches, filters, today)

    # Statistiken
    serial_stats = {
        'total': SerialNumber.objects.filter(expiry_date__isnull=False).count(),
        'expired': SerialNumber.objects.filter(expiry_date__lt=today).count(),
        'expiring_soon': SerialNumber.objects.filter(
            expiry_date__gte=today,
            expiry_date__lte=today + timedelta(days=days_threshold)
        ).count(),
        'valid': SerialNumber.objects.filter(
            expiry_date__gt=today + timedelta(days=days_threshold)
        ).count(),
    }

    batch_stats = {
        'total': BatchNumber.objects.filter(expiry_date__isnull=False).count(),
        'expired': BatchNumber.objects.filter(expiry_date__lt=today).count(),
        'expiring_soon': BatchNumber.objects.filter(
            expiry_date__gte=today,
            expiry_date__lte=today + timedelta(days=days_threshold)
        ).count(),
        'valid': BatchNumber.objects.filter(
            expiry_date__gt=today + timedelta(days=days_threshold)
        ).count(),
    }

    serials = paginate_queryset(serials, request.GET.get('serials_page'), per_page=25)
    batches = paginate_queryset(batches, request.GET.get('batches_page'), per_page=25)

    context = {
        'serials': serials,
        'batches': batches,
        'serial_stats': serial_stats,
        'batch_stats': batch_stats,
        'categories': Category.objects.all(),
        'expiry_filter': filters['expiry'],
        'days_threshold': filters['days'],
        'search_query': filters['search'],
        'category_filter': filters['category'],
        'today': today,
    }

    return render(request, 'core/product/expiry_management.html', context)


# ------------------------------------------------------------------------------
# Erweiterung der bestehenden Produkt-Views
# ------------------------------------------------------------------------------

# Aktualisieren der product_detail View, um die erweiterten Funktionen anzuzeigen
@login_required
@permission_required('product', 'view')
def product_detail(request, pk):
    """Show details for a specific product."""
    product = get_object_or_404(Product, pk=pk)

    # Lieferanteninformationen
    supplier_products = SupplierProduct.objects.filter(product=product).select_related('supplier')

    # Bestandsbewegungen
    movements = StockMovement.objects.filter(product=product).select_related('created_by', 'warehouse').order_by('-created_at')[:20]

    # Primärbild abrufen, falls vorhanden
    primary_photo = product.photos.filter(is_primary=True).first()

    # Anzahl der Varianten, Seriennummern und Chargen
    variant_count = product.variants.count()
    serial_count = product.serial_numbers.count()
    batch_count = product.batches.count()

    # Lager, auf die der Benutzer Zugriff hat
    if request.user.is_superuser:
        accessible_warehouses = Warehouse.objects.filter(is_active=True)
    else:
        accessible_warehouses = [w for w in Warehouse.objects.filter(is_active=True)
                                 if WarehouseAccess.has_access(request.user, w, 'view')]

    # Varianten mit kritischem Bestand
    # Hier holen wir zuerst alle Varianten
    all_variants = product.variants.all()
    
    # Spezieller Schwellwert für "kritischen Bestand"
    low_stock_threshold = 3
    
    # Varianten mit kritischem Bestand
    low_stock_variants = []
    
    # Für jede Variante prüfen, ob der Gesamtbestand niedrig ist
    for variant in all_variants:
        # Versuchen, das total_stock über VariantWarehouse zu berechnen
        try:
            from django.db.models import Sum
            from inventory.models import VariantWarehouse
            
            total_stock = VariantWarehouse.objects.filter(
                variant=variant,
                warehouse__in=accessible_warehouses
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            # Wenn es niedrig ist, zur Liste hinzufügen
            if total_stock < low_stock_threshold and total_stock > 0:
                # Wir fügen das berechnete total_stock als Attribut hinzu
                variant.total_stock = total_stock
                low_stock_variants.append(variant)
                
                # Nur max. 5 Varianten in die Liste
                if len(low_stock_variants) >= 5:
                    break
        except (ImportError, AttributeError):
            # Wenn es keine VariantWarehouse-Tabelle gibt, leere Liste verwenden
            pass

    # Ablaufende Seriennummern und Chargen
    today = timezone.now().date()
    expiring_serials = product.serial_numbers.filter(
        expiry_date__isnull=False,
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30)
    ).order_by('expiry_date')[:5]

    expiring_batches = product.batches.filter(
        expiry_date__isnull=False,
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30)
    ).order_by('expiry_date')[:5]

    # Seriennummern nach Status gruppieren
    serial_status_counts = {}
    if product.has_serial_numbers:
        status_choices = SerialNumber.status_choices
        for status_code, status_name in status_choices:
            serial_status_counts[status_code] = product.serial_numbers.filter(status=status_code).count()
    
    # NEU: Lagerbestandsinformationen
    # Bestand in allen zugänglichen Lagern für dieses Produkt
    warehouse_stocks = ProductWarehouse.objects.filter(
        product=product,
        warehouse__in=accessible_warehouses
    ).select_related('warehouse').order_by('warehouse__name')
    
    # Gesamtbestand über alle zugänglichen Lager
    total_accessible_stock = warehouse_stocks.aggregate(total=models.Sum('quantity'))['total'] or 0

    context = {
        'product': product,
        'supplier_products': supplier_products,
        'movements': movements,
        'primary_photo': primary_photo,
        'variant_count': variant_count,
        'serial_count': serial_count,
        'batch_count': batch_count,
        'low_stock_variants': low_stock_variants,
        'expiring_serials': expiring_serials,
        'expiring_batches': expiring_batches,
        'serial_status_counts': serial_status_counts,
        'warehouse_stocks': warehouse_stocks,
        'total_accessible_stock': total_accessible_stock,
        'accessible_warehouses': accessible_warehouses,
    }

    return render(request, 'core/product_detail.html', context)


# ------------------------------------------------------------------------------
# Globale Seriennummern-Verwaltung
# ------------------------------------------------------------------------------

@login_required
@permission_required('serialnumber', 'view')
def serialnumber_list(request):
    """Zeigt alle Seriennummern im System an."""
    # Filteroptionen
    status_filter = request.GET.get('status', '')
    warehouse_filter = request.GET.get('warehouse', '')
    product_filter = request.GET.get('product', '')
    search_query = request.GET.get('search', '')

    filters = {
        'status': request.GET.get('status', ''),
        'warehouse': request.GET.get('warehouse', ''),
        'product': request.GET.get('product', ''),
        'search': request.GET.get('search', ''),
    }

    serials = SerialNumber.objects.select_related('product', 'warehouse', 'variant')
    serials = filter_product_serials(serials, filters)

    # Sortierung
    sort_by = request.GET.get('sort', '-created_at')
    serials = serials.order_by(sort_by)

    # Paginierung
    serials = paginate_queryset(serials, request.GET.get('page'), per_page=50)

    # Status-Statistiken
    status_counts = SerialNumber.objects.values('status').annotate(
        count=Count('status')).order_by('status')

    # In Dictionary umwandeln für leichteren Zugriff
    status_stats = {item['status']: item['count'] for item in status_counts}

    context = {
        'serials': serials,
        'status_stats': status_stats,
        'warehouses': Warehouse.objects.filter(is_active=True),
        'products': Product.objects.filter(has_serial_numbers=True),
        'status_choices': SerialNumber.status_choices,
        'status_filter': status_filter,
        'warehouse_filter': warehouse_filter,
        'product_filter': product_filter,
        'search_query': search_query,
        'sort_by': sort_by,
    }

    return render(request, 'core/serialnumber/serialnumber_list.html', context)


@login_required
@permission_required('serialnumber', 'view')
def serialnumber_detail(request, serial_id):
    """Zeigt Details zu einer Seriennummer an."""
    serial = get_object_or_404(SerialNumber.objects.select_related(
        'product', 'warehouse', 'variant'), pk=serial_id)

    # Hier könnten später Bewegungshistorie, Wartungsprotokolle etc. geladen werden

    context = {
        'serial': serial,
    }

    return render(request, 'core/serialnumber/serialnumber_detail.html', context)


@login_required
@permission_required('serialnumber', 'view')
def serialnumber_history(request, serial_id):
    """Zeigt die Historie einer Seriennummer an."""
    serial = get_object_or_404(SerialNumber.objects.select_related(
        'product', 'warehouse', 'variant'), pk=serial_id)

    # Hier würde man die Historiendaten laden
    # Beispiel: movements = SerialNumberMovement.objects.filter(serial_number=serial).order_by('-timestamp')

    context = {
        'serial': serial,
        'movements': [],  # Platzhalter für die tatsächlichen Bewegungsdaten
    }

    return render(request, 'core/serialnumber/serialnumber_history.html', context)


@login_required
@permission_required('serialnumber', 'add')
def serialnumber_add(request):
    """Fügt eine neue Seriennummer hinzu (produktunabhängig)."""
    if request.method == 'POST':
        form = SerialNumberForm(request.POST)
        if form.is_valid():
            serial = form.save()

            # Produkt für Seriennummern aktivieren, falls noch nicht geschehen
            product = serial.product
            if not product.has_serial_numbers:
                product.has_serial_numbers = True
                product.save()

            messages.success(request, 'Seriennummer wurde erfolgreich hinzugefügt.')
            return redirect('serialnumber_list')
    else:
        form = SerialNumberForm()

    context = {
        'form': form,
    }

    return render(request, 'core/serialnumber/serialnumber_form.html', context)


@login_required
@permission_required('serialnumber', 'transfer')
def serialnumber_transfer(request):
    """Transferiert eine Seriennummer von einem Lager zu einem anderen."""
    if request.method == 'POST':
        serial_number = request.POST.get('serial_number')
        target_warehouse_id = request.POST.get('target_warehouse')

        try:
            serial = SerialNumber.objects.get(serial_number=serial_number)
            old_warehouse = serial.warehouse
            new_warehouse = Warehouse.objects.get(pk=target_warehouse_id)

            # Prüfe Benutzerberechtigungen für beide Lager
            if not request.user.is_superuser:
                if (old_warehouse and not WarehouseAccess.has_access(request.user, old_warehouse)) or \
                        not WarehouseAccess.has_access(request.user, new_warehouse):
                    messages.error(request, 'Sie haben keine Berechtigung für eines der Lager.')
                    return redirect('serialnumber_transfer')

            # Formatierte Meldung erstellen, je nachdem ob ein altes Lager vorhanden war
            if old_warehouse:
                success_message = f'Seriennummer {serial_number} wurde erfolgreich von ' \
                                  f'{old_warehouse.name} nach {new_warehouse.name} transferiert.'
            else:
                success_message = f'Seriennummer {serial_number} wurde erfolgreich ' \
                                  f'dem Lager {new_warehouse.name} zugewiesen.'

            # Speichere den Transfer
            serial.warehouse = new_warehouse
            serial.last_modified = timezone.now()
            serial.save()

            # Hier könnte man auch einen Eintrag in einer Transfer-Historie speichern

            messages.success(request, success_message)
            return redirect('serialnumber_list')

        except SerialNumber.DoesNotExist:
            messages.error(request, f'Seriennummer {serial_number} wurde nicht gefunden.')
        except Warehouse.DoesNotExist:
            messages.error(request, 'Das Ziellager existiert nicht.')
        except Exception as e:
            messages.error(request, f'Fehler beim Transferieren: {str(e)}')

    # Für GET-Anfragen oder bei Fehlern im POST
    warehouses = Warehouse.objects.filter(is_active=True)

    # Wenn eine Seriennummer als Parameter übergeben wurde (z.B. aus der Detailansicht)
    initial_serial = request.GET.get('serial', '')

    context = {
        'warehouses': warehouses,
        'initial_serial': initial_serial,
    }

    return render(request, 'core/serialnumber/serialnumber_transfer.html', context)


@login_required
@permission_required('serialnumber', 'import')
def serialnumber_import(request):
    return handle_csv_import(
        form_class=SerialNumberImportForm,
        importer_class=SerialNumberImporter,
        request=request,
        template_name='core/serialnumber/serialnumber_import.html',
        success_redirect='serialnumber_import_log_detail',  # oder z. B. 'serialnumber_list' falls kein DetailView
        extra_context={
            'title': 'Seriennummern importieren',
            'description': 'Importieren Sie Seriennummern aus einer CSV-Datei.',
            'expected_format': 'product_sku,serial_number,status,warehouse_name,purchase_date,expiry_date',
            'example': 'P1001,SN12345,in_stock,Hauptlager,2023-01-01,2025-01-01',
            'required_columns': ['product_sku', 'serial_number'],
            'optional_columns': ['status', 'warehouse_name', 'purchase_date', 'expiry_date', 'notes'],
        }
    )


@login_required
@permission_required('serialnumber', 'export')
def serialnumber_export(request):
    """Exportiert Seriennummern in eine CSV- oder Excel-Datei."""
    # Filteroptionen für den Export
    status_filter = request.GET.get('status', '')
    warehouse_filter = request.GET.get('warehouse', '')
    product_filter = request.GET.get('product', '')

    filters = {
        'status': request.GET.get('status'),
        'warehouse': request.GET.get('warehouse'),
        'product': request.GET.get('product'),
    }

    serials = SerialNumber.objects.select_related('product', 'warehouse', 'variant')
    serials = filter_product_serials(serials, filters)

    # Export-Format bestimmen
    export_format = request.GET.get('format', '')

    if export_format == 'csv':
        # CSV-Export
        response = HttpResponse(content_type='text/csv')
        response[
            'Content-Disposition'] = f'attachment; filename="seriennummern_{timezone.now().strftime("%Y%m%d")}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Seriennummer', 'Produkt', 'Variante', 'Status', 'Lager',
            'Kaufdatum', 'Ablaufdatum', 'Notizen', 'Erstellt am'
        ])

        for serial in serials:
            writer.writerow([
                serial.serial_number,
                serial.product.name,
                serial.variant.name if serial.variant else '',
                serial.get_status_display(),
                serial.warehouse.name if serial.warehouse else '',
                serial.purchase_date.strftime('%Y-%m-%d') if serial.purchase_date else '',
                serial.expiry_date.strftime('%Y-%m-%d') if serial.expiry_date else '',
                serial.notes,
                serial.created_at.strftime('%Y-%m-%d %H:%M')
            ])

        return response

    elif export_format == 'excel':
        # Excel-Export würde hier implementiert werden
        # Ähnlich wie bei der export_import_logs-Funktion
        messages.error(request, 'Excel-Export ist derzeit nicht verfügbar.')
        return redirect('serialnumber_list')

    # Wenn kein Export angefordert wurde, zeige Exportformular
    context = {
        'warehouses': Warehouse.objects.filter(is_active=True),
        'products': Product.objects.filter(has_serial_numbers=True),
        'status_choices': SerialNumber.status_choices,
    }

    return render(request, 'core/serialnumber/serialnumber_export.html', context)


@login_required
@permission_required('serialnumber', 'view')
def serialnumber_scan(request):
    """Scannt eine Seriennummer (z.B. mit Barcode-Scanner) und zeigt Details an."""
    scanned_number = request.GET.get('scan', '')
    found_serial = None

    # Start with an empty list of recent scans
    recent_scans = []

    if scanned_number:
        try:
            # Versuche, die Seriennummer zu finden mit explizitem select_related für warehouse
            found_serial = SerialNumber.objects.select_related('product', 'warehouse').get(serial_number=scanned_number)

            # Hole die alten Suchanfragen
            recent_searches = request.session.get('recent_serial_searches', [])

            # Aktuelles Datum formatiert (deutsches Format)
            current_time = timezone.now()
            formatted_time = current_time.strftime('%d.%m.%Y %H:%M')

            # Warehouse bestimmen
            warehouse_name = 'Nicht zugewiesen'
            if hasattr(found_serial, 'warehouse') and found_serial.warehouse:
                warehouse_name = found_serial.warehouse.name

            # Create a new search entry
            search_info = {
                'id': found_serial.id,
                'serial_number': found_serial.serial_number,
                'product_name': found_serial.product.name,
                'timestamp': formatted_time,
                'warehouse_name': warehouse_name
            }

            # Remove duplicates
            filtered_searches = []
            for old_search in recent_searches:
                if old_search.get('id') != found_serial.id:
                    filtered_searches.append(old_search)

            # Add the new search at the beginning
            filtered_searches.insert(0, search_info)

            # Limit to 5 entries
            filtered_searches = filtered_searches[:5]

            # Save to session
            request.session['recent_serial_searches'] = filtered_searches

            # Redirect to detail page
            return redirect('serialnumber_detail', serial_id=found_serial.pk)
        except SerialNumber.DoesNotExist:
            messages.error(request, f'Seriennummer {scanned_number} wurde nicht gefunden.')
        except Exception as e:
            messages.error(request, f'Fehler bei der Suche: {str(e)}')

    # Get recent searches from session
    session_searches = request.session.get('recent_serial_searches', [])

    # Format for template
    for item in session_searches:
        recent_scans.append({
            'id': item.get('id', 0),
            'serial_number': item.get('serial_number', ''),
            'product': {'name': item.get('product_name', '')},
            'timestamp': item.get('timestamp', ''),
            'warehouse': {'name': item.get('warehouse_name', 'Nicht zugewiesen')}
        })

    context = {
        'scanned_number': scanned_number,
        'recent_scans': recent_scans,
    }

    return render(request, 'core/serialnumber/serialnumber_scan.html', context)


@login_required
@permission_required('serialnumber', 'view')
def serialnumber_search(request):
    """Erweiterte Suchfunktion für Seriennummern."""
    search_query = request.GET.get('q', '')

    results = []
    if search_query:
        # Suche nach Seriennummer, Produktname, verschiedenen Feldern
        results = SerialNumber.objects.filter(
            Q(serial_number__icontains=search_query) |
            Q(product__name__icontains=search_query) |
            Q(product__sku__icontains=search_query) |
            Q(product__barcode__icontains=search_query) |
            Q(variant__name__icontains=search_query) |
            Q(notes__icontains=search_query)
        ).select_related('product', 'warehouse', 'variant')

    context = {
        'search_query': search_query,
        'results': results,
    }

    return render(request, 'core/serialnumber/serialnumber_search.html', context)


@login_required
@permission_required('serialnumber', 'edit')
def serialnumber_batch_actions(request):
    """Massenaktionen für mehrere Seriennummern gleichzeitig."""
    if request.method == 'POST':
        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_serials')

        if not selected_ids:
            messages.error(request, 'Keine Seriennummern ausgewählt.')
            return redirect('serialnumber_list')

        # Seriennummern abrufen
        serials = SerialNumber.objects.filter(pk__in=selected_ids)

        if action == 'change_status':
            new_status = request.POST.get('new_status')
            if new_status:
                serials.update(status=new_status)
                messages.success(request,
                                 f'{len(selected_ids)} Seriennummern auf Status "{dict(SerialNumber.status_choices)[new_status]}" gesetzt.')

        elif action == 'change_warehouse':
            new_warehouse_id = request.POST.get('new_warehouse')
            if new_warehouse_id:
                try:
                    warehouse = Warehouse.objects.get(pk=new_warehouse_id)
                    serials.update(warehouse=warehouse)
                    messages.success(request, f'{len(selected_ids)} Seriennummern nach "{warehouse.name}" verschoben.')
                except Warehouse.DoesNotExist:
                    messages.error(request, 'Das ausgewählte Lager existiert nicht.')

        elif action == 'delete':
            count = serials.count()
            serials.delete()
            messages.success(request, f'{count} Seriennummern wurden gelöscht.')

        return redirect('serialnumber_list')

    # Standardwerte für GET-Anfrage
    context = {
        'status_choices': SerialNumber.status_choices,
        'warehouses': Warehouse.objects.filter(is_active=True),
    }

    return render(request, 'core/serialnumber/serialnumber_batch_actions.html', context)


@login_required
@permission_required('serialnumber', 'import')
def import_serialnumbers(request):
    """Importiert Seriennummern aus einer CSV-Datei."""
    if request.method == 'POST':
        form = request.POST
        csv_file = request.FILES.get('csv_file')

        if not csv_file:
            messages.error(request, 'Bitte wählen Sie eine CSV-Datei aus.')
            return redirect('import_serialnumbers')

        # Überprüfen des Dateityps
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Bitte laden Sie eine CSV-Datei hoch.')
            return redirect('import_serialnumbers')

        # Datei in Speicher lesen
        file_data = csv_file.read().decode('utf-8')

        # Import-Log erstellen
        import_log = ImportLog.objects.create(
            file_name=csv_file.name,
            import_type='serialnumbers',
            started_by=request.user,
            status='processing'
        )

        try:
            lines = file_data.split('\n')
            # CSV-Header überprüfen
            header = lines[0].strip().split(',')
            required_headers = ['product_sku', 'serial_number']
            optional_headers = ['status', 'warehouse_name', 'purchase_date', 'expiry_date', 'variant_name', 'notes']

            # Prüfen, ob alle erforderlichen Header vorhanden sind
            if not all(column in header for column in required_headers):
                import_log.status = 'failed'
                import_log.notes = f'Ungültiger CSV-Header. Erforderliche Felder: {", ".join(required_headers)}'
                import_log.save()
                messages.error(request, import_log.notes)
                return redirect('import_serialnumbers')

            # Indizes für die Spalten ermitteln
            product_sku_idx = header.index('product_sku')
            serial_number_idx = header.index('serial_number')

            # Optionale Felder
            status_idx = header.index('status') if 'status' in header else None
            warehouse_name_idx = header.index('warehouse_name') if 'warehouse_name' in header else None
            purchase_date_idx = header.index('purchase_date') if 'purchase_date' in header else None
            expiry_date_idx = header.index('expiry_date') if 'expiry_date' in header else None
            variant_name_idx = header.index('variant_name') if 'variant_name' in header else None
            notes_idx = header.index('notes') if 'notes' in header else None

            # Import starten
            created_count = 0
            updated_count = 0
            error_count = 0
            error_msgs = []

            # Update-Modus ermitteln
            update_existing = form.get('update_existing', 'False') == 'True'
            default_status = form.get('default_status', 'in_stock')
            default_warehouse_id = form.get('default_warehouse', None)

            # Standard-Lager ermitteln, falls angegeben
            default_warehouse = None
            if default_warehouse_id:
                try:
                    default_warehouse = Warehouse.objects.get(pk=default_warehouse_id)
                except Warehouse.DoesNotExist:
                    pass

            # CSV-Zeilen verarbeiten (ohne Header)
            for i, line in enumerate(lines[1:], 1):
                if not line.strip():  # Leere Zeile überspringen
                    continue

                try:
                    # Zeile parsen
                    fields = line.strip().split(',')

                    product_sku = fields[product_sku_idx].strip()
                    serial_number = fields[serial_number_idx].strip()

                    # Produkt finden
                    try:
                        product = Product.objects.get(sku=product_sku)
                    except Product.DoesNotExist:
                        error_count += 1
                        error_msg = f'Fehler in Zeile {i + 1}: Produkt mit SKU "{product_sku}" nicht gefunden.'
                        error_msgs.append(error_msg)
                        continue

                    # Prüfen, ob Seriennummer bereits existiert
                    existing_serial = SerialNumber.objects.filter(serial_number=serial_number).first()
                    if existing_serial and not update_existing:
                        error_count += 1
                        error_msg = f'Fehler in Zeile {i + 1}: Seriennummer "{serial_number}" existiert bereits.'
                        error_msgs.append(error_msg)
                        continue

                    # Optionale Felder extrahieren
                    status = fields[status_idx].strip() if status_idx is not None and status_idx < len(
                        fields) else default_status

                    # Lager finden, falls angegeben
                    warehouse = default_warehouse
                    if warehouse_name_idx is not None and warehouse_name_idx < len(fields) and fields[
                        warehouse_name_idx].strip():
                        warehouse_name = fields[warehouse_name_idx].strip()
                        try:
                            warehouse = Warehouse.objects.get(name=warehouse_name)
                        except Warehouse.DoesNotExist:
                            error_count += 1
                            error_msg = f'Fehler in Zeile {i + 1}: Lager "{warehouse_name}" nicht gefunden.'
                            error_msgs.append(error_msg)
                            continue

                    # Variante finden, falls angegeben
                    variant = None
                    if variant_name_idx is not None and variant_name_idx < len(fields) and fields[
                        variant_name_idx].strip():
                        variant_name = fields[variant_name_idx].strip()
                        try:
                            variant = ProductVariant.objects.get(parent_product=product, name=variant_name)
                        except ProductVariant.DoesNotExist:
                            error_count += 1
                            error_msg = f'Fehler in Zeile {i + 1}: Variante "{variant_name}" für Produkt nicht gefunden.'
                            error_msgs.append(error_msg)
                            continue

                    # Datum parsen, falls angegeben
                    purchase_date = None
                    if purchase_date_idx is not None and purchase_date_idx < len(fields) and fields[
                        purchase_date_idx].strip():
                        try:
                            purchase_date_str = fields[purchase_date_idx].strip()
                            purchase_date = datetime.strptime(purchase_date_str, '%Y-%m-%d').date()
                        except ValueError:
                            error_count += 1
                            error_msg = f'Fehler in Zeile {i + 1}: Ungültiges Kaufdatum "{fields[purchase_date_idx]}".'
                            error_msgs.append(error_msg)
                            continue

                    expiry_date = None
                    if expiry_date_idx is not None and expiry_date_idx < len(fields) and fields[
                        expiry_date_idx].strip():
                        try:
                            expiry_date_str = fields[expiry_date_idx].strip()
                            expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
                        except ValueError:
                            error_count += 1
                            error_msg = f'Fehler in Zeile {i + 1}: Ungültiges Ablaufdatum "{fields[expiry_date_idx]}".'
                            error_msgs.append(error_msg)
                            continue

                    # Notizen extrahieren
                    notes = fields[notes_idx].strip() if notes_idx is not None and notes_idx < len(fields) else ''

                    # Seriennummer erstellen oder aktualisieren
                    if existing_serial and update_existing:
                        # Aktualisieren
                        existing_serial.product = product
                        existing_serial.status = status
                        existing_serial.warehouse = warehouse
                        existing_serial.variant = variant
                        existing_serial.purchase_date = purchase_date
                        existing_serial.expiry_date = expiry_date
                        existing_serial.notes = notes
                        existing_serial.save()
                        updated_count += 1
                    else:
                        # Neu erstellen
                        SerialNumber.objects.create(
                            product=product,
                            serial_number=serial_number,
                            status=status,
                            warehouse=warehouse,
                            variant=variant,
                            purchase_date=purchase_date,
                            expiry_date=expiry_date,
                            notes=notes
                        )
                        created_count += 1

                    # Produkt für Seriennummern aktivieren, falls noch nicht geschehen
                    if not product.has_serial_numbers:
                        product.has_serial_numbers = True
                        product.save()

                except Exception as e:
                    error_count += 1
                    error_msg = f'Fehler in Zeile {i + 1}: {str(e)}'
                    error_msgs.append(error_msg)

            # Import-Log aktualisieren
            import_log.status = 'completed' if error_count == 0 else 'completed_with_errors'
            import_log.rows_processed = created_count + updated_count + error_count
            import_log.rows_created = created_count
            import_log.rows_updated = updated_count
            import_log.rows_error = error_count
            import_log.notes = f"Import abgeschlossen: {created_count} erstellt, {updated_count} aktualisiert, {error_count} Fehler"

            if error_msgs:
                import_log.error_details = '\n'.join(error_msgs)

            import_log.save()

            if error_count > 0:
                messages.warning(request,
                                 f'Import mit Fehlern abgeschlossen. {created_count} erstellt, {updated_count} aktualisiert, {error_count} Fehler.')
            else:
                messages.success(request,
                                 f'Import erfolgreich abgeschlossen. {created_count} erstellt, {updated_count} aktualisiert.')

            return redirect('import_log_detail', pk=import_log.pk)

        except Exception as e:
            import_log.status = 'failed'
            import_log.notes = f'Fehler beim Import: {str(e)}'
            import_log.save()
            messages.error(request, import_log.notes)
            return redirect('import_serialnumbers')

    # Für GET-Anfragen oder bei Fehlern im POST
    context = {
        'title': 'Seriennummern importieren',
        'description': 'Importieren Sie Seriennummern aus einer CSV-Datei.',
        'expected_format': 'product_sku,serial_number,status,warehouse_name,purchase_date,expiry_date,variant_name,notes',
        'example': 'P1001,SN12345,in_stock,Hauptlager,2023-01-01,2025-01-01,Standard,Neue Lieferung',
        'required_columns': ['product_sku', 'serial_number'],
        'optional_columns': ['status', 'warehouse_name', 'purchase_date', 'expiry_date', 'variant_name', 'notes'],
        'warehouses': Warehouse.objects.filter(is_active=True),
        'status_choices': SerialNumber.status_choices,
    }

    return render(request, 'core/serialnumber/serialnumber_import.html', context)


@login_required
@permission_required('core', 'view')
def currency_list(request):
    """List all currencies."""
    currencies = Currency.objects.all().order_by('code')

    context = {
        'currencies': currencies,
    }

    return render(request, 'core/currency/currency_list.html', context)


@login_required
@permission_required('core', 'create')
def currency_create(request):
    """Create a new currency."""
    if request.method == 'POST':
        form = CurrencyForm(request.POST)
        if form.is_valid():
            currency = form.save()
            messages.success(request, f'Währung "{currency.name}" wurde erfolgreich erstellt.')
            return redirect('currency_list')
    else:
        form = CurrencyForm()

    context = {
        'form': form,
        'title': 'Neue Währung erstellen',
    }

    return render(request, 'core/currency/currency_form.html', context)


@login_required
@permission_required('core', 'edit')
def currency_update(request, pk):
    """Update an existing currency."""
    currency = get_object_or_404(Currency, pk=pk)

    if request.method == 'POST':
        form = CurrencyForm(request.POST, instance=currency)
        if form.is_valid():
            currency = form.save()
            messages.success(request, f'Währung "{currency.name}" wurde erfolgreich aktualisiert.')
            return redirect('currency_list')
    else:
        form = CurrencyForm(instance=currency)

    context = {
        'form': form,
        'currency': currency,
        'title': f'Währung "{currency.name}" bearbeiten',
    }

    return render(request, 'core/currency/currency_form.html', context)


@login_required
@permission_required('core', 'delete')
def currency_delete(request, pk):
    """Delete a currency."""
    currency = get_object_or_404(Currency, pk=pk)

    # Check if this is the default currency
    if currency.is_default:
        messages.error(request, 'Die Standardwährung kann nicht gelöscht werden.')
        return redirect('currency_list')

    # Check if the currency is used in the system
    # This would need to check any foreign key relationships
    # Example: product_count = Product.objects.filter(currency=currency).count()
    # For now, we'll just check if it's in use or not
    is_used = False  # Replace with actual check

    if request.method == 'POST':
        try:
            currency.delete()
            messages.success(request, f'Währung "{currency.name}" wurde erfolgreich gelöscht.')
        except Exception as e:
            messages.error(request, f'Fehler beim Löschen der Währung: {str(e)}')

        return redirect('currency_list')

    context = {
        'currency': currency,
        'is_used': is_used,
    }

    return render(request, 'core/currency/currency_confirm_delete.html', context)


@login_required
def products_search(request):
    """API-Endpunkt für die Suche nach Produkten mit Bestandsinformationen und bevorzugtem Lieferanten."""
    search_query = request.GET.get('q', '')

    if not search_query or len(search_query) < 2:
        return JsonResponse([], safe=False)

    # Produkte suchen
    products = Product.objects.filter(
        Q(name__icontains=search_query) |
        Q(sku__icontains=search_query) |
        Q(barcode__icontains=search_query)
    ).prefetch_related('supplier_products', 'productwarehouse_set')

    # Maximal 50 Ergebnisse zurückgeben
    products = products[:50]

    results = []
    for product in products:
        # Aktuellen Bestand berechnen
        stock = product.productwarehouse_set.aggregate(total=Sum('quantity'))['total'] or 0

        # Bevorzugten Lieferanten ermitteln
        preferred_supplier = None
        try:
            supplier_product = SupplierProduct.objects.filter(
                product=product,
                is_preferred=True
            ).select_related('supplier').first()

            if supplier_product:
                preferred_supplier = {
                    'id': supplier_product.supplier.id,
                    'name': supplier_product.supplier.name
                }
            else:
                # Fallback: Den ersten verfügbaren Lieferanten verwenden
                supplier_product = SupplierProduct.objects.filter(
                    product=product
                ).select_related('supplier').first()

                if supplier_product:
                    preferred_supplier = {
                        'id': supplier_product.supplier.id,
                        'name': supplier_product.supplier.name
                    }
        except:
            pass

        # Produktdaten zusammenstellen
        product_data = {
            'id': product.id,
            'name': product.name,
            'sku': product.sku,
            'stock': float(stock),
            'minimum_stock': float(product.minimum_stock),
            'unit': product.unit,
            'preferred_supplier': preferred_supplier
        }

        results.append(product_data)

    return JsonResponse(results, safe=False)


@login_required
@permission_required('product', 'view')
def batch_number_list(request):
    """Liste aller Chargen im System."""
    # Filteroptionen
    filters = {
        'warehouse': request.GET.get('warehouse', ''),
        'product': request.GET.get('product', ''),
        'search': request.GET.get('search', ''),
        'expiry': request.GET.get('expiry', ''),
    }

    # Basisdaten
    batches = BatchNumber.objects.select_related('product', 'warehouse', 'variant')

    # Filter anwenden mit den Utility-Funktionen
    batches = apply_exact_filter(batches, 'warehouse_id', filters['warehouse'])
    batches = apply_exact_filter(batches, 'product_id', filters['product'])

    if filters['search']:
        batches = batches.filter(
            Q(batch_number__icontains=filters['search']) |
            Q(product__name__icontains=filters['search']) |
            Q(product__sku__icontains=filters['search'])
        )

    # Filtern nach Verfallsdatum mit der Utility-Funktion
    today = timezone.now().date()
    batches = apply_expiry_filter(batches, filters['expiry'], today, 30)

    # Sortierung
    sort_by = request.GET.get('sort', '-created_at')
    batches = batches.order_by(sort_by)

    # Paginierung mit der Utility-Funktion
    batches = paginate_queryset(batches, request.GET.get('page'), per_page=25)

    # Statistiken für die Verfallsdaten
    expired_count = BatchNumber.objects.filter(expiry_date__lt=today).count()
    expiring_soon_count = BatchNumber.objects.filter(
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30)
    ).count()
    valid_count = BatchNumber.objects.filter(
        expiry_date__gt=today + timedelta(days=30)
    ).count()
    total_count = BatchNumber.objects.count()

    context = {
        'batches': batches,
        'warehouses': Warehouse.objects.filter(is_active=True),
        'products': Product.objects.filter(has_batch_tracking=True),
        'warehouse_filter': filters['warehouse'],
        'product_filter': filters['product'],
        'search_query': filters['search'],
        'expiry_filter': filters['expiry'],
        'sort_by': sort_by,
        'expired_count': expired_count,
        'expiring_soon_count': expiring_soon_count,
        'valid_count': valid_count,
        'total_count': total_count,
        'today': today,
    }

    return render(request, 'core/batch/batch_number_list.html', context)


@login_required
@permission_required('product', 'view')
def batch_number_detail(request, batch_id):
    """Detailansicht einer Charge."""
    batch = get_object_or_404(BatchNumber.objects.select_related(
        'product', 'warehouse', 'variant', 'supplier'), pk=batch_id)

    context = {
        'batch': batch,
    }

    return render(request, 'core/batch/batch_number_detail.html', context)


@login_required
@permission_required('product', 'create')
def batch_number_add(request):
    """Fügt eine neue Charge hinzu (produktunabhängig)."""

    # Nutze die handle_form_view Utility-Funktion mit einem custom Post-Save-Hook
    def post_save_hook(batch, _):
        # Produkt für Chargenverfolgung aktivieren, falls noch nicht geschehen
        product = batch.product
        if not product.has_batch_tracking:
            product.has_batch_tracking = True
            product.save()
        messages.success(request, 'Charge wurde erfolgreich hinzugefügt.')

    return handle_form_view(
        request=request,
        form_class=BatchNumberForm,
        template='core/batch/batch_number_form.html',
        redirect_url='batch_number_list',
        context_extra={'title': 'Neue Charge anlegen'},
        success_message='Charge wurde erfolgreich erstellt.',
        post_save_hook=post_save_hook
    )


@login_required
@permission_required('product', 'edit')
def batch_number_edit(request, batch_id):
    """Bearbeitet eine bestehende Charge."""
    batch = get_object_or_404(BatchNumber, pk=batch_id)

    # Verwende die handle_model_update Utility-Funktion
    context = handle_model_update(
        request=request,
        instance=batch,
        form_class=BatchNumberForm,
        model_name_singular='Charge',
        success_redirect='batch_number_detail',
        context_extra={
            'title': f'Charge {batch.batch_number} bearbeiten',
            'batch': batch
        }
    )

    if isinstance(context, dict):
        # Wenn die Funktion ein Context-Dict zurückgibt, rendern wir das Template
        return render(request, 'core/batch/batch_number_form.html', context)
    else:
        # Ansonsten ist es eine Redirect-Response
        return context


@login_required
@permission_required('product', 'delete')
def batch_number_delete(request, batch_id):
    """Löscht eine Charge."""
    batch = get_object_or_404(BatchNumber, pk=batch_id)
    product = batch.product  # Produkt speichern, um später zu prüfen

    # Verwende die handle_model_delete Utility-Funktion
    if request.method == 'POST':
        context = handle_model_delete(
            request=request,
            instance=batch,
            model_name_singular='Charge',
            success_redirect='batch_number_list'
        )

        # Wenn keine Chargen mehr übrig sind, has_batch_tracking deaktivieren
        if not product.batches.exists():
            product.has_batch_tracking = False
            product.save()

        return redirect('batch_number_list')

    context = {
        'batch': batch,
    }

    return render(request, 'core/batch/batch_number_confirm_delete.html', context)


@login_required
@permission_required('product', 'view')
def batch_number_scan(request):
    """Scannt eine Chargennummer und zeigt Details an."""
    scanned_number = request.GET.get('scan', '')
    found_batch = None

    # Start with an empty list of recent scans
    recent_scans = []

    if scanned_number:
        try:
            # Versuche, die Chargennummer zu finden
            found_batch = BatchNumber.objects.select_related('product', 'warehouse').get(batch_number=scanned_number)

            # Hole die alten Suchanfragen
            recent_searches = request.session.get('recent_batch_searches', [])

            # Aktuelles Datum formatiert (deutsches Format)
            current_time = timezone.now()
            formatted_time = current_time.strftime('%d.%m.%Y %H:%M')

            # Warehouse bestimmen
            warehouse_name = 'Nicht zugewiesen'
            if hasattr(found_batch, 'warehouse') and found_batch.warehouse:
                warehouse_name = found_batch.warehouse.name

            # Create a new search entry
            search_info = {
                'id': found_batch.id,
                'batch_number': found_batch.batch_number,
                'product_name': found_batch.product.name,
                'timestamp': formatted_time,
                'warehouse_name': warehouse_name,
                'quantity': str(found_batch.quantity),
            }

            # Remove duplicates
            filtered_searches = []
            for old_search in recent_searches:
                if old_search.get('id') != found_batch.id:
                    filtered_searches.append(old_search)

            # Add the new search at the beginning
            filtered_searches.insert(0, search_info)

            # Limit to 5 entries
            filtered_searches = filtered_searches[:5]

            # Save to session
            request.session['recent_batch_searches'] = filtered_searches

            # Redirect to detail page
            return redirect('batch_number_detail', batch_id=found_batch.pk)
        except BatchNumber.DoesNotExist:
            messages.error(request, f'Charge {scanned_number} wurde nicht gefunden.')
        except Exception as e:
            messages.error(request, f'Fehler bei der Suche: {str(e)}')

    # Get recent searches from session
    session_searches = request.session.get('recent_batch_searches', [])

    # Format for template
    for item in session_searches:
        recent_scans.append({
            'id': item.get('id', 0),
            'batch_number': item.get('batch_number', ''),
            'product': {'name': item.get('product_name', '')},
            'timestamp': item.get('timestamp', ''),
            'warehouse': {'name': item.get('warehouse_name', 'Nicht zugewiesen')},
            'quantity': item.get('quantity', '0'),
        })

    context = {
        'scanned_number': scanned_number,
        'recent_scans': recent_scans,
    }

    return render(request, 'core/batch/batch_number_scan.html', context)


@login_required
@permission_required('product', 'edit')
def batch_number_transfer(request):
    """Transferiert eine Charge von einem Lager zu einem anderen."""
    if request.method == 'POST':
        batch_number = request.POST.get('batch_number')
        target_warehouse_id = request.POST.get('target_warehouse')
        transfer_quantity = request.POST.get('quantity', None)

        try:
            batch = BatchNumber.objects.get(batch_number=batch_number)
            old_warehouse = batch.warehouse
            new_warehouse = Warehouse.objects.get(pk=target_warehouse_id)

            # Wenn eine Menge angegeben wurde, teile die Charge
            if transfer_quantity:
                transfer_quantity = float(transfer_quantity)

                # Prüfe, ob die zu transferierende Menge gültig ist
                if transfer_quantity <= 0 or transfer_quantity > batch.quantity:
                    messages.error(request, f'Ungültige Menge. Muss größer als 0 und maximal {batch.quantity} sein.')
                    return redirect('batch_number_transfer')

                # Wenn nicht die gesamte Menge transferiert wird, teile die Charge
                if transfer_quantity < batch.quantity:
                    # Erstelle eine neue Charge im Ziellager
                    new_batch = BatchNumber.objects.create(
                        batch_number=batch.batch_number,
                        product=batch.product,
                        variant=batch.variant,
                        quantity=transfer_quantity,
                        production_date=batch.production_date,
                        expiry_date=batch.expiry_date,
                        supplier=batch.supplier,
                        warehouse=new_warehouse,
                        notes=f'Transferiert von {old_warehouse.name if old_warehouse else "unbekannt"} am {timezone.now().strftime("%d.%m.%Y")}'
                    )

                    # Reduziere die Menge der Originalcharge
                    batch.quantity -= transfer_quantity
                    batch.save()

                    messages.success(request,
                                     f'{transfer_quantity} Einheiten der Charge {batch_number} wurden erfolgreich von '
                                     f'{old_warehouse.name if old_warehouse else "unbekanntem Lager"} nach {new_warehouse.name} transferiert.')
                else:
                    # Wenn die gesamte Menge transferiert wird, verschiebe die Charge
                    batch.warehouse = new_warehouse
                    batch.save()

                    messages.success(request, f'Charge {batch_number} wurde vollständig von '
                                              f'{old_warehouse.name if old_warehouse else "unbekanntem Lager"} nach {new_warehouse.name} transferiert.')
            else:
                # Wenn keine Menge angegeben wurde, verschiebe die gesamte Charge
                batch.warehouse = new_warehouse
                batch.save()

                messages.success(request, f'Charge {batch_number} wurde erfolgreich von '
                                          f'{old_warehouse.name if old_warehouse else "unbekanntem Lager"} nach {new_warehouse.name} transferiert.')

            return redirect('batch_number_list')

        except BatchNumber.DoesNotExist:
            messages.error(request, f'Charge {batch_number} wurde nicht gefunden.')
        except Warehouse.DoesNotExist:
            messages.error(request, 'Das Ziellager existiert nicht.')
        except ValueError:
            messages.error(request, 'Ungültige Mengenangabe.')
        except Exception as e:
            messages.error(request, f'Fehler beim Transferieren: {str(e)}')

    # Für GET-Anfragen oder bei Fehlern im POST
    warehouses = Warehouse.objects.filter(is_active=True)

    # Wenn eine Chargennummer als Parameter übergeben wurde
    initial_batch = request.GET.get('batch', '')

    context = {
        'warehouses': warehouses,
        'initial_batch': initial_batch,
    }

    return render(request, 'core/batch/batch_number_transfer.html', context)


@login_required
@permission_required('product', 'import')
def batch_number_import(request):
    """Importiert Chargen aus einer CSV-Datei."""
    # Verwendet die handle_csv_import Utility-Funktion
    return handle_csv_import(
        form_class=BatchNumberImportForm,
        importer_class=BatchNumberImporter,
        request=request,
        template_name='core/batch/batch_number_import.html',
        success_redirect='import_log_detail',  # oder 'batch_number_list'
        extra_context={
            'title': 'Chargen importieren',
            'description': 'Importieren Sie Chargen aus einer CSV-Datei.',
            'expected_format': 'product_sku,batch_number,quantity,warehouse_name,production_date,expiry_date,supplier_name',
            'example': 'P1001,LOT2023001,100,Hauptlager,2023-01-01,2025-01-01,Beispiel GmbH',
            'required_columns': ['product_sku', 'batch_number', 'quantity'],
            'optional_columns': ['warehouse_name', 'production_date', 'expiry_date', 'supplier_name', 'notes'],
        }
    )


@login_required
@permission_required('product', 'export')
def batch_number_export(request):
    """Exportiert Chargen in eine CSV- oder Excel-Datei."""
    # Filteroptionen für den Export
    filters = {
        'warehouse': request.GET.get('warehouse', ''),
        'product': request.GET.get('product', ''),
        'expiry': request.GET.get('expiry', ''),
        'search': request.GET.get('search', ''),
    }

    # Basisdaten
    batches = BatchNumber.objects.select_related('product', 'warehouse', 'variant', 'supplier')

    # Filter anwenden mit den Utility-Funktionen
    batches = apply_exact_filter(batches, 'warehouse_id', filters['warehouse'])
    batches = apply_exact_filter(batches, 'product_id', filters['product'])

    if filters['search']:
        batches = batches.filter(
            Q(batch_number__icontains=filters['search']) |
            Q(product__name__icontains=filters['search']) |
            Q(product__sku__icontains=filters['search'])
        )

    # Filtern nach Verfallsdatum mit apply_expiry_filter
    today = timezone.now().date()
    batches = apply_expiry_filter(batches, filters['expiry'], today, 30)

    # Export-Format bestimmen
    export_format = request.GET.get('format', '')

    if export_format == 'csv':
        # CSV-Export
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="chargen_{timezone.now().strftime("%Y%m%d")}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Chargennummer', 'Produkt', 'SKU', 'Variante', 'Menge', 'Einheit', 'Lager',
            'Lieferant', 'Produktionsdatum', 'Ablaufdatum', 'Notizen', 'Erstellt am'
        ])

        for batch in batches:
            writer.writerow([
                batch.batch_number,
                batch.product.name,
                batch.product.sku,
                batch.variant.name if batch.variant else '',
                str(batch.quantity),
                batch.product.unit,
                batch.warehouse.name if batch.warehouse else '',
                batch.supplier.name if batch.supplier else '',
                batch.production_date.strftime('%Y-%m-%d') if batch.production_date else '',
                batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else '',
                batch.notes,
                batch.created_at.strftime('%Y-%m-%d %H:%M')
            ])

        return response

    elif export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            # Funktion zum Entfernen der Zeitzone
            def remove_timezone(dt):
                if dt and hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
                    return dt.replace(tzinfo=None)
                return dt

            # Neue Arbeitsmappe erstellen
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Chargen"

            # Überschriften formatieren
            headers = [
                'Chargennummer', 'Produkt', 'SKU', 'Variante', 'Menge', 'Einheit', 'Lager',
                'Lieferant', 'Produktionsdatum', 'Ablaufdatum', 'Notizen', 'Erstellt am'
            ]

            # Überschriften-Stil festlegen
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Überschriften schreiben und formatieren
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Daten einfügen
            for row_idx, batch in enumerate(batches, 2):
                ws.cell(row=row_idx, column=1, value=batch.batch_number)
                ws.cell(row=row_idx, column=2, value=batch.product.name)
                ws.cell(row=row_idx, column=3, value=batch.product.sku)
                ws.cell(row=row_idx, column=4, value=batch.variant.name if batch.variant else '')
                ws.cell(row=row_idx, column=5, value=float(batch.quantity))
                ws.cell(row=row_idx, column=6, value=batch.product.unit)
                ws.cell(row=row_idx, column=7, value=batch.warehouse.name if batch.warehouse else '')
                ws.cell(row=row_idx, column=8, value=batch.supplier.name if batch.supplier else '')

                if batch.production_date:
                    ws.cell(row=row_idx, column=9, value=batch.production_date)

                if batch.expiry_date:
                    ws.cell(row=row_idx, column=10, value=batch.expiry_date)
                    # Abgelaufene Chargen rot markieren
                    if batch.expiry_date < today:
                        ws.cell(row=row_idx, column=10).fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB",
                                                                           fill_type="solid")
                    # Bald ablaufende Chargen gelb markieren
                    elif batch.expiry_date <= today + timedelta(days=30):
                        ws.cell(row=row_idx, column=10).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC",
                                                                           fill_type="solid")

                ws.cell(row=row_idx, column=11, value=batch.notes)
                # Zeitzone aus dem created_at entfernen, um Excel-Fehler zu vermeiden
                created_at_value = remove_timezone(batch.created_at) if batch.created_at else None
                ws.cell(row=row_idx, column=12, value=created_at_value)

            # Spaltenbreiten anpassen
            for col_idx, header in enumerate(headers, 1):
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                max_length = len(header) + 2

                # Überprüfe Zellinhalte, um die optimale Spaltenbreite zu bestimmen
                for row_idx in range(2, len(batches) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)) + 2)

                # Spaltenbreite setzen (maximale Breite: 50)
                ws.column_dimensions[column_letter].width = min(max_length, 50)

            # Excel-Datei als Response zurückgeben
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="chargen_{timezone.now().strftime("%Y%m%d")}.xlsx"'

            return response

        except ImportError:
            messages.error(request, 'Excel-Export ist nicht verfügbar. Bitte installieren Sie openpyxl.')
            return redirect('batch_number_list')
        except Exception as e:
            messages.error(request, f'Fehler beim Excel-Export: {str(e)}')
            return redirect('batch_number_list')


    elif export_format == 'pdf':
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                    leftMargin=1 * cm, rightMargin=1 * cm,
                                    topMargin=1 * cm, bottomMargin=1 * cm)
            elements = []

            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            title_style.alignment = 1

            cell_style = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontSize=8.5,
                leading=10,
                wordWrap='CJK',
                alignment=0,
            )

            header_style = ParagraphStyle(
                'HeaderStyle',
                parent=styles['Normal'],
                fontSize=9,
                fontName='Helvetica-Bold',
                alignment=1,
                textColor=colors.white,
            )

            elements.append(Paragraph("Chargenexport", title_style))
            elements.append(Spacer(1, 10))

            # Dynamische Seitenbreite
            PAGE_WIDTH = landscape(A4)[0] - 2 * cm

            col_widths = [
                0.12 * PAGE_WIDTH,  # Chargennummer
                0.15 * PAGE_WIDTH,  # Produkt
                0.08 * PAGE_WIDTH,  # SKU
                0.08 * PAGE_WIDTH,  # Variante
                0.06 * PAGE_WIDTH,  # Menge
                0.07 * PAGE_WIDTH,  # Lager
                0.10 * PAGE_WIDTH,  # Lieferant
                0.08 * PAGE_WIDTH,  # Produktionsdatum
                0.08 * PAGE_WIDTH,  # Ablaufdatum
                0.14 * PAGE_WIDTH,  # Notizen
                0.04 * PAGE_WIDTH,  # Erstellt am
            ]

            headers = [
                Paragraph('<font color="white">Chargennummer</font>', header_style),
                Paragraph('<font color="white">Produkt</font>', header_style),
                Paragraph('<font color="white">SKU</font>', header_style),
                Paragraph('<font color="white">Variante</font>', header_style),
                Paragraph('<font color="white">Menge</font>', header_style),
                Paragraph('<font color="white">Lager</font>', header_style),
                Paragraph('<font color="white">Lieferant</font>', header_style),
                Paragraph('<font color="white">Produktionsdatum</font>', header_style),
                Paragraph('<font color="white">Ablaufdatum</font>', header_style),
                Paragraph('<font color="white">Notizen</font>', header_style),
                Paragraph('<font color="white">Erstellt am</font>', header_style),
            ]

            data = [headers]

            for batch in batches:
                row = [
                    Paragraph(batch.batch_number, cell_style),
                    Paragraph(batch.product.name, cell_style),
                    Paragraph(batch.product.sku, cell_style),
                    Paragraph(batch.variant.name if batch.variant else '', cell_style),
                    Paragraph(f"{batch.quantity} {batch.product.unit}", cell_style),
                    Paragraph(batch.warehouse.name if batch.warehouse else '', cell_style),
                    Paragraph(batch.supplier.name if batch.supplier else '', cell_style),
                    Paragraph(batch.production_date.strftime('%d.%m.%Y') if batch.production_date else '', cell_style),
                    Paragraph(batch.expiry_date.strftime('%d.%m.%Y') if batch.expiry_date else '', cell_style),
                    Paragraph(batch.notes if batch.notes else '', cell_style),
                    Paragraph(batch.created_at.strftime('%d.%m.%Y %H:%M') if batch.created_at else '', cell_style)
                ]
                data.append(row)

            table = Table(data, colWidths=col_widths, repeatRows=1)

            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
                ('TOPPADDING', (0, 0), (-1, 0), 7),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
                ('TOPPADDING', (0, 1), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
                ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
            ])

            for i, batch in enumerate(batches, 1):
                if batch.expiry_date:
                    if batch.expiry_date < today:
                        table_style.add('BACKGROUND', (8, i), (8, i), colors.HexColor('#FFCCCB'))
                    elif batch.expiry_date <= today + timedelta(days=30):
                        table_style.add('BACKGROUND', (8, i), (8, i), colors.HexColor('#FFFFCC'))

                if i % 2 == 0:
                    table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F5F5F5'))

            table.setStyle(table_style)
            elements.append(table)

            doc.build(elements)
            buffer.seek(0)

            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="chargen_{timezone.now().strftime("%Y%m%d")}.pdf"'
            return response

        except ImportError:
            messages.error(request, 'PDF-Export ist nicht verfügbar. Bitte installieren Sie reportlab.')
            return redirect('batch_number_list')

        except Exception as e:
            messages.error(request, f'Fehler beim PDF-Export: {str(e)}')
            return redirect('batch_number_list')

    # Statistiken für die Verfallsdaten
    expired_count = BatchNumber.objects.filter(expiry_date__lt=today).count()
    expiring_soon_count = BatchNumber.objects.filter(
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30)
    ).count()
    valid_count = BatchNumber.objects.filter(
        expiry_date__gt=today + timedelta(days=30)
    ).count()

    # Filter-Werte für das Template bereitstellen
    context = {
        'warehouses': Warehouse.objects.filter(is_active=True),
        'products': Product.objects.filter(has_batch_tracking=True),
        'batches': batches,
        'today': today,
        'date_today': timezone.now().strftime("%Y%m%d"),
        'expired_count': expired_count,
        'expiring_soon_count': expiring_soon_count,
        'valid_count': valid_count,
        'warehouse_filter': filters['warehouse'],
        'product_filter': filters['product'],
        'expiry_filter': filters['expiry'],
        'search_query': filters['search'],
    }

    return render(request, 'core/batch/batch_number_export.html', context)

@login_required
def api_product_variants(request):
    """API-Endpunkt, um Varianten für ein bestimmtes Produkt zu erhalten."""
    product_id = request.GET.get('product_id', None)

    if not product_id:
        return JsonResponse([], safe=False)

    variants = ProductVariant.objects.filter(
        parent_product_id=product_id,
        is_active=True
    ).values('id', 'name', 'value')

    return JsonResponse(list(variants), safe=False)