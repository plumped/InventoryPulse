import csv
import io
from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import permission_required, login_required
from django.http.response import HttpResponse
from django.shortcuts import redirect

from data_operations.models.importers_models import ImportLog


@login_required
@permission_required('report', 'view')
def export_import_logs(request):
    """Export import logs to CSV or Excel."""
    # Base queryset
    queryset = ImportLog.objects.all()

    # Apply filters
    import_type = request.GET.get('type')
    if import_type:
        queryset = queryset.filter(import_type=import_type)

    # Check for specific IDs
    ids_str = request.GET.get('ids', '')
    if ids_str:
        ids = [int(id_str.strip()) for id_str in ids_str.split(',') if id_str.strip()]
        queryset = queryset.filter(pk__in=ids)

    # Determine format
    export_format = request.GET.get('format', 'csv').lower()

    # CSV export
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="import_logs_{datetime.now().strftime("%Y%m%d")}.csv"'

        writer = csv.writer(response)

        # Write header
        writer.writerow(['ID', 'Type', 'Filename', 'Status', 'Total Rows', 'Successful Rows',
                         'Error Rows', 'Success Rate', 'User', 'Created At'])

        # Write data
        for log in queryset:
            writer.writerow([
                log.pk,
                log.import_type,
                log.file_name,
                log.status,
                log.total_records,
                log.success_count,
                log.error_count,
                f"{log.success_rate()}%",
                log.created_by.username if log.created_by else 'System',
                log.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])

        return response

    # Excel export
    elif export_format == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            messages.error(request, 'Excel-Export ist nicht verfügbar. Bitte installieren Sie openpyxl.')
            return redirect('import_log_list')

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Logs"

        # Define headers
        headers = ['ID', 'Type', 'Filename', 'Status', 'Total Rows', 'Successful Rows',
                   'Error Rows', 'Success Rate', 'User', 'Created At']

        # Define styles for header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Write and format header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for row_num, log in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=log.pk)
            ws.cell(row=row_num, column=2, value=log.import_type)
            ws.cell(row=row_num, column=3, value=log.file_name)
            ws.cell(row=row_num, column=4, value=log.status)
            ws.cell(row=row_num, column=5, value=log.total_records)
            ws.cell(row=row_num, column=6, value=log.success_count)
            ws.cell(row=row_num, column=7, value=log.error_count)
            ws.cell(row=row_num, column=8, value=f"{log.success_rate()}%")
            ws.cell(row=row_num, column=9, value=log.created_by.username if log.created_by else 'System')
            ws.cell(row=row_num, column=10, value=log.created_at.strftime('%Y-%m-%d %H:%M:%S'))

        # Adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            # Minimum width based on header + some space
            max_length = len(header) + 2

            # Check width based on data
            for row_num in range(2, len(queryset) + 2):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)) + 2)

            ws.column_dimensions[column_letter].width = max_length

        # Write Excel file to a BytesIO stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="import_logs_{datetime.now().strftime("%Y%m%d")}.xlsx"'

        return response

    # Unsupported format
    else:
        messages.error(request, f'Das Format "{export_format}" wird nicht unterstützt.')
        return redirect('import_log_list')
