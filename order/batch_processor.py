# order/batch_processor.py

import csv
import io
from decimal import Decimal

from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from master_data.models.currency_models import Currency
from master_data.models.systemsettings_models import SystemSettings
from order.models import PurchaseOrder, PurchaseOrderItem
from order.workflow import get_initial_order_status
from suppliers.models import Supplier, SupplierProduct


class BatchOrderProcessor:
    """Processor for handling batch order creation from CSV files."""

    REQUIRED_HEADERS = ['supplier_name', 'product_sku', 'quantity']
    OPTIONAL_HEADERS = ['notes', 'expected_delivery', 'shipping_address', 'unit_price', 'supplier_sku']

    def __init__(self, file_obj, delimiter=',', encoding='utf-8', skip_header=True, user=None):
        self.file_obj = file_obj
        self.delimiter = delimiter
        self.encoding = encoding
        self.skip_header = skip_header
        self.user = user
        self.errors = []
        self.warnings = []
        self.orders_created = 0
        self.orders_updated = 0
        self.line_items_processed = 0

    def validate_headers(self, headers):
        """Validate that all required headers are present in the CSV."""
        missing_headers = [h for h in self.REQUIRED_HEADERS if h not in headers]
        if missing_headers:
            self.errors.append(f"Missing required headers: {', '.join(missing_headers)}")
            return False
        return True

    def process_excel_file(file_obj):
        """Process an Excel file and convert it to CSV-like rows."""
        try:
            import openpyxl

            # Load the workbook
            wb = openpyxl.load_workbook(file_obj, data_only=True)

            # Use the first sheet
            ws = wb.active

            # Convert worksheet to rows
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Convert any None values to empty strings
                cleaned_row = ['' if cell is None else str(cell) for cell in row]
                rows.append(cleaned_row)

            return rows
        except Exception as e:
            raise ValueError(f"Error processing Excel file: {str(e)}")

    def read_csv(self):
        """Read the CSV or Excel file and return a list of dictionaries."""
        try:
            # Ensure we're at the beginning of the file
            self.file_obj.seek(0)

            # Check if this is an Excel file
            if hasattr(self.file_obj, 'name') and self.file_obj.name.lower().endswith(('.xlsx', '.xls')):
                rows = self.process_excel_file(self.file_obj)
            else:
                # Process as CSV
                content = self.file_obj.read().decode(self.encoding)
                csv_reader = csv.reader(io.StringIO(content), delimiter=self.delimiter)
                rows = list(csv_reader)

            if not rows:
                self.errors.append("File is empty")
                return []

            # Extract header and use it as keys for dictionaries
            header = rows[0]

            # Validate headers
            if not self.validate_headers(header):
                return []

            # Skip header if needed
            if self.skip_header:
                data = rows[1:]
            else:
                data = rows

            # Convert to list of dictionaries
            result = []
            for i, row in enumerate(data):
                # Skip empty rows
                if not row or (len(row) == 1 and not row[0]):
                    continue

                # Pad row with empty strings if it's shorter than the header
                if len(row) < len(header):
                    row.extend([''] * (len(header) - len(row)))
                # Truncate row if it's longer than the header
                elif len(row) > len(header):
                    row = row[:len(header)]

                # Create dictionary
                row_dict = dict(zip(header, row))
                result.append(row_dict)

            return result

        except Exception as e:
            self.errors.append(f"Error reading file: {str(e)}")
            return []

    def group_by_supplier(self, rows):
        """Group rows by supplier."""
        suppliers = {}

        for i, row in enumerate(rows, start=1):
            # Extract supplier name
            supplier_name = row.get('supplier_name', '').strip()
            if not supplier_name:
                self.errors.append(f"Row {i}: Missing supplier name")
                continue

            # Find supplier in database
            try:
                supplier = Supplier.objects.get(name=supplier_name)
            except Supplier.DoesNotExist:
                self.errors.append(f"Row {i}: Supplier '{supplier_name}' not found")
                continue

            # Add to supplier group
            if supplier.id not in suppliers:
                suppliers[supplier.id] = {
                    'supplier': supplier,
                    'items': [],
                    'notes': row.get('notes', ''),
                    'expected_delivery': row.get('expected_delivery', ''),
                    'shipping_address': row.get('shipping_address', '')
                }

            # Add line item to supplier
            suppliers[supplier.id]['items'].append({
                'row': i,
                'product_sku': row.get('product_sku', '').strip(),
                'quantity': row.get('quantity', '').strip(),
                'unit_price': row.get('unit_price', '').strip(),
                'supplier_sku': row.get('supplier_sku', '').strip()
            })

        return suppliers

    @transaction.atomic
    def process_orders(self):
        """Process the CSV and create purchase orders."""
        # Read CSV
        rows = self.read_csv()
        if not rows:
            return False

        # Group by supplier
        supplier_groups = self.group_by_supplier(rows)
        if not supplier_groups:
            self.errors.append("No valid supplier data found")
            return False

        # Create orders for each supplier
        for supplier_id, data in supplier_groups.items():
            supplier = data['supplier']
            items = data['items']

            # Skip suppliers with no valid items
            if not items:
                continue

            try:
                # Create order for this supplier
                order = self._create_purchase_order(supplier, data)

                # Add line items
                self._process_line_items(order, items)

                # Update order totals
                order.update_totals()

                # Increment counter
                self.orders_created += 1

            except Exception as e:
                self.errors.append(f"Error creating order for {supplier.name}: {str(e)}")
                # Roll back this order but continue with others
                transaction.set_rollback(True)
                continue

        return True

    def _create_purchase_order(self, supplier, data):
        """Create a purchase order for a supplier."""
        # Generate order number
        today = timezone.now().date()

        try:
            system_settings = SystemSettings.objects.first()
            prefix = system_settings.order_number_prefix if system_settings else "ORD-"
            next_seq = system_settings.next_order_number if system_settings else 1
        except:
            # Fallback if no system settings
            prefix = "ORD-"
            next_seq = 1

        # Find next sequence number
        order_number_prefix = f"{prefix}{today.strftime('%Y%m%d')}-"
        last_order = PurchaseOrder.objects.filter(
            order_number__startswith=order_number_prefix
        ).order_by('-order_number').first()

        if last_order:
            try:
                last_seq = int(last_order.order_number.split('-')[-1])
                next_seq = last_seq + 1
            except ValueError:
                pass

        order_number = f"{order_number_prefix}{next_seq:03d}"

        # Parse expected delivery date
        expected_delivery = None
        if data['expected_delivery']:
            try:
                from datetime import datetime
                expected_delivery = datetime.strptime(data['expected_delivery'], '%Y-%m-%d').date()
            except ValueError:
                self.warnings.append(
                    f"Invalid date format for expected delivery: {data['expected_delivery']}. Format should be YYYY-MM-DD.")

        # Create order
        order = PurchaseOrder.objects.create(
            order_number=order_number,
            supplier=supplier,
            status=get_initial_order_status(None),
            shipping_address=data['shipping_address'],
            notes=data['notes'],
            expected_delivery=expected_delivery,
            created_by=self.user
        )

        # Update system settings if they exist
        if 'system_settings' in locals() and system_settings:
            system_settings.next_order_number = next_seq + 1
            system_settings.save()

        return order

    def _process_line_items(self, order, items):
        """Process line items for a purchase order."""
        for item in items:
            try:
                # Get product
                product_sku = item['product_sku']
                if not product_sku:
                    self.errors.append(f"Row {item['row']}: Missing product SKU")
                    continue

                try:
                    product = Product.objects.get(sku=product_sku)
                except Product.DoesNotExist:
                    self.errors.append(f"Row {item['row']}: Product with SKU '{product_sku}' not found")
                    continue

                # Parse quantity
                quantity_str = item['quantity']
                try:
                    quantity = Decimal(quantity_str)
                    if quantity <= 0:
                        self.errors.append(f"Row {item['row']}: Quantity must be greater than zero")
                        continue
                except (ValueError, TypeError):
                    self.errors.append(f"Row {item['row']}: Invalid quantity '{quantity_str}'")
                    continue

                # Get unit price from CSV or supplier product
                unit_price = None
                if item['unit_price']:
                    try:
                        unit_price = Decimal(item['unit_price'])
                    except (ValueError, TypeError):
                        self.warnings.append(
                            f"Row {item['row']}: Invalid unit price '{item['unit_price']}'. Will use supplier price if available.")

                # If unit price not specified or invalid, get from supplier product
                if unit_price is None:
                    try:
                        supplier_product = SupplierProduct.objects.get(
                            supplier=order.supplier,
                            product=product
                        )
                        unit_price = supplier_product.purchase_price
                        # Use supplier_sku from database if not specified in CSV
                        if not item['supplier_sku']:
                            item['supplier_sku'] = supplier_product.supplier_sku
                        # Use currency from supplier product
                        currency = supplier_product.currency or order.supplier.default_currency
                    except SupplierProduct.DoesNotExist:
                        # Default to 0 if no price available
                        unit_price = Decimal('0.00')
                        self.warnings.append(
                            f"Row {item['row']}: No supplier product found for {product.name}. Using price 0.00.")
                        currency = None

                # Get currency
                if not currency:
                    currency = Currency.get_default_currency()

                # Create order item
                PurchaseOrderItem.objects.create(
                    purchase_order=order,
                    product=product,
                    quantity_ordered=quantity,
                    unit_price=unit_price,
                    supplier_sku=item['supplier_sku'],
                    tax=product.tax,
                    currency=currency
                )

                # Increment counter
                self.line_items_processed += 1

            except Exception as e:
                self.errors.append(f"Row {item['row']}: Error processing line item: {str(e)}")

    def generate_template(self):
        """Generate a CSV template for batch orders."""
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        header = self.REQUIRED_HEADERS + self.OPTIONAL_HEADERS
        writer.writerow(header)

        # Write a few example rows
        writer.writerow(
            ['Supplier A', 'PROD001', '10', 'Regular monthly order', '2023-12-31', 'Main Warehouse, 123 Street', '',
             'SUP001'])
        writer.writerow(['Supplier A', 'PROD002', '5', '', '', '', '15.99', ''])
        writer.writerow(['Supplier B', 'PROD003', '20', 'Urgent order', '2023-11-30', '', '', ''])

        output.seek(0)
        return output.read()

    def generate_template_excel(self):
        """Generate an Excel template for batch orders."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Order Template"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write header
        header = self.REQUIRED_HEADERS + self.OPTIONAL_HEADERS
        for col_num, column_title in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add sample data
        sample_data = [
            ['Supplier A', 'PROD001', '10', 'Regular monthly order', '2023-12-31', 'Main Warehouse, 123 Street', '',
             'SUP001'],
            ['Supplier A', 'PROD002', '5', '', '', '', '15.99', ''],
            ['Supplier B', 'PROD003', '20', 'Urgent order', '2023-11-30', '', '', '']
        ]

        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.border = thin_border

        # Add instructions in a new sheet
        ws_instructions = wb.create_sheet(title="Instructions")

        instructions = [
            ["Batch Order Import Instructions"],
            [""],
            ["Required Fields:"],
            ["supplier_name - The name of the supplier exactly as it appears in the system"],
            ["product_sku - The SKU (Stock Keeping Unit) of the product"],
            ["quantity - The quantity to order"],
            [""],
            ["Optional Fields:"],
            ["notes - Any notes for the order"],
            ["expected_delivery - Expected delivery date in YYYY-MM-DD format"],
            ["shipping_address - The shipping address for the order"],
            ["unit_price - Override the default supplier price"],
            ["supplier_sku - The supplier's SKU for the product"],
            [""],
            ["Notes:"],
            ["- Multiple lines for the same supplier will be combined into a single order"],
            ["- If unit_price is not specified, the system will use the price from the supplier product"],
            ["- If expected_delivery is not in YYYY-MM-DD format, it will be ignored"]
        ]

        for row_num, row_data in enumerate(instructions, 1):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws_instructions.cell(row=row_num, column=col_num, value=cell_value)
                if row_num == 1:
                    cell.font = Font(bold=True, size=14)
                elif "Required Fields:" in cell_value:
                    cell.font = Font(bold=True)
                elif "Optional Fields:" in cell_value:
                    cell.font = Font(bold=True)
                elif "Notes:" in cell_value:
                    cell.font = Font(bold=True)

        # Adjust column widths
        for ws_name in [ws, ws_instructions]:
            for column in ws_name.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_name.column_dimensions[column_letter].width = adjusted_width

        # Return Excel file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


def batch_order_import_view(request):
    """View for importing batch orders from CSV or Excel."""
    if request.method == 'POST':
        if 'csv_file' not in request.FILES:
            messages.error(request, "Please upload a CSV or Excel file.")
            return redirect('batch_order_import')

        uploaded_file = request.FILES['csv_file']

        # Check file type
        file_extension = uploaded_file.name.split('.')[-1].lower()
        if file_extension not in ['csv', 'xlsx', 'xls']:
            messages.error(request, "Unsupported file format. Please upload a CSV or Excel file.")
            return redirect('batch_order_import')

        # Get form parameters
        delimiter = request.POST.get('delimiter', ',')
        encoding = request.POST.get('encoding', 'utf-8')
        skip_header = 'skip_header' in request.POST

        # Process the file
        processor = BatchOrderProcessor(
            file_obj=uploaded_file,
            delimiter=delimiter,
            encoding=encoding,
            skip_header=skip_header,
            user=request.user
        )

        success = processor.process_orders()

        # Store results in session for result page
        result_data = {
            'success': success,
            'orders_created': processor.orders_created,
            'line_items_processed': processor.line_items_processed,
            'errors': processor.errors,
            'warnings': processor.warnings,
        }
        request.session['batch_import_result'] = result_data

        # Redirect to result page if there are errors or warnings to display
        if processor.errors or processor.warnings:
            return redirect('batch_order_import_result')

        # Otherwise, show a success message and redirect to orders list
        if processor.orders_created > 0:
            messages.success(
                request,
                f"Successfully created {processor.orders_created} orders with {processor.line_items_processed} line items."
            )
        else:
            messages.warning(request, "No orders were created.")

        return redirect('purchase_order_list')

    context = {
        'title': 'Batch Order Import',
    }

    return render(request, 'order/batch_order_import.html', context)


def download_order_template(request, format_type='csv'):
    """View for downloading the batch order template."""
    processor = BatchOrderProcessor(None)

    if format_type == 'csv':
        # Generate CSV template
        content = processor.generate_template()
        response = HttpResponse(content, content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="batch_order_template.csv"'
    else:
        # Generate Excel template
        content = processor.generate_template_excel()
        response = HttpResponse(content,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="batch_order_template.xlsx"'

    return response


class BatchImportResult:
    """Class to store and report batch import results."""

    def __init__(self):
        self.orders_created = 0
        self.orders_updated = 0
        self.line_items_processed = 0
        self.errors = []
        self.warnings = []
        self.supplier_orders = {}  # Store orders by supplier

    def add_error(self, error_msg, row_number=None):
        """Add an error with optional row number."""
        if row_number:
            self.errors.append(f"Row {row_number}: {error_msg}")
        else:
            self.errors.append(error_msg)

    def add_warning(self, warning_msg, row_number=None):
        """Add a warning with optional row number."""
        if row_number:
            self.warnings.append(f"Row {row_number}: {warning_msg}")
        else:
            self.warnings.append(warning_msg)

    def add_order(self, supplier_name, order):
        """Add an order to the results."""
        if supplier_name not in self.supplier_orders:
            self.supplier_orders[supplier_name] = []

        self.supplier_orders[supplier_name].append(order)
        self.orders_created += 1

    def add_line_item(self, count=1):
        """Increment the line items count."""
        self.line_items_processed += count

    def has_errors(self):
        """Check if there are any errors."""
        return len(self.errors) > 0

    def has_warnings(self):
        """Check if there are any warnings."""
        return len(self.warnings) > 0

    def get_summary(self):
        """Get a summary of the import results."""
        return {
            'orders_created': self.orders_created,
            'line_items_processed': self.line_items_processed,
            'error_count': len(self.errors),
            'warning_count': len(self.warnings),
            'supplier_count': len(self.supplier_orders),
        }


def generate_error_report(result, file_format='csv'):
    """Generate a report of errors and warnings."""
    if file_format == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow(['Type', 'Message'])

        # Write errors
        for error in result.errors:
            writer.writerow(['Error', error])

        # Write warnings
        for warning in result.warnings:
            writer.writerow(['Warning', warning])

        output.seek(0)
        return output.getvalue()
    else:
        # Excel format
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Errors"

        # Define styles
        error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        warning_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        header_font = Font(bold=True)

        # Write header
        ws.cell(row=1, column=1, value="Type").font = header_font
        ws.cell(row=1, column=2, value="Message").font = header_font

        # Write errors
        row_num = 2
        for error in result.errors:
            ws.cell(row=row_num, column=1, value="Error").fill = error_fill
            ws.cell(row=row_num, column=2, value=error).fill = error_fill
            row_num += 1

        # Write warnings
        for warning in result.warnings:
            ws.cell(row=row_num, column=1, value="Warning").fill = warning_fill
            ws.cell(row=row_num, column=2, value=warning).fill = warning_fill
            row_num += 1

        # Adjust column width
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 100

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


def batch_order_import_result(request):
    """View for displaying batch import results with detailed error information."""
    # Get result data from session
    result_data = request.session.get('batch_import_result', {})

    # Clear session data
    if 'batch_import_result' in request.session:
        del request.session['batch_import_result']

    if not result_data:
        messages.error(request, "No import results found.")
        return redirect('batch_order_import')

    context = {
        'title': 'Batch Import Results',
        'result': result_data,
        'show_download': bool(result_data.get('errors') or result_data.get('warnings')),
    }

    return render(request, 'order/batch_order_import_result.html', context)


def download_error_report(request, format_type='csv'):
    """Download the error report as CSV or Excel."""
    # Get result data from session
    result_data = request.session.get('batch_import_result', {})

    if not result_data:
        messages.error(request, "No import results found.")
        return redirect('batch_order_import')

    # Create a result object
    result = BatchImportResult()
    result.errors = result_data.get('errors', [])
    result.warnings = result_data.get('warnings', [])

    # Generate report
    if format_type == 'xlsx':
        content = generate_error_report(result, 'xlsx')
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        filename = 'batch_import_errors.xlsx'
    else:
        content = generate_error_report(result, 'csv')
        content_type = 'text/csv'
        filename = 'batch_import_errors.csv'

    # Create response
    response = HttpResponse(content, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response