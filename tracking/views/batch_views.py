import csv
import io
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models.query_utils import Q
from django.http.response import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from core.forms import BatchNumberForm, BatchNumberImportForm
from core.models import BatchNumber
from core.utils.filters import apply_exact_filter, apply_expiry_filter
from core.utils.forms import handle_form_view
from core.utils.imports import handle_csv_import
from core.utils.pagination import paginate_queryset
from core.utils.view_helpers import handle_model_update, handle_model_delete
from data_operations.importers.batch_numbers import BatchNumberImporter
from inventory.models import Warehouse
from product_management.models.products import Product


@login_required
@permission_required('products.view_product', raise_exception=True)
def product_batches(request, pk):
    """Zeigt alle Chargen eines Produkts an."""
    product = get_object_or_404(Product, pk=pk)

    # Filteroptionen
    warehouse_filter = request.GET.get('warehouse', '')
    variant_filter = request.GET.get('variant', '')
    search_query = request.GET.get('search', '')
    expiry_filter = request.GET.get('expiry', '')

    batches = BatchNumber.objects.filter(product=product)

    if warehouse_filter:
        batches = batches.filter(warehouse_id=warehouse_filter)

    if variant_filter:
        batches = batches.filter(variant_id=variant_filter)

    if search_query:
        batches = batches.filter(batch_number__icontains=search_query)

    # Filtern nach Verfallsdatum
    today = timezone.now().date()
    if expiry_filter == 'expired':
        batches = batches.filter(expiry_date__lt=today)
    elif expiry_filter == 'expiring_soon':
        batches = batches.filter(expiry_date__gte=today, expiry_date__lte=today + timedelta(days=30))
    elif expiry_filter == 'valid':
        batches = batches.filter(expiry_date__gt=today + timedelta(days=30))

    # Paginierung
    batches = paginate_queryset(batches, request.GET.get('page'), per_page=25)

    # Verfallsstatistiken
    expired_count = BatchNumber.objects.filter(product=product, expiry_date__lt=today).count()
    expiring_soon_count = BatchNumber.objects.filter(
        product=product,
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30)
    ).count()
    valid_count = BatchNumber.objects.filter(
        product=product,
        expiry_date__gt=today + timedelta(days=30)
    ).count()

    context = {
        'product': product,
        'batches': batches,
        'warehouses': Warehouse.objects.filter(is_active=True),
        'variants': product.variants.all(),
        'warehouse_filter': warehouse_filter,
        'variant_filter': variant_filter,
        'search_query': search_query,
        'expiry_filter': expiry_filter,
        'expired_count': expired_count,
        'expiring_soon_count': expiring_soon_count,
        'valid_count': valid_count,
        'today': today,
    }

    return render(request, 'core/product/product_batches.html', context)


@login_required
@permission_required('product', 'create')
def product_batch_add(request, pk):
    """Fügt eine Charge zu einem Produkt hinzu."""
    product = get_object_or_404(Product, pk=pk)

    # Prüfen, ob das Produkt für Chargen konfiguriert ist
    if not product.has_batch_tracking:
        product.has_batch_tracking = True
        product.save()
        messages.info(request, 'Das Produkt wurde für Chargenverfolgung aktiviert.')

    if request.method == 'POST':
        form = BatchNumberForm(request.POST, product=product)
        if form.is_valid():
            batch = form.save(commit=False)
            batch.product = product
            batch.save()

            messages.success(request, 'Charge wurde erfolgreich hinzugefügt.')
            return redirect('product_batches', pk=product.pk)
    else:
        form = BatchNumberForm(product=product)

    context = {
        'product': product,
        'form': form,
    }

    return render(request, 'core/product/product_batch_form.html', context)


@login_required
@permission_required('product', 'edit')
def product_batch_update(request, pk, batch_id):
    """Aktualisiert eine Charge."""
    product = get_object_or_404(Product, pk=pk)
    batch = get_object_or_404(BatchNumber, pk=batch_id, product=product)

    if request.method == 'POST':
        form = BatchNumberForm(request.POST, instance=batch, product=product)
        if form.is_valid():
            batch = form.save()
            messages.success(request, 'Charge wurde erfolgreich aktualisiert.')
            return redirect('product_batches', pk=product.pk)
    else:
        form = BatchNumberForm(instance=batch, product=product)

    context = {
        'product': product,
        'batch': batch,
        'form': form,
    }

    return render(request, 'core/product/product_batch_form.html', context)


@login_required
@permission_required('product', 'delete')
def product_batch_delete(request, pk, batch_id):
    """Löscht eine Charge."""
    product = get_object_or_404(Product, pk=pk)
    batch = get_object_or_404(BatchNumber, pk=batch_id, product=product)

    if request.method == 'POST':
        batch.delete()
        messages.success(request, 'Charge wurde erfolgreich gelöscht.')
        return redirect('product_batches', pk=product.pk)

    context = {
        'product': product,
        'batch': batch,
    }

    return render(request, 'core/product/product_batch_confirm_delete.html', context)


@login_required
@permission_required('products.view_product', raise_exception=True)
def batch_number_list(request):
    """Liste aller Chargen im System."""
    # Filteroptionen
    filters = {
        'warehouse': request.GET.get('warehouse', ''),
        'product': request.GET.get('product', ''),
        'search': request.GET.get('search', ''),
        'expiry': request.GET.get('expiry', ''),
    }

    # Basisdaten
    batches = BatchNumber.objects.select_related('product', 'warehouse', 'variant')

    # Filter anwenden mit den Utility-Funktionen
    batches = apply_exact_filter(batches, 'warehouse_id', filters['warehouse'])
    batches = apply_exact_filter(batches, 'product_id', filters['product'])

    if filters['search']:
        batches = batches.filter(
            Q(batch_number__icontains=filters['search']) |
            Q(product__name__icontains=filters['search']) |
            Q(product__sku__icontains=filters['search'])
        )

    # Filtern nach Verfallsdatum mit der Utility-Funktion
    today = timezone.now().date()
    batches = apply_expiry_filter(batches, filters['expiry'], today, 30)

    # Sortierung
    sort_by = request.GET.get('sort', '-created_at')
    batches = batches.order_by(sort_by)

    # Paginierung mit der Utility-Funktion
    batches = paginate_queryset(batches, request.GET.get('page'), per_page=25)

    # Statistiken für die Verfallsdaten
    expired_count = BatchNumber.objects.filter(expiry_date__lt=today).count()
    expiring_soon_count = BatchNumber.objects.filter(
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30)
    ).count()
    valid_count = BatchNumber.objects.filter(
        expiry_date__gt=today + timedelta(days=30)
    ).count()
    total_count = BatchNumber.objects.count()

    context = {
        'batches': batches,
        'warehouses': Warehouse.objects.filter(is_active=True),
        'products': Product.objects.filter(has_batch_tracking=True),
        'warehouse_filter': filters['warehouse'],
        'product_filter': filters['product'],
        'search_query': filters['search'],
        'expiry_filter': filters['expiry'],
        'sort_by': sort_by,
        'expired_count': expired_count,
        'expiring_soon_count': expiring_soon_count,
        'valid_count': valid_count,
        'total_count': total_count,
        'today': today,
    }

    return render(request, 'core/batch/batch_number_list.html', context)


@login_required
@permission_required('products.view_product', raise_exception=True)
def batch_number_detail(request, batch_id):
    """Detailansicht einer Charge."""
    batch = get_object_or_404(BatchNumber.objects.select_related(
        'product', 'warehouse', 'variant', 'supplier'), pk=batch_id)

    context = {
        'batch': batch,
    }

    return render(request, 'core/batch/batch_number_detail.html', context)


@login_required
@permission_required('product', 'create')
def batch_number_add(request):
    """Fügt eine neue Charge hinzu (produktunabhängig)."""

    # Nutze die handle_form_view Utility-Funktion mit einem custom Post-Save-Hook
    def post_save_hook(batch, _):
        # Produkt für Chargenverfolgung aktivieren, falls noch nicht geschehen
        product = batch.product
        if not product.has_batch_tracking:
            product.has_batch_tracking = True
            product.save()
        messages.success(request, 'Charge wurde erfolgreich hinzugefügt.')

    return handle_form_view(
        request=request,
        form_class=BatchNumberForm,
        template='core/batch/batch_number_form.html',
        redirect_url='batch_number_list',
        context_extra={'title': 'Neue Charge anlegen'},
        success_message='Charge wurde erfolgreich erstellt.',
        post_save_hook=post_save_hook
    )


@login_required
@permission_required('product', 'edit')
def batch_number_edit(request, batch_id):
    """Bearbeitet eine bestehende Charge."""
    batch = get_object_or_404(BatchNumber, pk=batch_id)

    # Verwende die handle_model_update Utility-Funktion
    context = handle_model_update(
        request=request,
        instance=batch,
        form_class=BatchNumberForm,
        model_name_singular='Charge',
        success_redirect='batch_number_detail',
        context_extra={
            'title': f'Charge {batch.batch_number} bearbeiten',
            'batch': batch
        }
    )

    if isinstance(context, dict):
        # Wenn die Funktion ein Context-Dict zurückgibt, rendern wir das Template
        return render(request, 'core/batch/batch_number_form.html', context)
    else:
        # Ansonsten ist es eine Redirect-Response
        return context


@login_required
@permission_required('product', 'delete')
def batch_number_delete(request, batch_id):
    """Löscht eine Charge."""
    batch = get_object_or_404(BatchNumber, pk=batch_id)
    product = batch.product  # Produkt speichern, um später zu prüfen

    # Verwende die handle_model_delete Utility-Funktion
    if request.method == 'POST':
        context = handle_model_delete(
            request=request,
            instance=batch,
            model_name_singular='Charge',
            success_redirect='batch_number_list'
        )

        # Wenn keine Chargen mehr übrig sind, has_batch_tracking deaktivieren
        if not product.batches.exists():
            product.has_batch_tracking = False
            product.save()

        return redirect('batch_number_list')

    context = {
        'batch': batch,
    }

    return render(request, 'core/batch/batch_number_confirm_delete.html', context)


@login_required
@permission_required('products.view_product', raise_exception=True)
def batch_number_scan(request):
    """Scannt eine Chargennummer und zeigt Details an."""
    scanned_number = request.GET.get('scan', '')
    found_batch = None

    # Start with an empty list of recent scans
    recent_scans = []

    if scanned_number:
        try:
            # Versuche, die Chargennummer zu finden
            found_batch = BatchNumber.objects.select_related('product', 'warehouse').get(batch_number=scanned_number)

            # Hole die alten Suchanfragen
            recent_searches = request.session.get('recent_batch_searches', [])

            # Aktuelles Datum formatiert (deutsches Format)
            current_time = timezone.now()
            formatted_time = current_time.strftime('%d.%m.%Y %H:%M')

            # Warehouse bestimmen
            warehouse_name = 'Nicht zugewiesen'
            if hasattr(found_batch, 'warehouse') and found_batch.warehouse:
                warehouse_name = found_batch.warehouse.name

            # Create a new search entry
            search_info = {
                'id': found_batch.id,
                'batch_number': found_batch.batch_number,
                'product_name': found_batch.product.name,
                'timestamp': formatted_time,
                'warehouse_name': warehouse_name,
                'quantity': str(found_batch.quantity),
            }

            # Remove duplicates
            filtered_searches = []
            for old_search in recent_searches:
                if old_search.get('id') != found_batch.id:
                    filtered_searches.append(old_search)

            # Add the new search at the beginning
            filtered_searches.insert(0, search_info)

            # Limit to 5 entries
            filtered_searches = filtered_searches[:5]

            # Save to session
            request.session['recent_batch_searches'] = filtered_searches

            # Redirect to detail page
            return redirect('batch_number_detail', batch_id=found_batch.pk)
        except BatchNumber.DoesNotExist:
            messages.error(request, f'Charge {scanned_number} wurde nicht gefunden.')
        except Exception as e:
            messages.error(request, f'Fehler bei der Suche: {str(e)}')

    # Get recent searches from session
    session_searches = request.session.get('recent_batch_searches', [])

    # Format for template
    for item in session_searches:
        recent_scans.append({
            'id': item.get('id', 0),
            'batch_number': item.get('batch_number', ''),
            'product': {'name': item.get('product_name', '')},
            'timestamp': item.get('timestamp', ''),
            'warehouse': {'name': item.get('warehouse_name', 'Nicht zugewiesen')},
            'quantity': item.get('quantity', '0'),
        })

    context = {
        'scanned_number': scanned_number,
        'recent_scans': recent_scans,
    }

    return render(request, 'core/batch/batch_number_scan.html', context)


@login_required
@permission_required('product', 'edit')
def batch_number_transfer(request):
    """Transferiert eine Charge von einem Lager zu einem anderen."""
    if request.method == 'POST':
        batch_number = request.POST.get('batch_number')
        target_warehouse_id = request.POST.get('target_warehouse')
        transfer_quantity = request.POST.get('quantity', None)

        try:
            batch = BatchNumber.objects.get(batch_number=batch_number)
            old_warehouse = batch.warehouse
            new_warehouse = Warehouse.objects.get(pk=target_warehouse_id)

            # Wenn eine Menge angegeben wurde, teile die Charge
            if transfer_quantity:
                transfer_quantity = float(transfer_quantity)

                # Prüfe, ob die zu transferierende Menge gültig ist
                if transfer_quantity <= 0 or transfer_quantity > batch.quantity:
                    messages.error(request, f'Ungültige Menge. Muss größer als 0 und maximal {batch.quantity} sein.')
                    return redirect('batch_number_transfer')

                # Wenn nicht die gesamte Menge transferiert wird, teile die Charge
                if transfer_quantity < batch.quantity:
                    # Erstelle eine neue Charge im Ziellager
                    new_batch = BatchNumber.objects.create(
                        batch_number=batch.batch_number,
                        product=batch.product,
                        variant=batch.variant,
                        quantity=transfer_quantity,
                        production_date=batch.production_date,
                        expiry_date=batch.expiry_date,
                        supplier=batch.supplier,
                        warehouse=new_warehouse,
                        notes=f'Transferiert von {old_warehouse.name if old_warehouse else "unbekannt"} am {timezone.now().strftime("%d.%m.%Y")}'
                    )

                    # Reduziere die Menge der Originalcharge
                    batch.quantity -= transfer_quantity
                    batch.save()

                    messages.success(request,
                                     f'{transfer_quantity} Einheiten der Charge {batch_number} wurden erfolgreich von '
                                     f'{old_warehouse.name if old_warehouse else "unbekanntem Lager"} nach {new_warehouse.name} transferiert.')
                else:
                    # Wenn die gesamte Menge transferiert wird, verschiebe die Charge
                    batch.warehouse = new_warehouse
                    batch.save()

                    messages.success(request, f'Charge {batch_number} wurde vollständig von '
                                              f'{old_warehouse.name if old_warehouse else "unbekanntem Lager"} nach {new_warehouse.name} transferiert.')
            else:
                # Wenn keine Menge angegeben wurde, verschiebe die gesamte Charge
                batch.warehouse = new_warehouse
                batch.save()

                messages.success(request, f'Charge {batch_number} wurde erfolgreich von '
                                          f'{old_warehouse.name if old_warehouse else "unbekanntem Lager"} nach {new_warehouse.name} transferiert.')

            return redirect('batch_number_list')

        except BatchNumber.DoesNotExist:
            messages.error(request, f'Charge {batch_number} wurde nicht gefunden.')
        except Warehouse.DoesNotExist:
            messages.error(request, 'Das Ziellager existiert nicht.')
        except ValueError:
            messages.error(request, 'Ungültige Mengenangabe.')
        except Exception as e:
            messages.error(request, f'Fehler beim Transferieren: {str(e)}')

    # Für GET-Anfragen oder bei Fehlern im POST
    warehouses = Warehouse.objects.filter(is_active=True)

    # Wenn eine Chargennummer als Parameter übergeben wurde
    initial_batch = request.GET.get('batch', '')

    context = {
        'warehouses': warehouses,
        'initial_batch': initial_batch,
    }

    return render(request, 'core/batch/batch_number_transfer.html', context)


@login_required
@permission_required('product', 'import')
def batch_number_import(request):
    """Importiert Chargen aus einer CSV-Datei."""
    # Verwendet die handle_csv_import Utility-Funktion
    return handle_csv_import(
        form_class=BatchNumberImportForm,
        importer_class=BatchNumberImporter,
        request=request,
        template_name='core/batch/batch_number_import.html',
        success_redirect='import_log_detail',  # oder 'batch_number_list'
        extra_context={
            'title': 'Chargen importieren',
            'description': 'Importieren Sie Chargen aus einer CSV-Datei.',
            'expected_format': 'product_sku,batch_number,quantity,warehouse_name,production_date,expiry_date,supplier_name',
            'example': 'P1001,LOT2023001,100,Hauptlager,2023-01-01,2025-01-01,Beispiel GmbH',
            'required_columns': ['product_sku', 'batch_number', 'quantity'],
            'optional_columns': ['warehouse_name', 'production_date', 'expiry_date', 'supplier_name', 'notes'],
        }
    )


@login_required
@permission_required('product', 'export')
def batch_number_export(request):
    """Exportiert Chargen in eine CSV- oder Excel-Datei."""
    # Filteroptionen für den Export
    filters = {
        'warehouse': request.GET.get('warehouse', ''),
        'product': request.GET.get('product', ''),
        'expiry': request.GET.get('expiry', ''),
        'search': request.GET.get('search', ''),
    }

    # Basisdaten
    batches = BatchNumber.objects.select_related('product', 'warehouse', 'variant', 'supplier')

    # Filter anwenden mit den Utility-Funktionen
    batches = apply_exact_filter(batches, 'warehouse_id', filters['warehouse'])
    batches = apply_exact_filter(batches, 'product_id', filters['product'])

    if filters['search']:
        batches = batches.filter(
            Q(batch_number__icontains=filters['search']) |
            Q(product__name__icontains=filters['search']) |
            Q(product__sku__icontains=filters['search'])
        )

    # Filtern nach Verfallsdatum mit apply_expiry_filter
    today = timezone.now().date()
    batches = apply_expiry_filter(batches, filters['expiry'], today, 30)

    # Export-Format bestimmen
    export_format = request.GET.get('format', '')

    if export_format == 'csv':
        # CSV-Export
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="chargen_{timezone.now().strftime("%Y%m%d")}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Chargennummer', 'Produkt', 'SKU', 'Variante', 'Menge', 'Einheit', 'Lager',
            'Lieferant', 'Produktionsdatum', 'Ablaufdatum', 'Notizen', 'Erstellt am'
        ])

        for batch in batches:
            writer.writerow([
                batch.batch_number,
                batch.product.name,
                batch.product.sku,
                batch.variant.name if batch.variant else '',
                str(batch.quantity),
                batch.product.unit,
                batch.warehouse.name if batch.warehouse else '',
                batch.supplier.name if batch.supplier else '',
                batch.production_date.strftime('%Y-%m-%d') if batch.production_date else '',
                batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else '',
                batch.notes,
                batch.created_at.strftime('%Y-%m-%d %H:%M')
            ])

        return response

    elif export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            # Funktion zum Entfernen der Zeitzone
            def remove_timezone(dt):
                if dt and hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
                    return dt.replace(tzinfo=None)
                return dt

            # Neue Arbeitsmappe erstellen
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Chargen"

            # Überschriften formatieren
            headers = [
                'Chargennummer', 'Produkt', 'SKU', 'Variante', 'Menge', 'Einheit', 'Lager',
                'Lieferant', 'Produktionsdatum', 'Ablaufdatum', 'Notizen', 'Erstellt am'
            ]

            # Überschriften-Stil festlegen
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Überschriften schreiben und formatieren
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Daten einfügen
            for row_idx, batch in enumerate(batches, 2):
                ws.cell(row=row_idx, column=1, value=batch.batch_number)
                ws.cell(row=row_idx, column=2, value=batch.product.name)
                ws.cell(row=row_idx, column=3, value=batch.product.sku)
                ws.cell(row=row_idx, column=4, value=batch.variant.name if batch.variant else '')
                ws.cell(row=row_idx, column=5, value=float(batch.quantity))
                ws.cell(row=row_idx, column=6, value=batch.product.unit)
                ws.cell(row=row_idx, column=7, value=batch.warehouse.name if batch.warehouse else '')
                ws.cell(row=row_idx, column=8, value=batch.supplier.name if batch.supplier else '')

                if batch.production_date:
                    ws.cell(row=row_idx, column=9, value=batch.production_date)

                if batch.expiry_date:
                    ws.cell(row=row_idx, column=10, value=batch.expiry_date)
                    # Abgelaufene Chargen rot markieren
                    if batch.expiry_date < today:
                        ws.cell(row=row_idx, column=10).fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB",
                                                                           fill_type="solid")
                    # Bald ablaufende Chargen gelb markieren
                    elif batch.expiry_date <= today + timedelta(days=30):
                        ws.cell(row=row_idx, column=10).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC",
                                                                           fill_type="solid")

                ws.cell(row=row_idx, column=11, value=batch.notes)
                # Zeitzone aus dem created_at entfernen, um Excel-Fehler zu vermeiden
                created_at_value = remove_timezone(batch.created_at) if batch.created_at else None
                ws.cell(row=row_idx, column=12, value=created_at_value)

            # Spaltenbreiten anpassen
            for col_idx, header in enumerate(headers, 1):
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                max_length = len(header) + 2

                # Überprüfe Zellinhalte, um die optimale Spaltenbreite zu bestimmen
                for row_idx in range(2, len(batches) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)) + 2)

                # Spaltenbreite setzen (maximale Breite: 50)
                ws.column_dimensions[column_letter].width = min(max_length, 50)

            # Excel-Datei als Response zurückgeben
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="chargen_{timezone.now().strftime("%Y%m%d")}.xlsx"'

            return response

        except ImportError:
            messages.error(request, 'Excel-Export ist nicht verfügbar. Bitte installieren Sie openpyxl.')
            return redirect('batch_number_list')
        except Exception as e:
            messages.error(request, f'Fehler beim Excel-Export: {str(e)}')
            return redirect('batch_number_list')


    elif export_format == 'pdf':
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                    leftMargin=1 * cm, rightMargin=1 * cm,
                                    topMargin=1 * cm, bottomMargin=1 * cm)
            elements = []

            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            title_style.alignment = 1

            cell_style = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontSize=8.5,
                leading=10,
                wordWrap='CJK',
                alignment=0,
            )

            header_style = ParagraphStyle(
                'HeaderStyle',
                parent=styles['Normal'],
                fontSize=9,
                fontName='Helvetica-Bold',
                alignment=1,
                textColor=colors.white,
            )

            elements.append(Paragraph("Chargenexport", title_style))
            elements.append(Spacer(1, 10))

            # Dynamische Seitenbreite
            PAGE_WIDTH = landscape(A4)[0] - 2 * cm

            col_widths = [
                0.12 * PAGE_WIDTH,  # Chargennummer
                0.15 * PAGE_WIDTH,  # Produkt
                0.08 * PAGE_WIDTH,  # SKU
                0.08 * PAGE_WIDTH,  # Variante
                0.06 * PAGE_WIDTH,  # Menge
                0.07 * PAGE_WIDTH,  # Lager
                0.10 * PAGE_WIDTH,  # Lieferant
                0.08 * PAGE_WIDTH,  # Produktionsdatum
                0.08 * PAGE_WIDTH,  # Ablaufdatum
                0.14 * PAGE_WIDTH,  # Notizen
                0.04 * PAGE_WIDTH,  # Erstellt am
            ]

            headers = [
                Paragraph('<font color="white">Chargennummer</font>', header_style),
                Paragraph('<font color="white">Produkt</font>', header_style),
                Paragraph('<font color="white">SKU</font>', header_style),
                Paragraph('<font color="white">Variante</font>', header_style),
                Paragraph('<font color="white">Menge</font>', header_style),
                Paragraph('<font color="white">Lager</font>', header_style),
                Paragraph('<font color="white">Lieferant</font>', header_style),
                Paragraph('<font color="white">Produktionsdatum</font>', header_style),
                Paragraph('<font color="white">Ablaufdatum</font>', header_style),
                Paragraph('<font color="white">Notizen</font>', header_style),
                Paragraph('<font color="white">Erstellt am</font>', header_style),
            ]

            data = [headers]

            for batch in batches:
                row = [
                    Paragraph(batch.batch_number, cell_style),
                    Paragraph(batch.product.name, cell_style),
                    Paragraph(batch.product.sku, cell_style),
                    Paragraph(batch.variant.name if batch.variant else '', cell_style),
                    Paragraph(f"{batch.quantity} {batch.product.unit}", cell_style),
                    Paragraph(batch.warehouse.name if batch.warehouse else '', cell_style),
                    Paragraph(batch.supplier.name if batch.supplier else '', cell_style),
                    Paragraph(batch.production_date.strftime('%d.%m.%Y') if batch.production_date else '', cell_style),
                    Paragraph(batch.expiry_date.strftime('%d.%m.%Y') if batch.expiry_date else '', cell_style),
                    Paragraph(batch.notes if batch.notes else '', cell_style),
                    Paragraph(batch.created_at.strftime('%d.%m.%Y %H:%M') if batch.created_at else '', cell_style)
                ]
                data.append(row)

            table = Table(data, colWidths=col_widths, repeatRows=1)

            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
                ('TOPPADDING', (0, 0), (-1, 0), 7),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
                ('TOPPADDING', (0, 1), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
                ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
            ])

            for i, batch in enumerate(batches, 1):
                if batch.expiry_date:
                    if batch.expiry_date < today:
                        table_style.add('BACKGROUND', (8, i), (8, i), colors.HexColor('#FFCCCB'))
                    elif batch.expiry_date <= today + timedelta(days=30):
                        table_style.add('BACKGROUND', (8, i), (8, i), colors.HexColor('#FFFFCC'))

                if i % 2 == 0:
                    table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F5F5F5'))

            table.setStyle(table_style)
            elements.append(table)

            doc.build(elements)
            buffer.seek(0)

            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="chargen_{timezone.now().strftime("%Y%m%d")}.pdf"'
            return response

        except ImportError:
            messages.error(request, 'PDF-Export ist nicht verfügbar. Bitte installieren Sie reportlab.')
            return redirect('batch_number_list')

        except Exception as e:
            messages.error(request, f'Fehler beim PDF-Export: {str(e)}')
            return redirect('batch_number_list')

    # Statistiken für die Verfallsdaten
    expired_count = BatchNumber.objects.filter(expiry_date__lt=today).count()
    expiring_soon_count = BatchNumber.objects.filter(
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30)
    ).count()
    valid_count = BatchNumber.objects.filter(
        expiry_date__gt=today + timedelta(days=30)
    ).count()

    # Filter-Werte für das Template bereitstellen
    context = {
        'warehouses': Warehouse.objects.filter(is_active=True),
        'products': Product.objects.filter(has_batch_tracking=True),
        'batches': batches,
        'today': today,
        'date_today': timezone.now().strftime("%Y%m%d"),
        'expired_count': expired_count,
        'expiring_soon_count': expiring_soon_count,
        'valid_count': valid_count,
        'warehouse_filter': filters['warehouse'],
        'product_filter': filters['product'],
        'expiry_filter': filters['expiry'],
        'search_query': filters['search'],
    }

    return render(request, 'core/batch/batch_number_export.html', context)
