import csv
import io
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models.aggregates import Count
from django.db.models.query_utils import Q
from django.http.response import HttpResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.utils import timezone

from accessmanagement.models import WarehouseAccess
from core.utils.filters import filter_product_serials
from core.utils.imports import handle_csv_import
from core.utils.pagination import paginate_queryset
from data_operations.forms.import_forms import SerialNumberImportForm
from data_operations.importers.serialnumbers import SerialNumberImporter
from inventory.models import Warehouse
from product_management.models.products_models import Product
from tracking.forms.tracking_forms import SerialNumberForm, BulkSerialNumberForm
from tracking.models.serial_numbers_models import SerialNumber


@login_required
@permission_required('tracking.view_serialnumber', raise_exception=True)
def serialnumber_list(request):
    """Zeigt alle Seriennummern im System an."""
    # Filteroptionen
    status_filter = request.GET.get('status', '')
    warehouse_filter = request.GET.get('warehouse', '')
    product_filter = request.GET.get('product', '')
    search_query = request.GET.get('search', '')

    filters = {
        'status': request.GET.get('status', ''),
        'warehouse': request.GET.get('warehouse', ''),
        'product': request.GET.get('product', ''),
        'search': request.GET.get('search', ''),
    }

    serials = SerialNumber.objects.select_related('product', 'warehouse', 'variant')
    serials = filter_product_serials(serials, filters)

    # Sortierung
    sort_by = request.GET.get('sort', '-created_at')
    serials = serials.order_by(sort_by)

    # Paginierung
    serials = paginate_queryset(serials, request.GET.get('page'), per_page=50)

    # Status-Statistiken
    status_counts = SerialNumber.objects.values('status').annotate(
        count=Count('status')).order_by('status')

    # In Dictionary umwandeln für leichteren Zugriff
    status_stats = {item['status']: item['count'] for item in status_counts}

    context = {
        'serials': serials,
        'status_stats': status_stats,
        'warehouses': Warehouse.objects.filter(is_active=True),
        'products': Product.objects.filter(has_serial_numbers=True),
        'status_choices': SerialNumber.status_choices,
        'status_filter': status_filter,
        'warehouse_filter': warehouse_filter,
        'product_filter': product_filter,
        'search_query': search_query,
        'sort_by': sort_by,
    }

    return render(request, 'core/serialnumber/serialnumber_list.html', context)


@login_required
@permission_required('tracking.view_serialnumber', raise_exception=True)
def serialnumber_detail(request, serial_id):
    """Zeigt Details zu einer Seriennummer an."""
    serial = get_object_or_404(SerialNumber.objects.select_related(
        'product', 'warehouse', 'variant'), pk=serial_id)

    # Hier könnten später Bewegungshistorie, Wartungsprotokolle etc. geladen werden

    context = {
        'serial': serial,
    }

    return render(request, 'core/serialnumber/serialnumber_detail.html', context)


@login_required
@permission_required('tracking.view_serialnumber', raise_exception=True)
def serialnumber_history(request, serial_id):
    """Zeigt die Historie einer Seriennummer an."""
    serial = get_object_or_404(SerialNumber.objects.select_related(
        'product', 'warehouse', 'variant'), pk=serial_id)

    # Hier würde man die Historiendaten laden
    # Beispiel: movements = SerialNumberMovement.objects.filter(serial_number=serial).order_by('-timestamp')

    context = {
        'serial': serial,
        'movements': [],  # Platzhalter für die tatsächlichen Bewegungsdaten
    }

    return render(request, 'core/serialnumber/serialnumber_history.html', context)


@login_required
@permission_required('tracking.add_serialnumber', raise_exception=True)
def serialnumber_add(request):
    """Fügt eine neue Seriennummer hinzu (produktunabhängig)."""
    if request.method == 'POST':
        form = SerialNumberForm(request.POST)
        if form.is_valid():
            serial = form.save()

            # Produkt für Seriennummern aktivieren, falls noch nicht geschehen
            product = serial.product
            if not product.has_serial_numbers:
                product.has_serial_numbers = True
                product.save()

            messages.success(request, 'Seriennummer wurde erfolgreich hinzugefügt.')
            return redirect('serialnumber_list')
    else:
        form = SerialNumberForm()

    context = {
        'form': form,
    }

    return render(request, 'core/serialnumber/serialnumber_form.html', context)


@login_required
@permission_required('tracking.transfer_serialnumber', raise_exception=True)
def serialnumber_transfer(request):
    """Transferiert eine Seriennummer von einem Lager zu einem anderen."""
    if request.method == 'POST':
        serial_number = request.POST.get('serial_number')
        target_warehouse_id = request.POST.get('target_warehouse')

        try:
            serial = SerialNumber.objects.get(serial_number=serial_number)
            old_warehouse = serial.warehouse
            new_warehouse = Warehouse.objects.get(pk=target_warehouse_id)

            # Prüfe Benutzerberechtigungen für beide Lager
            if not request.user.is_superuser:
                if (old_warehouse and not WarehouseAccess.has_access(request.user, old_warehouse)) or \
                        not WarehouseAccess.has_access(request.user, new_warehouse):
                    messages.error(request, 'Sie haben keine Berechtigung für eines der Lager.')
                    return redirect('serialnumber_transfer')

            # Formatierte Meldung erstellen, je nachdem ob ein altes Lager vorhanden war
            if old_warehouse:
                success_message = f'Seriennummer {serial_number} wurde erfolgreich von ' \
                                  f'{old_warehouse.name} nach {new_warehouse.name} transferiert.'
            else:
                success_message = f'Seriennummer {serial_number} wurde erfolgreich ' \
                                  f'dem Lager {new_warehouse.name} zugewiesen.'

            # Speichere den Transfer
            serial.warehouse = new_warehouse
            serial.last_modified = timezone.now()
            serial.save()

            # Hier könnte man auch einen Eintrag in einer Transfer-Historie speichern

            messages.success(request, success_message)
            return redirect('serialnumber_list')

        except SerialNumber.DoesNotExist:
            messages.error(request, f'Seriennummer {serial_number} wurde nicht gefunden.')
        except Warehouse.DoesNotExist:
            messages.error(request, 'Das Ziellager existiert nicht.')
        except Exception as e:
            messages.error(request, f'Fehler beim Transferieren: {str(e)}')

    # Für GET-Anfragen oder bei Fehlern im POST
    warehouses = Warehouse.objects.filter(is_active=True)

    # Wenn eine Seriennummer als Parameter übergeben wurde (z.B. aus der Detailansicht)
    initial_serial = request.GET.get('serial', '')

    context = {
        'warehouses': warehouses,
        'initial_serial': initial_serial,
    }

    return render(request, 'core/serialnumber/serialnumber_transfer.html', context)


@login_required
@permission_required('tracking.import_serialnumber', raise_exception=True)
def serialnumber_import(request):
    return handle_csv_import(
        form_class=SerialNumberImportForm,
        importer_class=SerialNumberImporter,
        request=request,
        template_name='core/serialnumber/serialnumber_import.html',
        success_redirect='serialnumber_import_log_detail',  # oder z. B. 'serialnumber_list' falls kein DetailView
        extra_context={
            'title': 'Seriennummern importieren',
            'description': 'Importieren Sie Seriennummern aus einer CSV-Datei.',
            'expected_format': 'product_sku,serial_number,status,warehouse_name,purchase_date,expiry_date',
            'example': 'P1001,SN12345,in_stock,Hauptlager,2023-01-01,2025-01-01',
            'required_columns': ['product_sku', 'serial_number'],
            'optional_columns': ['status', 'warehouse_name', 'purchase_date', 'expiry_date', 'notes'],
        }
    )


@login_required
@permission_required('tracking.export_serialnumber', raise_exception=True)
def serialnumber_export(request):
    """Exportiert Seriennummern in eine CSV- oder Excel-Datei."""
    # Filteroptionen für den Export
    filters = {
        'status': request.GET.get('status', ''),
        'warehouse': request.GET.get('warehouse', ''),
        'product': request.GET.get('product', ''),
        'search': request.GET.get('search', ''),
    }

    # Basisdaten
    serials = SerialNumber.objects.select_related('product', 'warehouse', 'variant')
    serials = filter_product_serials(serials, filters)

    # Export-Format bestimmen
    export_format = request.GET.get('format', '')

    if export_format == 'csv':
        # CSV-Export
        response = HttpResponse(content_type='text/csv')
        response[
            'Content-Disposition'] = f'attachment; filename="seriennummern_{timezone.now().strftime("%Y%m%d")}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Seriennummer', 'Produkt', 'SKU', 'Variante', 'Status', 'Lager',
            'Kaufdatum', 'Ablaufdatum', 'Notizen', 'Erstellt am'
        ])

        for serial in serials:
            writer.writerow([
                serial.serial_number,
                serial.product.name,
                serial.product.sku,
                serial.variant.name if serial.variant else '',
                serial.get_status_display(),
                serial.warehouse.name if serial.warehouse else '',
                serial.purchase_date.strftime('%Y-%m-%d') if serial.purchase_date else '',
                serial.expiry_date.strftime('%Y-%m-%d') if serial.expiry_date else '',
                serial.notes,
                serial.created_at.strftime('%Y-%m-%d %H:%M')
            ])

        return response

    elif export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            # Funktion zum Entfernen der Zeitzone
            def remove_timezone(dt):
                if dt and hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
                    return dt.replace(tzinfo=None)
                return dt

            # Neue Arbeitsmappe erstellen
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Seriennummern"

            # Überschriften formatieren
            headers = [
                'Seriennummer', 'Produkt', 'SKU', 'Variante', 'Status', 'Lager',
                'Kaufdatum', 'Ablaufdatum', 'Notizen', 'Erstellt am'
            ]

            # Überschriften-Stil festlegen
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Überschriften schreiben und formatieren
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Heute-Datum für Ablaufprüfung
            today = timezone.now().date()

            # Daten einfügen
            for row_idx, serial in enumerate(serials, 2):
                ws.cell(row=row_idx, column=1, value=serial.serial_number)
                ws.cell(row=row_idx, column=2, value=serial.product.name)
                ws.cell(row=row_idx, column=3, value=serial.product.sku)
                ws.cell(row=row_idx, column=4, value=serial.variant.name if serial.variant else '')
                ws.cell(row=row_idx, column=5, value=serial.get_status_display())
                ws.cell(row=row_idx, column=6, value=serial.warehouse.name if serial.warehouse else '')

                if serial.purchase_date:
                    ws.cell(row=row_idx, column=7, value=serial.purchase_date)

                if serial.expiry_date:
                    ws.cell(row=row_idx, column=8, value=serial.expiry_date)
                    # Abgelaufene Seriennummern rot markieren
                    if serial.expiry_date < today:
                        ws.cell(row=row_idx, column=8).fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB",
                                                                          fill_type="solid")
                    # Bald ablaufende Seriennummern gelb markieren
                    elif serial.expiry_date <= today + timedelta(days=30):
                        ws.cell(row=row_idx, column=8).fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC",
                                                                          fill_type="solid")

                ws.cell(row=row_idx, column=9, value=serial.notes)
                # Zeitzone aus dem created_at entfernen, um Excel-Fehler zu vermeiden
                created_at_value = remove_timezone(serial.created_at) if serial.created_at else None
                ws.cell(row=row_idx, column=10, value=created_at_value)

            # Spaltenbreiten anpassen
            for col_idx, header in enumerate(headers, 1):
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                max_length = len(header) + 2

                # Überprüfe Zellinhalte, um die optimale Spaltenbreite zu bestimmen
                for row_idx in range(2, len(serials) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)) + 2)

                # Spaltenbreite setzen (maximale Breite: 50)
                ws.column_dimensions[column_letter].width = min(max_length, 50)

            # Excel-Datei als Response zurückgeben
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response[
                'Content-Disposition'] = f'attachment; filename="seriennummern_{timezone.now().strftime("%Y%m%d")}.xlsx"'

            return response

        except ImportError:
            messages.error(request, 'Excel-Export ist nicht verfügbar. Bitte installieren Sie openpyxl.')
            return redirect('serialnumber_list')
        except Exception as e:
            messages.error(request, f'Fehler beim Excel-Export: {str(e)}')
            return redirect('serialnumber_list')

    elif export_format == 'pdf':
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm, cm
            from io import BytesIO

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                    leftMargin=1 * cm, rightMargin=1 * cm,
                                    topMargin=1 * cm, bottomMargin=1 * cm)
            elements = []

            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            title_style.alignment = 1

            cell_style = ParagraphStyle(
                'CellStyle',
                parent=styles['Normal'],
                fontSize=8.5,
                leading=10,
                wordWrap='CJK',
                alignment=0,
            )

            header_style = ParagraphStyle(
                'HeaderStyle',
                parent=styles['Normal'],
                fontSize=9,
                fontName='Helvetica-Bold',
                alignment=1,
                textColor=colors.white,
            )

            elements.append(Paragraph("Seriennummernexport", title_style))
            elements.append(Spacer(1, 10))

            # Dynamische Seitenbreite
            PAGE_WIDTH = landscape(A4)[0] - 2 * cm

            col_widths = [
                0.14 * PAGE_WIDTH,  # Seriennummer
                0.14 * PAGE_WIDTH,  # Produkt
                0.09 * PAGE_WIDTH,  # SKU
                0.09 * PAGE_WIDTH,  # Variante
                0.08 * PAGE_WIDTH,  # Status
                0.08 * PAGE_WIDTH,  # Lager
                0.08 * PAGE_WIDTH,  # Kaufdatum
                0.08 * PAGE_WIDTH,  # Ablaufdatum
                0.16 * PAGE_WIDTH,  # Notizen
                0.06 * PAGE_WIDTH,  # Erstellt am
            ]

            headers = [
                Paragraph('<font color="white">Seriennummer</font>', header_style),
                Paragraph('<font color="white">Produkt</font>', header_style),
                Paragraph('<font color="white">SKU</font>', header_style),
                Paragraph('<font color="white">Variante</font>', header_style),
                Paragraph('<font color="white">Status</font>', header_style),
                Paragraph('<font color="white">Lager</font>', header_style),
                Paragraph('<font color="white">Kaufdatum</font>', header_style),
                Paragraph('<font color="white">Ablaufdatum</font>', header_style),
                Paragraph('<font color="white">Notizen</font>', header_style),
                Paragraph('<font color="white">Erstellt am</font>', header_style),
            ]

            data = [headers]

            # Heute-Datum für Ablaufprüfung
            today = timezone.now().date()

            for serial in serials:
                row = [
                    Paragraph(serial.serial_number, cell_style),
                    Paragraph(serial.product.name, cell_style),
                    Paragraph(serial.product.sku, cell_style),
                    Paragraph(serial.variant.name if serial.variant else '', cell_style),
                    Paragraph(serial.get_status_display(), cell_style),
                    Paragraph(serial.warehouse.name if serial.warehouse else '', cell_style),
                    Paragraph(serial.purchase_date.strftime('%d.%m.%Y') if serial.purchase_date else '', cell_style),
                    Paragraph(serial.expiry_date.strftime('%d.%m.%Y') if serial.expiry_date else '', cell_style),
                    Paragraph(serial.notes if serial.notes else '', cell_style),
                    Paragraph(serial.created_at.strftime('%d.%m.%Y %H:%M') if serial.created_at else '', cell_style)
                ]
                data.append(row)

            table = Table(data, colWidths=col_widths, repeatRows=1)

            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
                ('TOPPADDING', (0, 0), (-1, 0), 7),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
                ('TOPPADDING', (0, 1), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1, colors.black),
                ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
                ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
            ])

            for i, serial in enumerate(serials, 1):
                if serial.expiry_date:
                    if serial.expiry_date < today:
                        table_style.add('BACKGROUND', (7, i), (7, i), colors.HexColor('#FFCCCB'))
                    elif serial.expiry_date <= today + timedelta(days=30):
                        table_style.add('BACKGROUND', (7, i), (7, i), colors.HexColor('#FFFFCC'))

                if i % 2 == 0:
                    table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F5F5F5'))

            table.setStyle(table_style)
            elements.append(table)

            doc.build(elements)
            buffer.seek(0)

            response = HttpResponse(buffer, content_type='application/pdf')
            response[
                'Content-Disposition'] = f'attachment; filename="seriennummern_{timezone.now().strftime("%Y%m%d")}.pdf"'
            return response

        except ImportError:
            messages.error(request, 'PDF-Export ist nicht verfügbar. Bitte installieren Sie reportlab.')
            return redirect('serialnumber_list')

        except Exception as e:
            messages.error(request, f'Fehler beim PDF-Export: {str(e)}')
            return redirect('serialnumber_list')

    # Für GET-Anfragen oder bei Fehlern im POST
    # Statistiken nach Status
    status_counts = {}
    for status_code, status_name in SerialNumber.status_choices:
        status_counts[status_code] = SerialNumber.objects.filter(status=status_code).count()

    # Statistiken für Verfallsdaten
    today = timezone.now().date()
    expired_count = SerialNumber.objects.filter(expiry_date__lt=today).count()
    expiring_soon_count = SerialNumber.objects.filter(
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30)
    ).count()
    valid_count = SerialNumber.objects.filter(
        expiry_date__gt=today + timedelta(days=30)
    ).count()

    context = {
        'warehouses': Warehouse.objects.filter(is_active=True),
        'products': Product.objects.filter(has_serial_numbers=True),
        'status_choices': SerialNumber.status_choices,
        'status_counts': status_counts,
        'serials': serials,
        'today': today,
        'date_today': timezone.now().strftime("%Y%m%d"),
        'expired_count': expired_count,
        'expiring_soon_count': expiring_soon_count,
        'valid_count': valid_count,
        'warehouse_filter': filters['warehouse'],
        'product_filter': filters['product'],
        'status_filter': filters['status'],
        'search_query': filters['search'],
    }

    return render(request, 'core/serialnumber/serialnumber_export.html', context)


@login_required
@permission_required('tracking.view_serialnumber', raise_exception=True)
def serialnumber_scan(request):
    """Scannt eine Seriennummer (z.B. mit Barcode-Scanner) und zeigt Details an."""
    scanned_number = request.GET.get('scan', '')
    found_serial = None

    # Start with an empty list of recent scans
    recent_scans = []

    if scanned_number:
        try:
            # Versuche, die Seriennummer zu finden mit explizitem select_related für warehouse
            found_serial = SerialNumber.objects.select_related('product', 'warehouse').get(serial_number=scanned_number)

            # Hole die alten Suchanfragen
            recent_searches = request.session.get('recent_serial_searches', [])

            # Aktuelles Datum formatiert (deutsches Format)
            current_time = timezone.now()
            formatted_time = current_time.strftime('%d.%m.%Y %H:%M')

            # Warehouse bestimmen
            warehouse_name = 'Nicht zugewiesen'
            if hasattr(found_serial, 'warehouse') and found_serial.warehouse:
                warehouse_name = found_serial.warehouse.name

            # Create a new search entry
            search_info = {
                'id': found_serial.id,
                'serial_number': found_serial.serial_number,
                'product_name': found_serial.product.name,
                'timestamp': formatted_time,
                'warehouse_name': warehouse_name
            }

            # Remove duplicates
            filtered_searches = []
            for old_search in recent_searches:
                if old_search.get('id') != found_serial.id:
                    filtered_searches.append(old_search)

            # Add the new search at the beginning
            filtered_searches.insert(0, search_info)

            # Limit to 5 entries
            filtered_searches = filtered_searches[:5]

            # Save to session
            request.session['recent_serial_searches'] = filtered_searches

            # Redirect to detail page
            return redirect('serialnumber_detail', serial_id=found_serial.pk)
        except SerialNumber.DoesNotExist:
            messages.error(request, f'Seriennummer {scanned_number} wurde nicht gefunden.')
        except Exception as e:
            messages.error(request, f'Fehler bei der Suche: {str(e)}')

    # Get recent searches from session
    session_searches = request.session.get('recent_serial_searches', [])

    # Format for template
    for item in session_searches:
        recent_scans.append({
            'id': item.get('id', 0),
            'serial_number': item.get('serial_number', ''),
            'product': {'name': item.get('product_name', '')},
            'timestamp': item.get('timestamp', ''),
            'warehouse': {'name': item.get('warehouse_name', 'Nicht zugewiesen')}
        })

    context = {
        'scanned_number': scanned_number,
        'recent_scans': recent_scans,
    }

    return render(request, 'core/serialnumber/serialnumber_scan.html', context)


@login_required
@permission_required('tracking.view_serialnumber', raise_exception=True)
def serialnumber_search(request):
    """Erweiterte Suchfunktion für Seriennummern."""
    search_query = request.GET.get('q', '')

    results = []
    if search_query:
        # Suche nach Seriennummer, Produktname, verschiedenen Feldern
        results = SerialNumber.objects.filter(
            Q(serial_number__icontains=search_query) |
            Q(product__name__icontains=search_query) |
            Q(product__sku__icontains=search_query) |
            Q(product__barcode__icontains=search_query) |
            Q(variant__name__icontains=search_query) |
            Q(notes__icontains=search_query)
        ).select_related('product', 'warehouse', 'variant')

    context = {
        'search_query': search_query,
        'results': results,
    }

    return render(request, 'core/serialnumber/serialnumber_search.html', context)


@login_required
@permission_required('tracking.change_serialnumber', raise_exception=True)
def serialnumber_batch_actions(request):
    """Massenaktionen für mehrere Seriennummern gleichzeitig."""
    if request.method == 'POST':
        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_serials')

        if not selected_ids:
            messages.error(request, 'Keine Seriennummern ausgewählt.')
            return redirect('serialnumber_list')

        # Seriennummern abrufen
        serials = SerialNumber.objects.filter(pk__in=selected_ids)

        if action == 'change_status':
            new_status = request.POST.get('new_status')
            if new_status:
                serials.update(status=new_status)
                messages.success(request,
                                 f'{len(selected_ids)} Seriennummern auf Status "{dict(SerialNumber.status_choices)[new_status]}" gesetzt.')

        elif action == 'change_warehouse':
            new_warehouse_id = request.POST.get('new_warehouse')
            if new_warehouse_id:
                try:
                    warehouse = Warehouse.objects.get(pk=new_warehouse_id)
                    serials.update(warehouse=warehouse)
                    messages.success(request, f'{len(selected_ids)} Seriennummern nach "{warehouse.name}" verschoben.')
                except Warehouse.DoesNotExist:
                    messages.error(request, 'Das ausgewählte Lager existiert nicht.')

        elif action == 'delete':
            count = serials.count()
            serials.delete()
            messages.success(request, f'{count} Seriennummern wurden gelöscht.')

        return redirect('serialnumber_list')

    # Standardwerte für GET-Anfrage
    context = {
        'status_choices': SerialNumber.status_choices,
        'warehouses': Warehouse.objects.filter(is_active=True),
    }

    return render(request, 'core/serialnumber/serialnumber_batch_actions.html', context)


@login_required
@permission_required('product_management.view_product', raise_exception=True)
def product_serials(request, pk):
    """Zeigt alle Seriennummern eines Produkts an."""
    product = get_object_or_404(Product, pk=pk)

    filters = {
        'status': request.GET.get('status', ''),
        'warehouse': request.GET.get('warehouse', ''),
        'variant': request.GET.get('variant', ''),
        'search': request.GET.get('search', ''),
    }

    serials = SerialNumber.objects.filter(product=product)
    serials = filter_product_serials(serials, filters)
    serials = paginate_queryset(serials, request.GET.get('page'), per_page=50)

    status_counts = SerialNumber.objects.filter(product=product).values('status').annotate(
        count=Count('status')).order_by('status')
    status_stats = {item['status']: item['count'] for item in status_counts}

    context = {
        'product': product,
        'serials': serials,
        'status_stats': status_stats,
        'warehouses': Warehouse.objects.filter(is_active=True),
        'variants': product.variants.all(),
        'status_choices': SerialNumber.status_choices,
        **filters,
    }

    return render(request, 'core/product/product_serials.html', context)


@login_required
@permission_required('product_management.add_product', raise_exception=True)
def product_serial_add(request, pk):
    """Fügt eine Seriennummer zu einem Produkt hinzu."""
    product = get_object_or_404(Product, pk=pk)

    # Prüfen, ob das Produkt für Seriennummern konfiguriert ist
    if not product.has_serial_numbers:
        product.has_serial_numbers = True
        product.save()
        messages.info(request, 'Das Produkt wurde für Seriennummern aktiviert.')

    if request.method == 'POST':
        form = SerialNumberForm(request.POST, product=product)
        if form.is_valid():
            serial = form.save(commit=False)
            serial.product = product
            serial.save()

            messages.success(request, 'Seriennummer wurde erfolgreich hinzugefügt.')
            return redirect('product_serials', pk=product.pk)
    else:
        form = SerialNumberForm(product=product)

    context = {
        'product': product,
        'form': form,
    }

    return render(request, 'core/product/product_serial_form.html', context)


@login_required
@permission_required('product_management.add_product', raise_exception=True)
def product_serial_bulk_add(request, pk):
    """Fügt mehrere Seriennummern zu einem Produkt hinzu."""
    product = get_object_or_404(Product, pk=pk)

    # Prüfen, ob das Produkt für Seriennummern konfiguriert ist
    if not product.has_serial_numbers:
        product.has_serial_numbers = True
        product.save()
        messages.info(request, 'Das Produkt wurde für Seriennummern aktiviert.')

    if request.method == 'POST':
        form = BulkSerialNumberForm(request.POST, product=product)
        if form.is_valid():
            prefix = form.cleaned_data.get('prefix', '')
            start_number = form.cleaned_data.get('start_number')
            count = form.cleaned_data.get('count')
            digits = form.cleaned_data.get('digits')
            status = form.cleaned_data.get('status')
            warehouse = form.cleaned_data.get('warehouse')
            purchase_date = form.cleaned_data.get('purchase_date')
            expiry_date = form.cleaned_data.get('expiry_date')
            variant = form.cleaned_data.get('variant')

            # Format für Seriennummern erstellen
            serial_format = f"{prefix}{{:0{digits}d}}"

            # Prüfen, ob Seriennummern bereits existieren
            existing_numbers = []
            for i in range(start_number, start_number + count):
                serial_number = serial_format.format(i)
                if SerialNumber.objects.filter(serial_number=serial_number).exists():
                    existing_numbers.append(serial_number)

            if existing_numbers:
                messages.error(request,
                               f'Folgende Seriennummern existieren bereits: {", ".join(existing_numbers[:10])}' +
                               (f' und {len(existing_numbers) - 10} weitere' if len(existing_numbers) > 10 else ''))
                context = {
                    'product': product,
                    'form': form,
                }
                return render(request, 'core/product/product_serial_bulk_form.html', context)

            # Seriennummern erstellen
            created_count = 0
            for i in range(start_number, start_number + count):
                serial_number = serial_format.format(i)
                SerialNumber.objects.create(
                    product=product,
                    serial_number=serial_number,
                    status=status,
                    warehouse=warehouse,
                    purchase_date=purchase_date,
                    expiry_date=expiry_date,
                    variant=variant
                )
                created_count += 1

            messages.success(request, f'{created_count} Seriennummern wurden erfolgreich erstellt.')
            return redirect('product_serials', pk=product.pk)
    else:
        form = BulkSerialNumberForm(product=product)

    context = {
        'product': product,
        'form': form,
    }

    return render(request, 'core/product/product_serial_bulk_form.html', context)


@login_required
@permission_required('product_management.change_product', raise_exception=True)
def product_serial_update(request, pk, serial_id):
    """Aktualisiert eine Seriennummer."""
    product = get_object_or_404(Product, pk=pk)
    serial = get_object_or_404(SerialNumber, pk=serial_id, product=product)

    if request.method == 'POST':
        form = SerialNumberForm(request.POST, instance=serial, product=product)
        if form.is_valid():
            serial = form.save()
            messages.success(request, 'Seriennummer wurde erfolgreich aktualisiert.')
            return redirect('product_serials', pk=product.pk)
    else:
        form = SerialNumberForm(instance=serial, product=product)

    context = {
        'product': product,
        'serial': serial,
        'form': form,
    }

    return render(request, 'core/product/product_serial_form.html', context)


@login_required
@permission_required('product_management.delete_product', raise_exception=True)
def product_serial_delete(request, pk, serial_id):
    """Löscht eine Seriennummer."""
    product = get_object_or_404(Product, pk=pk)
    serial = get_object_or_404(SerialNumber, pk=serial_id, product=product)

    if request.method == 'POST':
        serial.delete()
        messages.success(request, 'Seriennummer wurde erfolgreich gelöscht.')
        return redirect('product_serials', pk=product.pk)

    context = {
        'product': product,
        'serial': serial,
    }

    return render(request, 'core/product/product_serial_confirm_delete.html', context)
